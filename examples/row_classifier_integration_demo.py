#!/usr/bin/env python3
"""
Row Classifier Integration Demo
Shows how RowClassifier works with ColumnMapper and ExcelProcessor
"""

import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.file_processor import ExcelProcessor
from core.column_mapper import ColumnMapper
from core.row_classifier import RowClassifier
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')


def create_complex_boq_file():
    """Create a complex BOQ Excel file for testing"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Create complex Excel structure
    wb = Workbook()
    ws = wb.active
    ws.title = "Complex BOQ"
    
    # Add project info
    ws['A1'] = "PROJECT: Sample Construction Project"
    ws['A2'] = "BOQ REFERENCE: BOQ-2024-001"
    ws['A3'] = "DATE: 2024-01-15"
    
    # Add section headers
    ws['A5'] = "SECTION 1: EARTHWORKS"
    ws['A6'] = ""
    
    # Add line items
    ws['A7'] = "1.1"
    ws['B7'] = "Excavation for foundation"
    ws['C7'] = "m³"
    ws['D7'] = "100"
    ws['E7'] = "25.50"
    ws['F7'] = "2550.00"
    
    ws['A8'] = "1.2"
    ws['B8'] = "Backfilling"
    ws['C8'] = "m³"
    ws['D8'] = "50"
    ws['E8'] = "15.00"
    ws['F8'] = "750.00"
    
    # Add subtotal
    ws['A9'] = ""
    ws['B9'] = "Subtotal - Earthworks"
    ws['F9'] = "3300.00"
    
    ws['A10'] = ""
    
    # Add section 2
    ws['A11'] = "SECTION 2: CONCRETE WORKS"
    ws['A12'] = ""
    
    ws['A13'] = "2.1"
    ws['B13'] = "Concrete foundation"
    ws['C13'] = "m³"
    ws['D13'] = "50"
    ws['E13'] = "150.00"
    ws['F13'] = "7500.00"
    
    ws['A14'] = "2.2"
    ws['B14'] = "Concrete columns"
    ws['C14'] = "m³"
    ws['D14'] = "30"
    ws['E14'] = "180.00"
    ws['F14'] = "5400.00"
    
    # Add notes
    ws['A15'] = ""
    ws['B15'] = "Note: All concrete works include formwork and reinforcement"
    
    # Add subtotal
    ws['A16'] = ""
    ws['B16'] = "Subtotal - Concrete"
    ws['F16'] = "12900.00"
    
    ws['A17'] = ""
    
    # Add section 3
    ws['A18'] = "SECTION 3: STEEL & FORMWORK"
    ws['A19'] = ""
    
    ws['A20'] = "3.1"
    ws['B20'] = "Reinforcement steel"
    ws['C20'] = "kg"
    ws['D20'] = "2000"
    ws['E20'] = "2.50"
    ws['F20'] = "5000.00"
    
    ws['A21'] = "3.2"
    ws['B21'] = "Formwork"
    ws['C21'] = "m²"
    ws['D21'] = "200"
    ws['E21'] = "15.00"
    ws['F21'] = "3000.00"
    
    # Add subtotal
    ws['A22'] = ""
    ws['B22'] = "Subtotal - Steel & Formwork"
    ws['F22'] = "8000.00"
    
    ws['A23'] = ""
    
    # Add summary
    ws['A24'] = ""
    ws['B24'] = "TOTAL CONSTRUCTION"
    ws['F24'] = "24200.00"
    
    ws['A25'] = ""
    ws['B25'] = "Contingency (10%)"
    ws['F25'] = "2420.00"
    
    ws['A26'] = ""
    ws['B26'] = "GRAND TOTAL"
    ws['F26'] = "26620.00"
    
    # Add some invalid items
    ws['A27'] = ""
    ws['A28'] = "4.1"
    ws['B28'] = "Invalid item - missing unit"
    ws['C28'] = ""
    ws['D28'] = "100"
    ws['E28'] = "25.50"
    ws['F28'] = "2550.00"
    
    ws['A29'] = "4.2"
    ws['B29'] = "Invalid item - negative price"
    ws['C29'] = "m³"
    ws['D29'] = "50"
    ws['E29'] = "-150.00"
    ws['F29'] = "-7500.00"
    
    # Format headers
    for row in [1, 2, 3, 5, 11, 18]:
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save file
    test_file = Path("complex_boq_test.xlsx")
    wb.save(test_file)
    
    print(f"Created complex BOQ test file: {test_file}")
    return test_file


def demo_full_integration():
    """Demonstrate full integration of all components"""
    print("Row Classifier Integration Demo")
    print("=" * 60)
    
    # Create test file
    test_file = create_complex_boq_file()
    
    try:
        # Process Excel file
        processor = ExcelProcessor()
        if processor.load_file(test_file):
            print(f"\nProcessing file: {test_file}")
            print(f"File info: {processor.get_file_info()}")
            
            # Get sheet data
            sheets = processor.get_all_sheets_data()
            
            if not sheets:
                print("No sheets found in file")
                return
            
            # Process each sheet
            column_mapper = ColumnMapper()
            row_classifier = RowClassifier()
            
            for sheet_name, sheet_data in sheets.items():
                print(f"\n--- Processing Sheet: {sheet_name} ---")
                print(f"Sheet dimensions: {len(sheet_data)} rows x {len(sheet_data[0]) if sheet_data else 0} columns")
                
                # Step 1: Find header row and map columns
                header_info = column_mapper.find_header_row(sheet_data)
                print(f"Header row: {header_info.row_index} (confidence: {header_info.confidence:.2f})")
                
                if header_info and header_info.headers:
                    # Map columns
                    column_mapping_result = column_mapper.process_sheet_mapping(sheet_data)
                    column_mapping = {m.column_index: m.mapped_type for m in column_mapping_result.mappings}
                    
                    print(f"Column mapping confidence: {column_mapping_result.overall_confidence:.2f}")
                    print(f"Mapped columns: {len(column_mapping)}")
                    
                    # Step 2: Classify rows
                    row_classification_result = row_classifier.classify_rows(sheet_data, column_mapping)
                    
                    print(f"\nRow Classification Results:")
                    print(f"Overall quality score: {row_classification_result.overall_quality_score:.2f}")
                    
                    # Show summary
                    print("\nRow Type Summary:")
                    for row_type, count in row_classification_result.summary.items():
                        if count > 0:
                            print(f"  {row_type.value}: {count}")
                    
                    # Show detailed classifications for first 10 rows
                    print("\nDetailed Classifications (first 10 rows):")
                    print("-" * 80)
                    
                    for classification in row_classification_result.classifications[:10]:
                        row_data = sheet_data[classification.row_index]
                        print(f"Row {classification.row_index + 1}: {classification.row_type.value}")
                        print(f"  Confidence: {classification.confidence:.2f}")
                        print(f"  Completeness: {classification.completeness_score:.2f}")
                        
                        if classification.hierarchical_level is not None:
                            print(f"  Hierarchical Level: {classification.hierarchical_level}")
                        
                        if classification.section_title:
                            print(f"  Section: '{classification.section_title}'")
                        
                        if classification.reasoning:
                            print(f"  Reasoning: {'; '.join(classification.reasoning)}")
                        
                        if classification.validation_errors:
                            print(f"  Validation Errors: {'; '.join(classification.validation_errors)}")
                        
                        print(f"  Data: {row_data}")
                        print()
                    
                    # Show suggestions
                    if row_classification_result.suggestions:
                        print("Suggestions:")
                        for suggestion in row_classification_result.suggestions:
                            print(f"  - {suggestion}")
                    
                    # Show sample data with classifications
                    print("\nSample Data with Classifications:")
                    print("-" * 60)
                    
                    for classification in row_classification_result.classifications:
                        if classification.row_type.value in ['primary_line_item', 'subtotal_row', 'header_section_break']:
                            row_data = sheet_data[classification.row_index]
                            print(f"Row {classification.row_index + 1} [{classification.row_type.value}]: {row_data}")
                
                else:
                    print("No headers found in sheet")
        
        processor.close()
    
    except Exception as e:
        print(f"Error processing file: {e}")
    
    finally:
        # Clean up test file
        if test_file.exists():
            test_file.unlink()
            print(f"\nCleaned up test file: {test_file}")


def demo_validation_workflow():
    """Demonstrate validation workflow"""
    print("\n" + "=" * 60)
    print("Validation Workflow Demo")
    print("=" * 60)
    
    # Create sample data with validation issues
    sample_data = [
        ["Item Code", "Description", "Unit", "Quantity", "Unit Price", "Total"],
        ["1.1", "Valid item", "m³", "100", "25.50", "2550.00"],
        ["1.2", "Missing unit", "", "100", "25.50", "2550.00"],
        ["1.3", "Negative price", "m³", "50", "-150.00", "-7500.00"],
        ["1.4", "Zero quantity", "m³", "0", "150.00", "0.00"],
        ["1.5", "Invalid quantity", "m³", "abc", "25.50", "2550.00"],
        ["", "Subtotal", "", "", "", "3300.00"],
        ["", "Note: Prices include 10% contingency", "", "", "", ""],
    ]
    
    # Create column mapping
    column_mapping = {
        0: "code",
        1: "description",
        2: "unit",
        3: "quantity",
        4: "unit_price",
        5: "total_price"
    }
    
    print("Sample data with validation issues:")
    for i, row in enumerate(sample_data):
        print(f"  Row {i + 1}: {row}")
    
    # Process with both classifiers
    column_mapper = ColumnMapper()
    row_classifier = RowClassifier()
    
    # Map columns
    column_mapping_result = column_mapper.process_sheet_mapping(sample_data)
    enum_mapping = {m.column_index: m.mapped_type for m in column_mapping_result.mappings}
    
    print(f"\nColumn mapping confidence: {column_mapping_result.overall_confidence:.2f}")
    
    # Classify rows
    row_result = row_classifier.classify_rows(sample_data, enum_mapping)
    
    print(f"\nRow classification quality: {row_result.overall_quality_score:.2f}")
    
    # Show validation results
    print("\nValidation Results:")
    print("-" * 50)
    
    for classification in row_result.classifications:
        if classification.validation_errors:
            print(f"Row {classification.row_index + 1} [{classification.row_type.value}]:")
            for error in classification.validation_errors:
                print(f"  ✗ {error}")
        else:
            print(f"Row {classification.row_index + 1} [{classification.row_type.value}]: ✓ Valid")


def demo_performance_analysis():
    """Demonstrate performance analysis capabilities"""
    print("\n" + "=" * 60)
    print("Performance Analysis Demo")
    print("=" * 60)
    
    # Create sample data with various row types
    sample_data = [
        ["Code", "Description", "Unit", "Qty", "Rate", "Total"],
        ["1.1", "Excavation", "m³", "100", "25.50", "2550.00"],
        ["1.2", "Concrete", "m³", "50", "150.00", "7500.00"],
        ["", "Subtotal", "", "", "", "10050.00"],
        ["", "", "", "", "", ""],
        ["2.1", "Steel", "kg", "2000", "2.50", "5000.00"],
        ["2.2", "Formwork", "m²", "200", "15.00", "3000.00"],
        ["", "Subtotal", "", "", "", "8000.00"],
        ["", "Note: All prices include VAT", "", "", "", ""],
        ["", "GRAND TOTAL", "", "", "", "18050.00"],
    ]
    
    column_mapper = ColumnMapper()
    row_classifier = RowClassifier()
    
    # Process data
    column_result = column_mapper.process_sheet_mapping(sample_data)
    enum_mapping = {m.column_index: m.mapped_type for m in column_result.mappings}
    row_result = row_classifier.classify_rows(sample_data, enum_mapping)
    
    # Analyze performance
    print("Performance Analysis:")
    print("-" * 40)
    
    # Data completeness
    line_items = [c for c in row_result.classifications if c.row_type.value == 'primary_line_item']
    avg_completeness = sum(c.completeness_score for c in line_items) / len(line_items) if line_items else 0
    print(f"Average line item completeness: {avg_completeness:.2f}")
    
    # Data quality
    invalid_items = [c for c in row_result.classifications if c.row_type.value == 'invalid_line_item']
    total_items = len(line_items) + len(invalid_items)
    quality_score = len(line_items) / total_items if total_items > 0 else 0
    print(f"Data quality score: {quality_score:.2f} ({len(line_items)}/{total_items} valid)")
    
    # Structure analysis
    subtotals = [c for c in row_result.classifications if c.row_type.value == 'subtotal_row']
    headers = [c for c in row_result.classifications if c.row_type.value == 'header_section_break']
    notes = [c for c in row_result.classifications if c.row_type.value == 'notes_comments']
    
    print(f"Structure analysis:")
    print(f"  Line items: {len(line_items)}")
    print(f"  Subtotals: {len(subtotals)}")
    print(f"  Headers: {len(headers)}")
    print(f"  Notes: {len(notes)}")
    
    # Confidence analysis
    avg_confidence = sum(c.confidence for c in row_result.classifications) / len(row_result.classifications)
    print(f"Average classification confidence: {avg_confidence:.2f}")


def main():
    """Run all integration demos"""
    demo_full_integration()
    demo_validation_workflow()
    demo_performance_analysis()
    
    print("\n" + "=" * 60)
    print("Integration demos completed successfully!")


if __name__ == "__main__":
    main() 