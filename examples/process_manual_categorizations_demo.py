#!/usr/bin/env python3
"""
Demo script for processing manual categorization Excel files
"""

import sys
import os
from pathlib import Path
import pandas as pd

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.auto_categorizer import CategoryDictionary, auto_categorize_dataset, collect_unmatched_descriptions
from core.manual_categorizer import (
    generate_manual_categorization_excel, 
    process_manual_categorizations,
    validate_excel_file_structure,
    get_categorization_statistics
)


def main():
    """Demo the manual categorization processing workflow"""
    print("=== Manual Categorization Processing Demo ===\n")
    
    # Load category dictionary
    print("1. Loading category dictionary...")
    category_dict = CategoryDictionary(Path("config/category_dictionary.json"))
    print(f"   Loaded {len(category_dict.mappings)} category mappings")
    print(f"   Available categories: {list(category_dict.get_all_categories())}")
    print()
    
    # Load sample data
    print("2. Loading sample BOQ data...")
    sample_file = "examples/sample_boq.csv"
    if not os.path.exists(sample_file):
        print(f"   Sample file not found: {sample_file}")
        print("   Creating sample data...")
        
        # Create sample data
        sample_data = {
            'Description': [
                'Solar Panel Installation',
                'Inverter Setup',
                'Cable Management',
                'Mounting System',
                'Electrical Testing',
                'Unknown Component A',
                'Mystery Part B',
                'Uncategorized Item C',
                'Solar Panel Installation',  # Duplicate for frequency testing
                'Inverter Setup'  # Duplicate for frequency testing
            ],
            'Quantity': [10, 5, 100, 20, 1, 15, 8, 12, 5, 3],
            'Unit': ['pcs', 'pcs', 'm', 'sets', 'lot', 'pcs', 'pcs', 'pcs', 'pcs', 'pcs'],
            'Unit_Price': [150.00, 500.00, 2.50, 75.00, 1000.00, 25.00, 45.00, 30.00, 150.00, 500.00]
        }
        df = pd.DataFrame(sample_data)
        df.to_csv(sample_file, index=False)
        print(f"   Created sample file: {sample_file}")
    else:
        df = pd.read_csv(sample_file)
        print(f"   Loaded {len(df)} rows from {sample_file}")
    print()
    
    # Auto-categorize the data
    print("3. Auto-categorizing data...")
    result = auto_categorize_dataset(df, category_dict)
    categorized_df = result.dataframe
    unmatched_descriptions = result.unmatched_descriptions
    stats = result.match_statistics
    print(f"   Categorized {stats['matched_rows']} out of {stats['total_rows']} descriptions")
    print(f"   Unmatched descriptions: {len(unmatched_descriptions)}")
    print()
    
    # Collect unmatched descriptions
    print("4. Collecting unmatched descriptions...")
    unmatched_list = collect_unmatched_descriptions(categorized_df, 'Description')
    print(f"   Collected {len(unmatched_list)} unique unmatched descriptions")
    print()
    
    # Generate manual categorization Excel file
    print("5. Generating manual categorization Excel file...")
    available_categories = list(category_dict.get_all_categories())
    
    try:
        excel_filepath = generate_manual_categorization_excel(
            unmatched_descriptions=unmatched_list,
            available_categories=available_categories,
            output_dir=Path("examples")
        )
        print(f"   Excel file created: {excel_filepath}")
        print()
        
        # Validate the Excel file structure
        print("6. Validating Excel file structure...")
        validation_result = validate_excel_file_structure(excel_filepath)
        
        if validation_result['is_valid']:
            print("   ✓ Excel file structure is valid")
            print(f"   File size: {validation_result['file_info']['file_size']} bytes")
            print(f"   Last modified: {validation_result['file_info']['last_modified']}")
            print(f"   Data rows: {validation_result['sheet_info'].get('data_rows', 0)}")
        else:
            print("   ✗ Excel file structure has issues:")
            for error in validation_result['errors']:
                print(f"     - {error}")
        
        if validation_result['warnings']:
            print("   Warnings:")
            for warning in validation_result['warnings']:
                print(f"     - {warning}")
        print()
        
        # Try to process manual categorizations (will be empty if not filled yet)
        print("7. Processing manual categorizations...")
        try:
            categorization_mapping = process_manual_categorizations(excel_filepath)
            
            if categorization_mapping:
                print(f"   ✓ Found {len(categorization_mapping)} manual categorizations")
                
                # Get statistics
                stats = get_categorization_statistics(categorization_mapping)
                print(f"   Total categorizations: {stats['total_categorizations']}")
                print(f"   Unique categories used: {stats['unique_categories']}")
                print(f"   Average description length: {stats['average_description_length']} chars")
                print()
                
                print("   Category distribution:")
                for category, count in stats['category_distribution'].items():
                    print(f"     {category}: {count}")
                print()
                
                # Show some examples
                print("   Sample categorizations:")
                for i, (desc, cat) in enumerate(list(categorization_mapping.items())[:5]):
                    print(f"     '{desc}' → '{cat}'")
                print()
                
            else:
                print("   No manual categorizations found (file not yet filled)")
                print("   This is expected if the file hasn't been manually edited yet")
                print()
                
        except Exception as e:
            print(f"   Error processing manual categorizations: {e}")
            print("   This is expected if the file hasn't been manually edited yet")
            print()
        
        print("=== Demo completed ===")
        print("\nNext steps:")
        print("1. Open the generated Excel file:")
        print(f"   {excel_filepath}")
        print("2. Go to the 'Categorization' sheet")
        print("3. Select categories from the dropdown for each description")
        print("4. Add any notes if needed")
        print("5. Save the file")
        print("6. Run this demo again to process the manual categorizations")
        
    except Exception as e:
        print(f"   Error generating Excel file: {e}")
        return


def test_with_sample_categorizations():
    """Test the processing with sample categorizations"""
    print("\n=== Testing with Sample Categorizations ===\n")
    
    # Create a sample Excel file with some categorizations filled in
    sample_data = {
        'Description': [
            'Unknown Component A',
            'Mystery Part B', 
            'Uncategorized Item C'
        ],
        'Source_Sheet': [
            'Sheet1',
            'Sheet1',
            'Sheet2'
        ],
        'Frequency': [15, 8, 12],
        'Category': [
            'Electrical Components',
            'Mechanical Parts',
            'Tools & Equipment'
        ],
        'Notes': [
            'Looks like a connector',
            'Metal bracket type',
            'Hand tool category'
        ]
    }
    
    # Create a temporary Excel file
    temp_file = Path("examples/sample_categorized.xlsx")
    
    try:
        # Create workbook
        import openpyxl
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)
        
        # Create Categorization sheet
        ws = wb.create_sheet("Categorization")
        
        # Add headers
        headers = ["Description", "Source_Sheet", "Frequency", "Category", "Notes"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, (desc, sheet, freq, cat, note) in enumerate(zip(
            sample_data['Description'],
            sample_data['Source_Sheet'],
            sample_data['Frequency'],
            sample_data['Category'],
            sample_data['Notes']
        ), 2):
            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=sheet)
            ws.cell(row=row, column=3, value=freq)
            ws.cell(row=row, column=4, value=cat)
            ws.cell(row=row, column=5, value=note)
        
        # Save file
        wb.save(temp_file)
        print(f"Created sample categorized file: {temp_file}")
        
        # Process the sample file
        print("\nProcessing sample categorizations...")
        categorization_mapping = process_manual_categorizations(temp_file)
        
        print(f"Processed {len(categorization_mapping)} categorizations:")
        for desc, cat in categorization_mapping.items():
            print(f"  '{desc}' → '{cat}'")
        
        # Get statistics
        stats = get_categorization_statistics(categorization_mapping)
        print(f"\nStatistics:")
        print(f"  Total: {stats['total_categorizations']}")
        print(f"  Unique categories: {stats['unique_categories']}")
        print(f"  Category distribution: {stats['category_distribution']}")
        
        # Clean up
        temp_file.unlink()
        print(f"\nCleaned up temporary file")
        
    except Exception as e:
        print(f"Error in sample test: {e}")


if __name__ == "__main__":
    main()
    
    # Run additional test with sample data
    test_with_sample_categorizations() 