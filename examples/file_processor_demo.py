#!/usr/bin/env python3
"""
Excel File Processor Demo
Demonstrates the capabilities of the ExcelProcessor class
"""

import sys
import tempfile
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.file_processor import ExcelProcessor, analyze_excel_file
from openpyxl import Workbook
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')


def create_sample_excel_file() -> Path:
    """Create a sample Excel file for testing"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create BOQ Main sheet
    ws1 = wb.create_sheet("BOQ Main")
    ws1['A1'] = "Description"
    ws1['B1'] = "Qty"
    ws1['C1'] = "Unit"
    ws1['D1'] = "Unit Price"
    ws1['E1'] = "Total"
    
    # Add sample data
    data = [
        ["Excavation for foundation", 100, "m³", 25.50, 2550.00],
        ["Concrete foundation", 50, "m³", 150.00, 7500.00],
        ["Reinforcement steel", 2000, "kg", 2.50, 5000.00],
        ["Formwork", 200, "m²", 15.00, 3000.00],
        ["", "", "", "", ""],  # Empty row
        ["Subtotal", "", "", "", 18050.00],
        ["", "", "", "", ""],  # Empty row
        ["Contingency (10%)", "", "", "", 1805.00],
        ["Grand Total", "", "", "", 19855.00]
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Category"
    ws2['B1'] = "Amount"
    
    summary_data = [
        ["Substructure", 18050.00],
        ["Contingency", 1805.00],
        ["Total", 19855.00]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Create hidden sheet
    ws3 = wb.create_sheet("Hidden Sheet")
    ws3.sheet_state = 'hidden'
    ws3['A1'] = "This sheet is hidden"
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    wb.close()
    
    return Path(temp_file.name)


def demo_basic_functionality():
    """Demonstrate basic ExcelProcessor functionality"""
    print("=== Basic ExcelProcessor Functionality ===")
    
    # Create sample file
    sample_file = create_sample_excel_file()
    print(f"Created sample file: {sample_file}")
    
    try:
        # Use context manager for automatic cleanup
        with ExcelProcessor(max_memory_mb=100) as processor:
            # Load file
            success = processor.load_file(sample_file)
            print(f"File loaded successfully: {success}")
            
            # Get visible sheets
            visible_sheets = processor.get_visible_sheets()
            print(f"Visible sheets: {visible_sheets}")
            
            # Get metadata for each sheet
            for sheet_name in visible_sheets:
                metadata = processor.get_sheet_metadata(sheet_name)
                print(f"\nSheet: {metadata.name}")
                print(f"  Rows: {metadata.row_count}, Columns: {metadata.column_count}")
                print(f"  Data density: {metadata.data_density:.1%}")
                print(f"  Data range: {metadata.first_data_row}-{metadata.last_data_row}, "
                      f"{metadata.first_data_column}-{metadata.last_data_column}")
                print(f"  Empty rows: {metadata.empty_rows_count}")
                print(f"  Empty columns: {metadata.empty_columns_count}")
                print(f"  Estimated size: {metadata.estimated_size_mb:.2f} MB")
            
            # Sample content from first sheet
            if visible_sheets:
                sample = processor.sample_sheet_content(visible_sheets[0], rows=5)
                print(f"\nContent sample from '{sample.sheet_name}':")
                print(f"  Headers: {sample.headers}")
                print(f"  Sample rows: {sample.sample_size}")
                print(f"  Has data: {sample.has_data}")
                
                if sample.rows:
                    print("  First few rows:")
                    for i, row in enumerate(sample.rows[:3]):
                        print(f"    Row {i+1}: {row}")
    
    finally:
        # Clean up sample file
        sample_file.unlink()
        print(f"\nCleaned up sample file")


def demo_quick_analysis():
    """Demonstrate quick file analysis function"""
    print("\n=== Quick File Analysis ===")
    
    # Create sample file
    sample_file = create_sample_excel_file()
    
    try:
        # Analyze file
        analysis = analyze_excel_file(sample_file, sample_rows=10)
        
        print(f"File: {analysis['file_path']}")
        print(f"Format: {analysis['file_format']}")
        print(f"Total sheets: {analysis['total_sheets']}")
        print(f"Visible sheets: {analysis['visible_sheets']}")
        
        print("\nSheets metadata:")
        for name, meta in analysis['sheets_metadata'].items():
            print(f"  {name}: {meta['row_count']} rows, {meta['column_count']} columns, "
                  f"density: {meta['data_density']:.1%}")
        
        if analysis['first_sheet_sample']:
            sample = analysis['first_sheet_sample']
            print(f"\nFirst sheet sample:")
            print(f"  Headers: {sample['headers']}")
            print(f"  Sample rows: {len(sample['sample_rows'])}")
            print(f"  Has data: {sample['has_data']}")
    
    finally:
        # Clean up sample file
        sample_file.unlink()


def demo_error_handling():
    """Demonstrate error handling"""
    print("\n=== Error Handling Demo ===")
    
    # Test with non-existent file
    try:
        with ExcelProcessor() as processor:
            processor.load_file("non_existent_file.xlsx")
    except FileNotFoundError as e:
        print(f"✓ Correctly caught FileNotFoundError: {e}")
    
    # Test with invalid file
    temp_file = tempfile.NamedTemporaryFile(suffix='.txt', delete=False)
    temp_file.write(b"This is not an Excel file")
    temp_file.close()
    
    try:
        with ExcelProcessor() as processor:
            processor.load_file(temp_file.name)
    except Exception as e:
        print(f"✓ Correctly caught invalid file error: {type(e).__name__}")
    
    # Clean up
    Path(temp_file.name).unlink()


def demo_memory_management():
    """Demonstrate memory management"""
    print("\n=== Memory Management Demo ===")
    
    # Create a large sample file
    wb = Workbook()
    ws = wb.active
    ws.title = "Large Sheet"
    
    # Add headers
    for col in range(1, 11):
        ws.cell(row=1, column=col, value=f"Column {col}")
    
    # Add data (1000 rows)
    for row in range(2, 1002):
        for col in range(1, 11):
            ws.cell(row=row, column=col, value=f"Data {row}-{col}")
    
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    wb.close()
    
    try:
        # Test with memory limit
        with ExcelProcessor(max_memory_mb=1) as processor:  # Very low limit
            try:
                processor.load_file(temp_file.name)
                print("✓ File loaded within memory limit")
                
                metadata = processor.get_sheet_metadata("Large Sheet")
                print(f"  Rows: {metadata.row_count}, Columns: {metadata.column_count}")
                print(f"  Data density: {metadata.data_density:.1%}")
                
            except MemoryError as e:
                print(f"✓ Memory limit enforced: {e}")
    
    finally:
        Path(temp_file.name).unlink()


def main():
    """Run all demos"""
    print("Excel File Processor Demo")
    print("=" * 50)
    
    demo_basic_functionality()
    demo_quick_analysis()
    demo_error_handling()
    demo_memory_management()
    
    print("\n" + "=" * 50)
    print("All demos completed successfully!")


if __name__ == "__main__":
    main() 