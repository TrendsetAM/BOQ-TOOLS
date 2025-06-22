#!/usr/bin/env python3
"""
BOQ Processor Demo
Demonstrates the complete BOQ processing pipeline
"""

import sys
import tempfile
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.boq_processor import BOQProcessor
from openpyxl import Workbook
import logging
import json

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')


def create_sample_boq_file() -> Path:
    """Create a sample BOQ Excel file for testing"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create BOQ Main sheet
    ws1 = wb.create_sheet("BOQ Main")
    
    # Headers
    headers = ["Item Code", "Description", "Unit", "Quantity", "Unit Price", "Total Amount", "Remarks"]
    for col, header in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=header)
    
    # Sample BOQ data
    data = [
        ["001", "Excavation for foundation", "m³", 100, 25.50, 2550.00, "Manual excavation"],
        ["002", "Concrete foundation", "m³", 50, 150.00, 7500.00, "C25 concrete"],
        ["003", "Reinforcement steel", "kg", 2000, 2.50, 5000.00, "High yield steel"],
        ["004", "Formwork", "m²", 200, 15.00, 3000.00, "Plywood formwork"],
        ["005", "Backfilling", "m³", 80, 12.00, 960.00, "Selected fill"],
        ["", "", "", "", "", "", ""],  # Empty row
        ["", "Subtotal", "", "", "", 18010.00, ""],
        ["", "", "", "", "", "", ""],  # Empty row
        ["", "Contingency (10%)", "", "", "", 1801.00, ""],
        ["", "Grand Total", "", "", "", 19811.00, ""]
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Category"
    ws2['B1'] = "Description"
    ws2['C1'] = "Amount"
    
    summary_data = [
        ["Substructure", "Foundation works", 18010.00],
        ["Contingency", "10% contingency", 1801.00],
        ["", "Total", 19811.00]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Preliminaries sheet
    ws3 = wb.create_sheet("Preliminaries")
    ws3['A1'] = "Item"
    ws3['B1'] = "Description"
    ws3['C1'] = "Unit"
    ws3['D1'] = "Quantity"
    ws3['E1'] = "Rate"
    ws3['F1'] = "Amount"
    
    prelim_data = [
        ["Site setup", "Temporary site office", "lump sum", 1, 5000.00, 5000.00],
        ["Site security", "Security fencing", "m", 100, 25.00, 2500.00],
        ["Utilities", "Temporary power and water", "lump sum", 1, 3000.00, 3000.00]
    ]
    
    for row_idx, row_data in enumerate(prelim_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws3.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    wb.close()
    
    return Path(temp_file.name)


def demo_basic_processing():
    """Demonstrate basic BOQ processing"""
    print("=== Basic BOQ Processing ===")
    
    # Create sample file
    sample_file = create_sample_boq_file()
    print(f"Created sample BOQ file: {sample_file}")
    
    try:
        # Use context manager for automatic cleanup
        with BOQProcessor() as processor:
            # Load file
            success = processor.load_excel(sample_file)
            print(f"File loaded successfully: {success}")
            
            if success:
                # Process the file
                results = processor.process()
                
                print(f"\nProcessing Results:")
                print(f"  File: {results['file_path']}")
                print(f"  Format: {results['file_format']}")
                print(f"  Sheets processed: {results['sheets_processed']}")
                
                # Show summary
                summary = results['summary']
                print(f"\nSummary:")
                print(f"  Total sheets: {summary['total_sheets']}")
                print(f"  Valid sheets: {summary['valid_sheets']}")
                print(f"  Total items: {summary['total_items']}")
                print(f"  Total value: ${summary['total_value']:,.2f}")
                
                # Show sheet types
                print(f"\nSheet types:")
                for sheet_type, count in summary['sheet_types'].items():
                    print(f"  {sheet_type}: {count}")
                
                # Show column coverage
                print(f"\nColumn coverage:")
                for col_type, count in summary['column_coverage'].items():
                    print(f"  {col_type}: {count} sheets")
                
                # Show detailed results for each sheet
                print(f"\nDetailed Results:")
                for sheet_name, sheet_data in results['sheets_data'].items():
                    print(f"\n  Sheet: {sheet_name}")
                    print(f"    Type: {sheet_data['sheet_type']} (confidence: {sheet_data['confidence']:.2f})")
                    print(f"    Valid: {sheet_data['validation']['is_valid']}")
                    print(f"    Score: {sheet_data['validation']['score']:.2f}")
                    
                    if sheet_data['validation']['errors']:
                        print(f"    Errors: {sheet_data['validation']['errors']}")
                    
                    if sheet_data['validation']['warnings']:
                        print(f"    Warnings: {sheet_data['validation']['warnings']}")
                    
                    print(f"    Columns mapped: {len(sheet_data['column_mappings'])}")
                    for col_idx, col_type in sheet_data['column_mappings'].items():
                        print(f"      Column {col_idx}: {col_type}")
                    
                    if sheet_data['boq_data']:
                        print(f"    BOQ items: {len(sheet_data['boq_data'])}")
                        print(f"    Sample items:")
                        for i, item in enumerate(sheet_data['boq_data'][:3]):
                            print(f"      {i+1}. {item['description']} - {item['quantity']} {item['unit']} @ ${item['unit_price']}")
    
    finally:
        # Clean up sample file
        sample_file.unlink()
        print(f"\nCleaned up sample file")


def demo_error_handling():
    """Demonstrate error handling in BOQ processing"""
    print("\n=== Error Handling Demo ===")
    
    # Test with non-existent file
    try:
        with BOQProcessor() as processor:
            processor.load_excel(Path("non_existent_file.xlsx"))
    except Exception as e:
        print(f"✓ Correctly handled non-existent file: {type(e).__name__}")
    
    # Test with invalid file
    temp_file = tempfile.NamedTemporaryFile(suffix='.txt', delete=False)
    temp_file.write(b"This is not an Excel file")
    temp_file.close()
    
    try:
        with BOQProcessor() as processor:
            processor.load_excel(Path(temp_file.name))
    except Exception as e:
        print(f"✓ Correctly handled invalid file: {type(e).__name__}")
    
    # Clean up
    Path(temp_file.name).unlink()


def demo_configuration_integration():
    """Demonstrate configuration system integration"""
    print("\n=== Configuration Integration Demo ===")
    
    from utils.config import get_config, ColumnType
    
    config = get_config()
    
    print("Configuration system integration:")
    print(f"  Column types: {len(config.get_all_column_types())}")
    print(f"  Required columns: {[col.value for col in config.get_required_columns()]}")
    print(f"  Sheet classifications: {len(config.sheet_classifications)}")
    
    # Test column mapping
    test_headers = ["Description", "Qty", "Unit Price", "Total", "Remarks"]
    print(f"\nTesting column mapping with headers: {test_headers}")
    
    for header in test_headers:
        best_match = None
        best_score = 0
        
        for col_type in config.get_all_column_types():
            mapping = config.get_column_mapping(col_type)
            if mapping:
                for keyword in mapping.keywords:
                    if keyword.lower() in header.lower():
                        score = mapping.weight
                        if score > best_score:
                            best_score = score
                            best_match = col_type.value
        
        if best_match:
            print(f"  '{header}' -> {best_match} (score: {best_score})")
        else:
            print(f"  '{header}' -> no match")


def demo_export_results():
    """Demonstrate exporting processing results"""
    print("\n=== Export Results Demo ===")
    
    # Create sample file
    sample_file = create_sample_boq_file()
    
    try:
        with BOQProcessor() as processor:
            if processor.load_excel(sample_file):
                results = processor.process()
                
                # Export to JSON
                export_file = sample_file.with_suffix('.json')
                with open(export_file, 'w') as f:
                    json.dump(results, f, indent=2, default=str)
                
                print(f"Results exported to: {export_file}")
                print(f"Export file size: {export_file.stat().st_size} bytes")
                
                # Show export structure
                print(f"\nExport structure:")
                print(f"  file_path: {results['file_path']}")
                print(f"  file_format: {results['file_format']}")
                print(f"  sheets_processed: {results['sheets_processed']}")
                print(f"  sheets_data: {len(results['sheets_data'])} sheets")
                print(f"  summary: {len(results['summary'])} summary items")
                
                # Clean up export file
                export_file.unlink()
    
    finally:
        sample_file.unlink()


def main():
    """Run all demos"""
    print("BOQ Processor Demo")
    print("=" * 60)
    
    demo_basic_processing()
    demo_error_handling()
    demo_configuration_integration()
    demo_export_results()
    
    print("\n" + "=" * 60)
    print("All demos completed successfully!")


if __name__ == "__main__":
    main() 