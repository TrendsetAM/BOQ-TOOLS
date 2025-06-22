#!/usr/bin/env python3
"""
Column Mapper Integration Demo
Shows how ColumnMapper works with ExcelProcessor
"""

import sys
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.file_processor import ExcelProcessor
from core.column_mapper import ColumnMapper
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')


def create_test_excel_file():
    """Create a test Excel file for demonstration"""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    # Create test data
    data = {
        'Item Code': ['001', '002', '003', '004'],
        'Description': [
            'Excavation for foundation',
            'Concrete foundation',
            'Reinforcement steel',
            'Formwork'
        ],
        'Unit': ['m³', 'm³', 'kg', 'm²'],
        'Quantity': [100, 50, 2000, 200],
        'Unit Price': [25.50, 150.00, 2.50, 15.00],
        'Total Amount': [2550.00, 7500.00, 5000.00, 3000.00]
    }
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Sheet"
    
    # Add headers with formatting
    headers = list(data.keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(zip(*data.values()), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save file
    test_file = Path("test_boq.xlsx")
    wb.save(test_file)
    
    print(f"Created test file: {test_file}")
    return test_file


def demo_integration():
    """Demonstrate ColumnMapper integration with ExcelProcessor"""
    print("Column Mapper Integration Demo")
    print("=" * 50)
    
    # Create test file
    test_file = create_test_excel_file()
    
    try:
        # Process Excel file
        processor = ExcelProcessor()
        if processor.load_file(test_file):
            print(f"\nProcessing file: {test_file}")
            print(f"File info: {processor.get_file_info()}")
            
            # Get sheet data
            sheets = processor.get_all_sheets_data()
            
            if not sheets:
                print("No sheets found in file")
                return
            
            # Process each sheet
            mapper = ColumnMapper()
            
            for sheet_name, sheet_data in sheets.items():
                print(f"\n--- Processing Sheet: {sheet_name} ---")
                print(f"Sheet dimensions: {len(sheet_data)} rows x {len(sheet_data[0]) if sheet_data else 0} columns")
                
                # Find header row
                header_info = mapper.find_header_row(sheet_data)
                print(f"Header row: {header_info.row_index} (confidence: {header_info.confidence:.2f})")
                
                if header_info.headers:
                    print("Headers found:")
                    for i, header in enumerate(header_info.headers):
                        print(f"  Column {i + 1}: '{header}'")
                    
                    # Map columns
                    result = mapper.process_sheet_mapping(sheet_data)
                    
                    print(f"\nMapping Results:")
                    print(f"Overall confidence: {result.overall_confidence:.2f}")
                    print(f"Mapped columns: {len(result.mappings)}")
                    print(f"Unmapped columns: {len(result.unmapped_columns)}")
                    
                    if result.mappings:
                        print("\nColumn Mappings:")
                        for mapping in result.mappings:
                            print(f"  Column {mapping.column_index + 1}: '{mapping.original_header}' -> {mapping.mapped_type.value}")
                            print(f"    Confidence: {mapping.confidence:.2f}")
                            print(f"    Normalized: '{mapping.normalized_header}'")
                    
                    if result.unmapped_columns:
                        print("\nUnmapped Columns:")
                        for col_idx in result.unmapped_columns:
                            if col_idx < len(header_info.headers):
                                header = header_info.headers[col_idx]
                                print(f"  Column {col_idx + 1}: '{header}'")
                    
                    if result.suggestions:
                        print("\nSuggestions:")
                        for suggestion in result.suggestions:
                            print(f"  - {suggestion}")
                    
                    # Show sample data with mapped columns
                    if len(sheet_data) > header_info.row_index + 1:
                        print(f"\nSample Data (first 2 rows after header):")
                        for row_idx in range(header_info.row_index + 1, min(header_info.row_index + 3, len(sheet_data))):
                            row = sheet_data[row_idx]
                            print(f"  Row {row_idx + 1}: {row}")
                else:
                    print("No headers found in sheet")
        
        processor.close()
    
    except Exception as e:
        print(f"Error processing file: {e}")
    
    finally:
        # Clean up test file
        if test_file.exists():
            test_file.unlink()
            print(f"\nCleaned up test file: {test_file}")


def demo_with_real_excel_structure():
    """Demo with more complex Excel structure"""
    print("\n" + "=" * 50)
    print("Complex Excel Structure Demo")
    print("=" * 50)
    
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Create complex Excel structure
    wb = Workbook()
    ws = wb.active
    ws.title = "Complex BOQ"
    
    # Add project info
    ws['A1'] = "PROJECT: Sample Construction Project"
    ws['A2'] = "BOQ REFERENCE: BOQ-2024-001"
    ws['A3'] = "DATE: 2024-01-15"
    
    # Add merged header
    ws['A5'] = "Item Details"
    ws['B5'] = "Measurement"
    ws['C5'] = "Pricing"
    ws.merge_cells('A5:A6')
    ws.merge_cells('B5:B6')
    ws.merge_cells('C5:F6')
    
    # Add detailed headers
    ws['A7'] = "Item Code"
    ws['B7'] = "Description"
    ws['C7'] = "Unit"
    ws['D7'] = "Quantity"
    ws['E7'] = "Unit Rate"
    ws['F7'] = "Total Amount"
    
    # Add data
    data = [
        ['001', 'Excavation for foundation', 'm³', 100, 25.50, 2550.00],
        ['002', 'Concrete foundation', 'm³', 50, 150.00, 7500.00],
        ['003', 'Reinforcement steel', 'kg', 2000, 2.50, 5000.00],
        ['004', 'Formwork', 'm²', 200, 15.00, 3000.00],
    ]
    
    for row_idx, row_data in enumerate(data, 8):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add summary
    ws['A13'] = "Subtotal:"
    ws['F13'] = 18050.00
    ws['A14'] = "Contingency (10%):"
    ws['F14'] = 1805.00
    ws['A15'] = "Total:"
    ws['F15'] = 19855.00
    
    # Format headers
    for row in [5, 7]:
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save file
    complex_file = Path("complex_boq.xlsx")
    wb.save(complex_file)
    
    print(f"Created complex test file: {complex_file}")
    
    try:
        # Process complex file
        processor = ExcelProcessor()
        if processor.load_file(complex_file):
            print(f"\nProcessing complex file: {complex_file}")
            
            sheets = processor.get_all_sheets_data()
            
            if sheets:
                sheet_name = list(sheets.keys())[0]
                sheet_data = sheets[sheet_name]
                
                print(f"\n--- Complex Sheet: {sheet_name} ---")
                print(f"Sheet dimensions: {len(sheet_data)} rows x {len(sheet_data[0]) if sheet_data else 0} columns")
                
                # Show first 10 rows
                print("\nFirst 10 rows:")
                for i, row in enumerate(sheet_data[:10]):
                    print(f"  Row {i + 1}: {row}")
                
                # Process with ColumnMapper
                mapper = ColumnMapper()
                result = mapper.process_sheet_mapping(sheet_data)
                
                print(f"\nComplex Mapping Results:")
                print(f"Header row: {result.header_row.row_index}")
                print(f"Header confidence: {result.header_row.confidence:.2f}")
                print(f"Overall confidence: {result.overall_confidence:.2f}")
                print(f"Is merged: {result.header_row.is_merged}")
                
                if result.mappings:
                    print("\nMappings:")
                    for mapping in result.mappings:
                        print(f"  Column {mapping.column_index + 1}: '{mapping.original_header}' -> {mapping.mapped_type.value}")
                        print(f"    Confidence: {mapping.confidence:.2f}")
                        if mapping.alternatives:
                            print(f"    Alternatives: {[(alt[0].value, alt[1]) for alt in mapping.alternatives[:2]]}")
        
        processor.close()
    
    except Exception as e:
        print(f"Error processing complex file: {e}")
    
    finally:
        # Clean up
        if complex_file.exists():
            complex_file.unlink()
            print(f"\nCleaned up complex test file: {complex_file}")


def main():
    """Run integration demos"""
    demo_integration()
    demo_with_real_excel_structure()
    
    print("\n" + "=" * 50)
    print("Integration demos completed successfully!")


if __name__ == "__main__":
    main() 