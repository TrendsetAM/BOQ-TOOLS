from openpyxl import Workbook
from pathlib import Path

# Create a new workbook
wb = Workbook()

# Sheet 1: Line Items
ws1 = wb.active
ws1.title = "Sheet1"
sheet1_data = [
    ["Item No.", "Description", "Unit", "Quantity", "Unit Price", "Total Price"],
    ["1.1", "Excavation for foundation", "m³", 150.00, 25.00, 3750.00],
    ["1.2", "Concrete foundation", "m³", 75.00, 120.00, 9000.00],
    ["", "", "", "", "Subtotal", 12750.00],
    ["2.1", "Brickwork", "m²", 200.00, 45.00, 9000.00],
    ["", "", "", "", "Total", 21750.00],
]
for row in sheet1_data:
    ws1.append(row)

# Sheet 2: Project Info + Table
ws2 = wb.create_sheet("Sheet2")
sheet2_data = [
    ["Project Information", "", "", "", "", ""],
    ["Project Name:", "Sample Building Project", "", "", "", ""],
    ["Client:", "ABC Construction Ltd", "", "", "", ""],
    ["Date:", "2024-01-15", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["Item", "Specification", "Qty", "Rate", "Amount", "Remarks"],
    [1, "Site preparation", 1, 5000.00, 5000.00, ""],
    [2, "Foundation work", 1, 15000.00, 15000.00, ""],
    ["", "", "", "", "Total", 20000.00],
]
for row in sheet2_data:
    ws2.append(row)

# Sheet 3: Notes
ws3 = wb.create_sheet("Sheet3")
sheet3_data = [
    ["Notes and Conditions", "", "", "", "", ""],
    ["1. All prices are inclusive of taxes", "", "", "", "", ""],
    ["2. Payment terms: 30 days", "", "", "", "", ""],
    ["3. Delivery: As per schedule", "", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["Contact Information", "", "", "", "", ""],
    ["Phone:", "+1-234-567-8900", "", "", "", ""],
    ["Email:", "info@sample.com", "", "", "", ""],
]
for row in sheet3_data:
    ws3.append(row)

# Save the file
output_path = Path(__file__).parent / "sample_boq.xlsx"
wb.save(output_path)
print(f"Demo Excel file created at: {output_path}") 