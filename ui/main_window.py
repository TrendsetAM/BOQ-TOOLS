"""
BOQ Tools Main Window
Comprehensive GUI for Excel file processing and BOQ analysis
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import platform
from pathlib import Path
from typing import List, Dict, Any, Tuple
import logging
import threading
import dataclasses
import pandas as pd
import openpyxl
from core.row_classifier import RowType
from datetime import datetime
import pickle
import re
import time

# Get a logger for this module
logger = logging.getLogger(__name__)

# Custom exceptions for validation failures
class ValidationError(Exception):
    """Custom exception for validation failures"""
    def __init__(self, message, validation_result=None):
        super().__init__(message)
        self.validation_result = validation_result

class PositionValidationError(ValidationError):
    """Specific exception for position-description validation failures"""
    pass

def _format_validation_error_message(validation_result, context=""):
    """
    Standardized error message formatting for position-description validation failures
    
    Args:
        validation_result: Validation result dictionary with errors and summary
        context: Additional context string (e.g., "Use Mapping", "Compare Full")
        
    Returns:
        str: Formatted error message for user display
    """
    if not validation_result or validation_result.get('is_valid', True):
        return "No validation errors found."
    
    # Base error message
    title = f"{context} - Structure Validation Failed" if context else "Structure Validation Failed"
    
    # Summary
    summary = validation_result.get('summary', 'Validation failed with unknown errors.')
    
    # Error details (limit to first 5 errors for readability)
    errors = validation_result.get('errors', [])
    error_details = "\n".join(errors[:5])
    if len(errors) > 5:
        error_details += f"\n... and {len(errors) - 5} more errors"
    
    # Mismatched positions for additional context
    mismatched_positions = validation_result.get('mismatched_positions', [])
    position_count = len(mismatched_positions)
    
    # Build comprehensive error message
    if context == "Use Mapping":
        action_guidance = (
            "This means the current file has a different structure than the saved mapping.\n"
            "You cannot apply this mapping to the current file.\n\n"
            "Possible solutions:\n"
            "• Use a file with the same structure as the original\n"
            "• Create a new mapping for this file structure\n"
            "• Verify you selected the correct mapping file"
        )
    elif context == "Compare Full":
        action_guidance = (
            "This means the comparison file has a different structure than the master BOQ.\n"
            "You cannot compare BOQs with different structures.\n\n"
            "Possible solutions:\n"
            "• Use a comparison file with the same structure\n"
            "• Create separate analyses for files with different structures\n"
            "• Verify you selected the correct comparison file"
        )
    else:
        action_guidance = (
            "The files have different structures and cannot be processed together.\n\n"
            "Please verify that you are using compatible files."
        )
    
    formatted_message = (
        f"{summary}\n\n"
        f"Details ({position_count} mismatches found):\n{error_details}\n\n"
        f"{action_guidance}"
    )
    
    return formatted_message

def _log_validation_failure(validation_result, context="", operation="validation"):
    """
    Standardized logging for validation failures
    
    Args:
        validation_result: Validation result dictionary
        context: Operation context (e.g., "Use Mapping", "Compare Full")
        operation: Type of operation being performed
    """
    if not validation_result:
        logger.error(f"{context} {operation}: No validation result provided")
        return
    
    if validation_result.get('is_valid', True):
        logger.info(f"{context} {operation}: Validation passed successfully")
        return
    
    # Log summary
    summary = validation_result.get('summary', 'Unknown validation failure')
    logger.error(f"{context} {operation} failed: {summary}")
    
    # Log detailed errors
    errors = validation_result.get('errors', [])
    logger.error(f"{context} {operation}: Found {len(errors)} validation errors:")
    for i, error in enumerate(errors[:10], 1):  # Log first 10 errors
        logger.error(f"  {i}. {error}")
    
    if len(errors) > 10:
        logger.error(f"  ... and {len(errors) - 10} more errors")
    
    # Log mismatched positions summary
    mismatched_positions = validation_result.get('mismatched_positions', [])
    if mismatched_positions:
        logger.debug(f"{context} {operation}: Mismatched positions details:")
        for position_info in mismatched_positions[:5]:  # Log first 5 for debugging
            if 'expected_description' in position_info:
                logger.debug(f"  Position {position_info.get('position', 'unknown')}: "
                           f"expected '{position_info.get('expected_description', '')}', "
                           f"got '{position_info.get('actual_description', '')}'")
            else:
                logger.debug(f"  Missing position {position_info.get('position', 'unknown')}: "
                           f"expected '{position_info.get('expected_description', '')}'")

def _handle_validation_failure(validation_result, context="", operation="validation", show_dialog=True):
    """
    Standardized validation failure handling with logging and user notification
    
    Args:
        validation_result: Validation result dictionary
        context: Operation context for error messages
        operation: Type of operation being performed
        show_dialog: Whether to show error dialog to user
        
    Returns:
        bool: False (indicating failure)
        
    Raises:
        PositionValidationError: Always raises this exception to terminate processing
    """
    # Log the failure
    _log_validation_failure(validation_result, context, operation)
    
    # Format user message
    error_message = _format_validation_error_message(validation_result, context)
    
    # Show dialog if requested
    if show_dialog:
        dialog_title = f"{context} - Structure Validation Failed" if context else "Structure Validation Failed"
        messagebox.showerror(dialog_title, error_message)
    
    # Log termination
    logger.warning(f"{context} {operation}: Process terminated due to validation failure")
    
    # Raise exception to terminate processing
    raise PositionValidationError(error_message, validation_result)

# Optional imports
try:
    from ttkthemes import ThemedTk
    THEME_AVAILABLE = True
except ImportError:
    THEME_AVAILABLE = False

try:
    import tkinterdnd2 as TkinterDnD
    DND_AVAILABLE = True
    DND_FILES = TkinterDnD.DND_FILES
except ImportError:
    DND_AVAILABLE = False

# Import settings dialog
try:
    from ui.settings_dialog import show_settings_dialog
    SETTINGS_AVAILABLE = True
except ImportError:
    SETTINGS_AVAILABLE = False

# Import sheet categorization dialog
try:
    from ui.sheet_categorization_dialog import show_sheet_categorization_dialog
    SHEET_CATEGORIZATION_AVAILABLE = True
except ImportError:
    SHEET_CATEGORIZATION_AVAILABLE = False

print('SHEET_CATEGORIZATION_AVAILABLE forced to True')

# Import comparison row review dialog
try:
    from ui.comparison_row_review_dialog import show_comparison_row_review
    COMPARISON_ROW_REVIEW_AVAILABLE = True
except ImportError:
    COMPARISON_ROW_REVIEW_AVAILABLE = False

# Import offer info dialog
try:
    from ui.offer_info_dialog import show_offer_info_dialog
    OFFER_INFO_AVAILABLE = True
except ImportError:
    OFFER_INFO_AVAILABLE = False

# Import row review dialog
try:
    from ui.row_review_dialog import show_row_review_dialog
    ROW_REVIEW_AVAILABLE = True
except ImportError:
    ROW_REVIEW_AVAILABLE = False

# Import preview dialog
try:
    from ui.preview_dialog import show_preview_dialog
    PREVIEW_AVAILABLE = True
except ImportError:
    PREVIEW_AVAILABLE = False

# Import categorization dialogs
try:
    from ui.categorization_dialog import show_categorization_dialog
    from ui.category_review_dialog import show_category_review_dialog
    from ui.categorization_stats_dialog import show_categorization_stats_dialog
    CATEGORIZATION_AVAILABLE = True
except ImportError:
    CATEGORIZATION_AVAILABLE = False

# Color coding for confidence
def confidence_color(score):
    if score >= 0.8:
        return '#4CAF50'  # Green
    elif score >= 0.6:
        return '#FFC107'  # Amber
    else:
        return '#F44336'  # Red


def tooltip(widget, text):
    """Simple tooltip for a widget"""
    def on_enter(event):
        widget._tip = tk.Toplevel()
        widget._tip.wm_overrideredirect(True)
        x = event.x_root + 10
        y = event.y_root + 10
        widget._tip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(widget._tip, text=text, background="#ffffe0", relief='solid', borderwidth=1, font=("TkDefaultFont", 9))
        label.pack()
    def on_leave(event):
        if hasattr(widget, '_tip') and widget._tip is not None:
            widget._tip.destroy()
            widget._tip = None
    widget.bind('<Enter>', on_enter)
    widget.bind('<Leave>', on_leave)


from utils.format_utils import format_number_eu, excel_column_letter

class MainWindow:
    def __init__(self, controller, root=None):
        self.controller = controller
        self.root = root or self._create_root()
        self.root.title("BOQ Tools - Excel Processor")
        self.root.geometry("1200x700")
        self.root.minsize(900, 600)
        self._setup_style()
        self._create_menu()
        # Initialize variables before creating widgets
        self.open_files = {}
        self.status_var = tk.StringVar(value="Ready.")
        self.progress_var = tk.DoubleVar(value=0)
        self.status_bar_visible = True
        self.file_mapping = None
        self.sheet_treeviews = {}
        self.column_mapper = None  # Will be set when processing files
        self.row_validity = {}  # Initialize row validity dictionary
        self.row_review_treeviews = {}  # Initialize row review treeviews dictionary
        self.row_review_frame = None  # Initialize row review frame
        # ComparisonProcessor will handle comparison logic
        self.comparison_processor = None
        # Track comparison state
        self.is_comparison_workflow = False
        self.master_file_mapping = None
        # Add robust tab-to-file-mapping
        self.tab_id_to_file_mapping = {}
        # Offer name for summary grid
        self.current_offer_name = None
        # Store comprehensive offer information
        self.current_offer_info = None
        self.previous_offer_info = None  # For subsequent BOQs in comparison
        self.current_sheet_categories = None
        # Create widgets
        self._create_main_widgets()
        self._setup_drag_and_drop()
        self._bind_shortcuts()
        self._update_status("Welcome to BOQ Tools!")

    def _create_root(self):
        if THEME_AVAILABLE:
            return ThemedTk(theme="arc")
        elif DND_AVAILABLE:
            return TkinterDnD.Tk()
        else:
            return tk.Tk()

    def _setup_style(self):
        style = ttk.Style(self.root)
        if platform.system() == 'Windows':
            style.theme_use('vista')
        elif platform.system() == 'Darwin':
            style.theme_use('aqua')
        else:
            style.theme_use('clam')
        style.configure('TNotebook.Tab', padding=[10, 5])
        style.configure('Treeview', rowheight=28)
        style.map('Treeview', 
                 background=[('selected', '#B3E5FC')])  # Light blue background (same as column mapping)

    def _create_menu(self):
        menubar = tk.Menu(self.root)
        # File
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open...", accelerator="Ctrl+O", command=self.open_file)
        file_menu.add_command(label="Export", accelerator="Ctrl+E", command=self.export_file)
        file_menu.add_separator()
        file_menu.add_command(label="Clear All Files", command=self._clear_all_files)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", accelerator="Ctrl+Q", command=self.root.quit)
        menubar.add_cascade(label="File", menu=file_menu)
        # Edit
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Undo", accelerator="Ctrl+Z", command=lambda: self._update_status('Undo is not implemented.'))
        edit_menu.add_command(label="Redo", accelerator="Ctrl+Y", command=lambda: self._update_status('Redo is not implemented.'))
        menubar.add_cascade(label="Edit", menu=edit_menu)
        # View
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Toggle Status Bar", command=self._toggle_status_bar)
        menubar.add_cascade(label="View", menu=view_menu)
        # Tools
        tools_menu = tk.Menu(menubar, tearoff=0)
        tools_menu.add_command(label="Settings", command=self.open_settings)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        self.root.config(menu=menubar)

    def _create_main_widgets(self):
        # Configure root window grid
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Main frame using grid
        main_frame = ttk.Frame(self.root)
        main_frame.grid(row=0, column=0, sticky=tk.NSEW)
        
        # Configure main_frame grid layout
        main_frame.grid_rowconfigure(0, weight=0)  # Buttons zone
        main_frame.grid_rowconfigure(1, weight=1)  # Notebook (expandable)
        main_frame.grid_rowconfigure(2, weight=0)  # Status bar
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Top: Button frame with three buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=0, column=0, sticky=tk.EW, padx=10, pady=8)
        
        # Configure button frame columns to be equal width
        button_frame.grid_columnconfigure(0, weight=1)  # New Analysis
        button_frame.grid_columnconfigure(1, weight=1)  # Load Analysis
        button_frame.grid_columnconfigure(2, weight=1)  # Use Mapping
        
        # Create the three buttons
        new_analysis_btn = ttk.Button(button_frame, text="New Analysis", command=self.open_file)
        new_analysis_btn.grid(row=0, column=0, sticky=tk.EW, padx=5)
        tooltip(new_analysis_btn, "Start a new BOQ analysis by selecting Excel files")
        
        load_analysis_btn = ttk.Button(button_frame, text="Load Analysis", command=self._load_analysis)
        load_analysis_btn.grid(row=0, column=1, sticky=tk.EW, padx=5)
        tooltip(load_analysis_btn, "Load a previously saved analysis")
        
        use_mapping_btn = ttk.Button(button_frame, text="Use Mapping", command=self._use_mapping)
        use_mapping_btn.grid(row=0, column=2, sticky=tk.EW, padx=5)
        tooltip(use_mapping_btn, "Apply a previously saved mapping to new files")
        
        # Center: Tabbed interface for files
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, sticky=tk.NSEW, padx=10, pady=5)
        
        # Bottom: Status bar and progress
        self.status_frame = ttk.Frame(main_frame)
        self.status_frame.grid(row=2, column=0, sticky=tk.EW)
        
        # Configure status frame grid
        self.status_frame.grid_columnconfigure(0, weight=1)
        self.status_frame.grid_columnconfigure(1, weight=0)
        
        self.status_label = ttk.Label(self.status_frame, textvariable=self.status_var, anchor="w")
        self.status_label.grid(row=0, column=0, sticky=tk.EW, padx=8)
        
        self.progress = ttk.Progressbar(self.status_frame, variable=self.progress_var, maximum=100, length=180)
        self.progress.grid(row=0, column=1, padx=8)

    def _setup_drag_and_drop(self):
        if DND_AVAILABLE:
            try:
                # Enable drag and drop if available
                if hasattr(self.root, 'drop_target_register'):
                    self.root.drop_target_register(DND_FILES)
                if hasattr(self.root, 'dnd_bind'):
                    self.root.dnd_bind('<<Drop>>', self._on_drop)
            except AttributeError:
                # This can happen if TkinterDnD is available but methods aren't patched correctly (rare)
                logger.warning("TkinterDnD methods not found on root. Drag and drop will not be fully functional.")
                pass
            except Exception:
                pass
        # No else clause needed - buttons are already set up in _create_main_widgets

    def _bind_shortcuts(self):
        self.root.bind('<Control-o>', lambda e: self.open_file())
        self.root.bind('<Control-e>', lambda e: self.export_file())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        self.root.bind('<Control-z>', lambda e: self._update_status('Undo is not implemented.'))
        self.root.bind('<Control-y>', lambda e: self._update_status('Redo is not implemented.'))

    def _update_status(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def _toggle_status_bar(self):
        if self.status_bar_visible:
            self.status_frame.grid_remove()
        else:
            self.status_frame.grid()
        self.status_bar_visible = not self.status_bar_visible

    def _not_implemented(self):
        # Instead of a dialog, just update the status bar
        self._update_status("This feature is not implemented yet.")

    def open_settings(self):
        if SETTINGS_AVAILABLE:
            show_settings_dialog(self.root)
        else:
            self._not_implemented()

    def export_file(self):
        # Get the currently selected tab
        current_tab = self.notebook.nametowidget(self.notebook.select())
        if hasattr(current_tab, 'file_mapping'):
            # Implement export logic here
            self._update_status("Export functionality to be implemented.")
        else:
            self._update_status("No data to export.")

    def open_file(self):
        """Open file with enhanced comparison workflow support"""
        # Clear all previous data when starting new analysis
        self._clear_all_files()
        
        # Check if we're in comparison workflow
        if self.is_comparison_workflow and self.master_file_mapping:
            # We're in comparison mode, prompt for comparison file
            offer_info = self._prompt_offer_info(is_first_boq=False)
            if offer_info is None:
                self._update_status("Comparison file selection cancelled.")
                return
            
            # Prompt for comparison offer information FIRST (same as master workflow)
            comparison_offer_info = self._prompt_offer_info(is_first_boq=False)
            if comparison_offer_info is None:
                self._update_status("Comparison cancelled (no offer information provided)")
                return
            
            # Prompt for comparison file SECOND (same as master workflow)
            filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            comparison_file = filedialog.askopenfilename(
                title="Select Comparison BoQ File",
                filetypes=filetypes
            )
            
            if not comparison_file:
                self._update_status("Comparison cancelled")
                return
            
            self._process_comparison_file(comparison_file, offer_info)
            return
        
        # Normal file opening workflow
        offer_info = self._prompt_offer_info(is_first_boq=True)
        if offer_info is None:
            self._update_status("File open cancelled (no offer information provided).")
            return
        
        filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        filenames = filedialog.askopenfilenames(title="Open Excel File", filetypes=filetypes)
        for file in filenames:
            # Check if there is already a file loaded in the current tab
            current_tab_id = self.notebook.select()
            if current_tab_id:
                current_tab = self.notebook.nametowidget(current_tab_id)
                # If the tab has a final_dataframe, trigger comparison logic
                if hasattr(current_tab, 'final_dataframe') and getattr(current_tab, 'final_dataframe', None) is not None:
                    self._compare_full(current_tab)
                    return
            # Otherwise, just open as new analysis
            self._open_excel_file(file, offer_info)

    def _on_drop(self, event):
        if DND_AVAILABLE:
            files = self.root.tk.splitlist(event.data)
            for file in files:
                if file.lower().endswith(('.xlsx', '.xls')):
                    # For drag and drop, prompt for offer info first
                    offer_info = self._prompt_offer_info(is_first_boq=True)
                    if offer_info:
                        self._open_excel_file(file, offer_info)
                    else:
                        self._update_status("File drop cancelled (no offer information provided).")
                else:
                    self._update_status(f"Unsupported file: {file}")

    def _open_excel_file(self, filepath, offer_info=None):
        """Handle the file processing workflow"""
        if not filepath:
            return

        # Process as master BoQ

        # Clear previous results, including treeview references
        self.sheet_treeviews.clear()
        if self.notebook:
            for tab in self.notebook.tabs():
                self.notebook.forget(tab)

        self._update_status(f"Processing {os.path.basename(filepath)}, please wait...")
        self.progress_var.set(0)

        # Create a new tab for the file
        filename = os.path.basename(filepath)
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=filename)
        self.notebook.select(tab)

        # Configure grid for the tab frame
        tab.grid_rowconfigure(0, weight=1)
        tab.grid_columnconfigure(0, weight=1)

        loading_label = ttk.Label(tab, text="Analyzing file...")
        loading_label.grid(row=0, column=0, pady=40, padx=100)
        self.root.update_idletasks()

        def process_in_thread():
            """Runs the file processing in a background thread."""
            try:
                # Step 1: Load the file and get visible sheets
                from core.file_processor import ExcelProcessor
                processor = ExcelProcessor()
                processor.load_file(Path(filepath))
                visible_sheets = processor.get_visible_sheets()
                if not visible_sheets:
                    self.root.after(0, self._on_processing_error, tab, filename, loading_label)
                    return
                
                print(f"Loaded {len(visible_sheets)} visible sheets: {visible_sheets}")
                
                # Step 2: Ask user to categorize sheets - schedule this on main thread
                def ask_categorization():
                    print('About to show sheet categorization dialog')
                    if SHEET_CATEGORIZATION_AVAILABLE:
                        categories = show_sheet_categorization_dialog(self.root, visible_sheets)
                        print('Returned from sheet categorization dialog')
                        if not categories:
                            # User cancelled
                            self._update_status("Sheet categorization cancelled.")
                            loading_label.destroy()
                            return
                        boq_sheets = [sheet for sheet, cat in categories.items() if cat == "BOQ"]
                        if not boq_sheets:
                            self._update_status("No sheets marked as BOQ. Processing aborted.")
                            loading_label.destroy()
                            return
                    else:
                        # Fallback: treat all sheets as BOQ
                        boq_sheets = visible_sheets
                        categories = {sheet: "BOQ" for sheet in visible_sheets}
                    
                    # Store the categories for later use in processing
                    self.current_sheet_categories = categories
                    
                    print(f"Processing {len(boq_sheets)} BOQ sheets: {boq_sheets}")
                    
                    # Step 3: Process only BOQ sheets in a separate thread
                    def process_boq_sheets(offer_info_param):
                        try:
                            file_mapping = self.controller.process_file(
                                Path(filepath),
                                progress_callback=lambda p, m: self.root.after(0, self.update_progress, p, m),
                                sheet_filter=boq_sheets,
                                sheet_types=categories
                            )
                            # After processing, show the main window with BOQ sheets for column mapping
                            self.file_mapping = file_mapping
                            self.column_mapper = file_mapping.column_mapper if hasattr(file_mapping, 'column_mapper') else None
                            
                            # Apply sheet categories to the file mapping
                            if hasattr(self, 'current_sheet_categories'):
                                for sheet in file_mapping.sheets:
                                    if sheet.sheet_name in self.current_sheet_categories:
                                        sheet.sheet_type = self.current_sheet_categories[sheet.sheet_name]
                                        logger.debug(f"Applied sheet type '{sheet.sheet_type}' to sheet '{sheet.sheet_name}'")
                            
                            # Schedule the completion callback on the main thread
                            try:
                                self.root.after(0, self._on_processing_complete, tab, filepath, self.file_mapping, loading_label, offer_info_param)
                            except RuntimeError:
                                # If the main loop is not running, just log the completion
                                logger.info(f"Processing completed for {filepath} but main loop not available")
                        except Exception as e:
                            logger.error(f"Failed to process BOQ sheets: {e}", exc_info=True)
                            self.root.after(0, self._on_processing_error, tab, filename, loading_label)
                    
                    # Start the BOQ processing in a separate thread
                    threading.Thread(target=lambda: process_boq_sheets(offer_info), daemon=True).start()
                
                # Schedule the categorization question on the main thread immediately
                try:
                    print("Scheduling ask_categorization on main thread")
                    self.root.after(0, ask_categorization)
                except RuntimeError:
                    logger.error("Could not schedule categorization dialog - main loop not available")
                    self._on_processing_error(tab, filename, loading_label)
                    
            except Exception as e:
                logger.error(f"Failed to process file {filepath}: {e}", exc_info=True)
                # Schedule the error callback on the main thread
                try:
                    self.root.after(0, self._on_processing_error, tab, filename, loading_label)
                except RuntimeError:
                    logger.error(f"Could not schedule error callback - main loop not available")

        # Start the processing thread
        try:
            threading.Thread(target=process_in_thread, daemon=True).start()
        except Exception as e:
            logger.error(f"Failed to start processing thread: {e}")
            self._on_processing_error(tab, filename, loading_label)

    def update_progress(self, percentage, message):
        """Thread-safe method to update the progress bar and status label."""
        self.progress_var.set(percentage)
        self._update_status(message)

    def _on_processing_complete(self, tab, filepath, file_mapping, loading_widget, offer_info=None):
        """Handle processing completion"""
        logger.info(f"Processing complete for file: {filepath}")
        
        # Prevent setting file_mapping during comparison workflow
        if getattr(self, 'is_comparison_workflow', False):
            logger.info("Skipping file_mapping assignment during comparison workflow")
            loading_widget.destroy()
            return
        
        # Store the file mapping and column mapper
        self.file_mapping = file_mapping
        self.column_mapper = file_mapping.column_mapper if hasattr(file_mapping, 'column_mapper') else None
        
        # Use the passed offer_info or prompt for it if not provided
        if offer_info is None:
            offer_info = self._prompt_offer_info(is_first_boq=True)
            if not offer_info:
                self._update_status("Offer information cancelled.")
                loading_widget.destroy()
                return
        
        # Store offer info for later use
        self.current_offer_info = offer_info
        
        # Store offer info in the controller's current_files for summary data collection
        file_key = str(Path(filepath).resolve())
        if file_key in self.controller.current_files:
            # Enhanced offer info creation with better fallbacks
            offer_info_enhanced = {
                'offer_name': offer_info.get('offer_name', 'Unknown'),
                'project_name': offer_info.get('project_name', 'Unknown'),
                'project_size': offer_info.get('project_size', 'N/A'),
                'date': offer_info.get('date', datetime.now().strftime('%Y-%m-%d'))
            }
            logger.debug(f"Original offer_info: {offer_info}")
            logger.debug(f"Enhanced offer_info: {offer_info_enhanced}")
            
            # Store under dynamic offer name for comparison datasets
            offer_name = offer_info_enhanced['offer_name']
            if 'offers' not in self.controller.current_files[file_key]:
                self.controller.current_files[file_key]['offers'] = {}
            self.controller.current_files[file_key]['offers'][offer_name] = offer_info_enhanced
            # For backward compatibility, also store the last offer as 'offer_info'
            self.controller.current_files[file_key]['offer_info'] = offer_info_enhanced
            logger.debug(f"Stored offer info for offer_name '{offer_name}': {offer_info_enhanced}")
        
        # Also store offer info directly in the file_mapping object for easier access
        file_mapping.offer_info = offer_info_enhanced
        logger.debug(f"Stored offer info directly in file_mapping: {offer_info_enhanced}")
        
        # Also store in the current instance for immediate access
        self.current_file_mapping = file_mapping
        logger.debug(f"Stored current_file_mapping reference")
        
        # Remove loading widget and populate tab
        loading_widget.destroy()
        self._populate_file_tab(tab, file_mapping)
        
        # Use centralized refresh method
        logger.debug("Calling centralized summary grid refresh")
        self._refresh_summary_grid_centralized()
        
        # Update status
        self._update_status(f"Processing complete: {os.path.basename(filepath)}")

    def _on_processing_error(self, tab, filename, loading_widget):
        """Handle processing errors"""
        logger.error(f"Processing error for file: {filename}")
        
        # Remove loading widget
        if loading_widget:
            loading_widget.destroy()
        
        # Show error message
        messagebox.showerror("Processing Error", f"Failed to process file: {filename}")
        
        # Use centralized refresh method
        logger.debug("Calling centralized summary grid refresh after error")
        self._refresh_summary_grid_centralized()

    def _populate_file_tab(self, tab, file_mapping):
        # print("[DEBUG] _populate_file_tab called for tab:", tab)
        """Populates a tab with the processed data from a file mapping."""
        
        # Prevent tab population during comparison workflow
        if getattr(self, 'is_comparison_workflow', False):
            logger.info("Tab population prevented during comparison workflow")
            return
        
        # Debug: print all sheet names and their types
        # print('DEBUG: Sheets in file_mapping:')
        for s in file_mapping.sheets:
            print(f'  {s.sheet_name} (sheet_type={getattr(s, "sheet_type", None)})')
        
        # Clear any existing widgets (like loading/error labels)
        for widget in tab.winfo_children():
            widget.destroy()
        
        # Create main container frame
        tab_frame = ttk.Frame(tab)
        tab_frame.grid(row=0, column=0, sticky=tk.NSEW)
        
        # Configure tab_frame's grid layout
        tab_frame.grid_rowconfigure(0, weight=1)  # For sheet_notebook (main content, expands vertically)
        tab_frame.grid_rowconfigure(1, weight=0)  # For confirm_col_btn
        tab_frame.grid_rowconfigure(2, weight=1)  # For row_review_container (expands vertically)
        tab_frame.grid_columnconfigure(0, weight=1)  # Only one column, expands horizontally

        # Create sheet notebook for individual sheet tabs
        sheet_notebook = ttk.Notebook(tab_frame)
        sheet_notebook.grid(row=0, column=0, sticky=tk.NSEW, padx=5, pady=5)
        
        # Populate each sheet as a tab in the sheet_notebook
        for sheet in file_mapping.sheets:
            sheet_frame = ttk.Frame(sheet_notebook)
            sheet_notebook.add(sheet_frame, text=sheet.sheet_name)
            self._populate_sheet_tab(sheet_frame, sheet)
        
        # Add confirmation button for column mappings
        confirm_frame = ttk.Frame(tab_frame)
        confirm_frame.grid(row=1, column=0, sticky=tk.EW, padx=5, pady=5)
        confirm_frame.grid_columnconfigure(0, weight=1)
        confirm_btn = ttk.Button(confirm_frame, text="Confirm Column Mappings", command=self._save_all_mappings_for_all_sheets)
        confirm_btn.grid(row=0, column=0, sticky=tk.EW, padx=5, pady=5)
        
        # Row Review section will be created only after column mapping is confirmed
        self.row_review_frame = None
        self.row_review_notebook = None
        self.row_review_treeviews = {}
        self.row_validity = {}

        # Ensure file_mapping knows its tab for later lookup
        file_mapping.tab = tab
        # Store mapping from tab ID to file_mapping for robust lookup
        tab_id = str(tab)
        self.tab_id_to_file_mapping[tab_id] = file_mapping

    def _populate_sheet_tab(self, sheet_frame, sheet):
        """Populate an individual sheet tab with its data and column mappings."""
        # Configure sheet frame grid
        sheet_frame.grid_rowconfigure(0, weight=0)  # Header row control
        sheet_frame.grid_rowconfigure(1, weight=1)  # Column mappings table
        sheet_frame.grid_columnconfigure(0, weight=1)
        
        # Header Row Control
        header_control_frame = ttk.Frame(sheet_frame)
        header_control_frame.grid(row=0, column=0, sticky=tk.EW, padx=5, pady=5)
        
        # Get current header row (convert from 0-based to 1-based for display)
        current_header_row = getattr(sheet, 'header_row_index', 0) + 1
        header_row_var = tk.IntVar(value=current_header_row)
        
        # Get sheet data length for validation (use a reasonable default)
        # We'll load the file only when the user actually clicks the +/- buttons
        sheet_data_length = 50  # Reasonable default for most Excel files
        
        # Header row control widgets
        ttk.Label(header_control_frame, text="Header Row:").pack(side=tk.LEFT, padx=(0, 5))
        
        decrease_btn = ttk.Button(header_control_frame, text="−", width=3,
                                 command=lambda: self._on_header_row_decrease(sheet, header_row_var, entry_widget, 
                                                                             decrease_btn, increase_btn))
        decrease_btn.pack(side=tk.LEFT, padx=(0, 2))
        
        entry_widget = ttk.Entry(header_control_frame, textvariable=header_row_var, width=4, 
                                justify=tk.CENTER, state='readonly')
        entry_widget.pack(side=tk.LEFT, padx=(0, 2))
        
        increase_btn = ttk.Button(header_control_frame, text="+", width=3,
                                 command=lambda: self._on_header_row_increase(sheet, header_row_var, entry_widget,
                                                                             decrease_btn, increase_btn))
        increase_btn.pack(side=tk.LEFT)
        
        # Update button states based on current row
        self._update_header_row_buttons_simple(current_header_row, decrease_btn, increase_btn)
        
        # Column mappings table
        mappings_frame = ttk.LabelFrame(sheet_frame, text="Column Mappings (Double-click to edit) - Required columns are highlighted")
        mappings_frame.grid(row=1, column=0, sticky=tk.NSEW, padx=5, pady=5)
        mappings_frame.grid_rowconfigure(0, weight=1)
        mappings_frame.grid_columnconfigure(0, weight=1)
        
        # Add the propagate button to the left, above the Treeview in the mappings_frame
        propagate_btn = ttk.Button(mappings_frame, text="Apply These Mappings to All Other Sheets", command=lambda s=sheet: self._apply_mappings_to_all_sheets(s))
        propagate_btn.grid(row=99, column=0, sticky=tk.W, padx=5, pady=(0, 5))
        
        # Create treeview for column mappings
        columns = ("Column", "Original Header", "Mapped Type", "Confidence", "Required", "Actions")
        tree = ttk.Treeview(mappings_frame, columns=columns, show="headings", height=10)
        
        for col in columns:
            tree.heading(col, text=col)
            if col == "Column":
                tree.column(col, width=60, anchor=tk.CENTER)
            else:
                tree.column(col, width=150)
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(mappings_frame, orient=tk.VERTICAL, command=tree.yview)
        h_scrollbar = ttk.Scrollbar(mappings_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout for treeview and scrollbars
        tree.grid(row=0, column=0, sticky=tk.NSEW)
        v_scrollbar.grid(row=0, column=1, sticky=tk.NS)
        h_scrollbar.grid(row=1, column=0, sticky=tk.EW)
        
        # Required types (base required + successfully mapped new columns)
        if self.column_mapper and hasattr(self.column_mapper, 'config'):
            base_required_types = {col_type.value for col_type in self.column_mapper.config.get_required_columns()}
        else:
            base_required_types = {"description", "quantity", "unit_price", "total_price", "unit", "code"}
        
        # Add new columns as "required" if they are successfully mapped (confidence > 0)
        required_types = base_required_types.copy()
        if hasattr(sheet, 'column_mappings'):
            for mapping in sheet.column_mappings:
                mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                confidence = getattr(mapping, 'confidence', 0)
                # Treat scope, manhours, wage as required if successfully mapped
                if mapped_type in ['scope', 'manhours', 'wage'] and confidence > 0:
                    required_types.add(mapped_type)
        
        # Populate treeview with column mappings
        if hasattr(sheet, 'column_mappings'):
            for mapping in sheet.column_mappings:
                col_index = getattr(mapping, 'column_index', None)
                # Column indices supplied by MappingGenerator are zero-based.
                column_letter = excel_column_letter(col_index) if col_index is not None else "--"
                confidence = getattr(mapping, 'confidence', 0)
                mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                required = mapped_type in required_types
                original_header = getattr(mapping, 'original_header', 'Unknown')
                # Determine if this mapping was user-edited
                if hasattr(mapping, 'is_user_edited'):
                    actions = "Edited" if getattr(mapping, 'is_user_edited', False) else "Auto-detected"
                else:
                    # Only show 'Edited' if confidence==1.0 and the Actions field was set to 'Edited' in the edit dialog, it's user-edited
                    actions = "Edited" if (confidence == 1.0 and hasattr(mapping, 'user_edited') and getattr(mapping, 'user_edited', False)) else "Auto-detected"
                tags = []
                if required:
                    tags.append('required')
                tree.insert("", tk.END, values=(
                    column_letter,
                    original_header,
                    mapped_type,
                    f"{confidence:.1%}",
                    "Yes" if required else "No",
                    actions
                ), tags=tags)
        
        # Only highlight required fields (light blue)
        tree.tag_configure('required', background='#e0f0ff')
        
        # Store treeview reference for later access
        self.sheet_treeviews[sheet.sheet_name] = tree
        
        # Bind double-click to edit
        tree.bind('<Double-1>', lambda e, t=tree, s=sheet: self._edit_column_mapping(t, s))
    
    def _update_header_row_buttons(self, current_row, decrease_btn, increase_btn, max_rows):
        """Update the state of header row buttons based on current row"""
        # Disable decrease button if at minimum (row 1)
        if current_row <= 1:
            decrease_btn.config(state='disabled')
        else:
            decrease_btn.config(state='normal')
        
        # Disable increase button if at maximum
        if current_row >= max_rows:
            increase_btn.config(state='disabled')
        else:
            increase_btn.config(state='normal')
    
    def _update_header_row_buttons_simple(self, current_row, decrease_btn, increase_btn):
        """Update the state of header row buttons based on current row (simple version)"""
        # Disable decrease button if at minimum (row 1)
        if current_row <= 1:
            decrease_btn.config(state='disabled')
        else:
            decrease_btn.config(state='normal')
        
        # Increase button is always enabled (validation happens in reprocess method)
        increase_btn.config(state='normal')
    
    def _on_header_row_decrease(self, sheet, header_row_var, entry_widget, decrease_btn, increase_btn):
        """Decrease header row number and immediately refresh"""
        current_row = header_row_var.get()
        if current_row > 1:  # Minimum row 1
            new_row = current_row - 1
            header_row_var.set(new_row)
            
            # Update entry widget
            entry_widget.config(state='normal')
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, str(new_row))
            entry_widget.config(state='readonly')
            
            # Immediately reprocess with new header row (validation happens inside)
            success = self._reprocess_sheet_with_header_row(sheet, new_row - 1)  # Convert to 0-based
            
            # Update button states based on success and new row
            if success:
                self._update_header_row_buttons_simple(new_row, decrease_btn, increase_btn)
            else:
                # Revert on failure
                header_row_var.set(current_row)
                entry_widget.config(state='normal')
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, str(current_row))
                entry_widget.config(state='readonly')
    
    def _on_header_row_increase(self, sheet, header_row_var, entry_widget, decrease_btn, increase_btn):
        """Increase header row number and immediately refresh"""
        current_row = header_row_var.get()
        new_row = current_row + 1
        header_row_var.set(new_row)
        
        # Update entry widget
        entry_widget.config(state='normal')
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, str(new_row))
        entry_widget.config(state='readonly')
        
        # Immediately reprocess with new header row (validation happens inside)
        success = self._reprocess_sheet_with_header_row(sheet, new_row - 1)  # Convert to 0-based
        
        # Update button states based on success and new row
        if success:
            self._update_header_row_buttons_simple(new_row, decrease_btn, increase_btn)
        else:
            # Revert on failure
            header_row_var.set(current_row)
            entry_widget.config(state='normal')
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, str(current_row))
            entry_widget.config(state='readonly')
    
    def _reprocess_sheet_with_header_row(self, sheet, new_header_row_index):
        """Reprocess sheet with new header row and update UI"""
        try:
            # Get the current file path from the file mapping
            if not hasattr(self, 'file_mapping') or not self.file_mapping:
                messagebox.showerror("Error", "No file mapping available")
                return False
            
            # Get the file path from the metadata
            file_path = Path(self.file_mapping.metadata.file_path)
            if not file_path.exists():
                messagebox.showerror("Error", f"Original file not found: {file_path}")
                return False
            
            # Reload the file temporarily to get sheet data
            try:
                from core.file_processor import ExcelProcessor
                temp_processor = ExcelProcessor()
                temp_processor.load_file(file_path)
                
                # Get sheet data for the specific sheet
                sheet_data = temp_processor.get_sheet_data(sheet.sheet_name)
                
                if not sheet_data:
                    messagebox.showerror("Error", "No sheet data available for reprocessing")
                    return False
                
                # Validate the new header row index
                if new_header_row_index < 0 or new_header_row_index >= len(sheet_data):
                    messagebox.showerror("Error", f"Invalid header row {new_header_row_index + 1}. Sheet has {len(sheet_data)} rows.")
                    return False
                
            except Exception as e:
                logger.error(f"Failed to reload file and get sheet data: {e}")
                messagebox.showerror("Error", f"Failed to reload file: {str(e)}")
                return False
            
            # Use the column mapper to reprocess with forced header row
            if not self.column_mapper:
                messagebox.showerror("Error", "Column mapper not available")
                return False
            
            # Process with forced header row
            mapping_result = self.column_mapper.process_sheet_mapping_with_forced_header(
                sheet_data, new_header_row_index
            )
            
            # Update sheet object with new mapping results
            sheet.header_row_index = new_header_row_index
            sheet.confidence = mapping_result.overall_confidence
            
            # Convert mappings to the format expected by the sheet
            new_column_mappings = []
            for mapping in mapping_result.mappings:
                # Create a simple object with the required attributes
                col_mapping = type('ColumnMapping', (), {})()
                col_mapping.column_index = mapping.column_index
                col_mapping.original_header = mapping.original_header
                col_mapping.mapped_type = mapping.mapped_type.value
                col_mapping.confidence = mapping.confidence
                col_mapping.user_edited = True  # Mark as user-edited since user manually changed header row
                col_mapping.reasoning = mapping.reasoning
                new_column_mappings.append(col_mapping)
            
            sheet.column_mappings = new_column_mappings
            
            # CRITICAL FIX: Update the cached sheet data in the controller to reflect the new header row
            # This prevents data corruption during row classification
            if hasattr(self, 'controller') and self.controller:
                # Find the file key by matching the file mapping
                file_key = None
                for key, file_data in self.controller.current_files.items():
                    if file_data.get('file_mapping') == self.file_mapping:
                        file_key = key
                        break
                
                if file_key and hasattr(self.controller, 'current_files') and file_key in self.controller.current_files:
                    processor_results = self.controller.current_files[file_key].get('processor_results', {})
                    original_sheet_data = processor_results.get('sheet_data', {})
                    
                    # Update the cached sheet data with the new header row information
                    if sheet.sheet_name in original_sheet_data:
                        # Store the new header row index as metadata for this sheet
                        # This will be used during row classification to skip the correct header row
                        if 'header_row_indices' not in processor_results:
                            processor_results['header_row_indices'] = {}
                        processor_results['header_row_indices'][sheet.sheet_name] = new_header_row_index
                        
                        print(f"[DEBUG] Updated cached header row index for sheet '{sheet.sheet_name}' to {new_header_row_index}")
            
            # CRITICAL FIX: Set the sheet.sheet_data to the FULL original data
            # The row classifications contain indices that reference the original data structure
            # We should NOT remove the header row from sheet.sheet_data because the row indices
            # in the classifications are adjusted to account for the header row position
            sheet.sheet_data = sheet_data
            print(f"[DEBUG] Set sheet.sheet_data for '{sheet.sheet_name}' with full original data")
            print(f"[DEBUG] Data length: {len(sheet_data)}")
            print(f"[DEBUG] Header row index: {new_header_row_index}")
            if new_header_row_index < len(sheet_data):
                print(f"[DEBUG] Header row content: {sheet_data[new_header_row_index]}")
            if len(sheet_data) > 0:
                print(f"[DEBUG] First row of sheet_data: {sheet_data[0]}")
            if len(sheet_data) > 1:
                print(f"[DEBUG] Second row of sheet_data: {sheet_data[1]}")
            
            # Refresh the UI for this sheet
            self._refresh_single_sheet_tab(sheet)
            
            # Update status
            self._update_status(f"Header row updated to row {new_header_row_index + 1} for sheet '{sheet.sheet_name}'")
            
            return True
            
        except Exception as e:
            logger.error(f"Error reprocessing sheet with header row {new_header_row_index}: {e}")
            messagebox.showerror("Error", f"Failed to reprocess sheet: {str(e)}")
            return False
    
    def _refresh_single_sheet_tab(self, sheet):
        """Refresh the UI for a single sheet tab"""
        try:
            # Find the sheet tab in the notebook
            if not hasattr(self, 'file_mapping') or not self.file_mapping:
                return
            
            # Get the sheet treeview
            tree = self.sheet_treeviews.get(sheet.sheet_name)
            if not tree:
                return
            
            # Clear existing items
            for item in tree.get_children():
                tree.delete(item)
            
            # Required types (base required + successfully mapped new columns)
            if self.column_mapper and hasattr(self.column_mapper, 'config'):
                base_required_types = {col_type.value for col_type in self.column_mapper.config.get_required_columns()}
            else:
                base_required_types = {"description", "quantity", "unit_price", "total_price", "unit", "code"}
            
            # Add new columns as "required" if they are successfully mapped (confidence > 0)
            required_types = base_required_types.copy()
            if hasattr(sheet, 'column_mappings'):
                for mapping in sheet.column_mappings:
                    mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                    confidence = getattr(mapping, 'confidence', 0)
                    # Treat scope, manhours, wage as required if successfully mapped
                    if mapped_type in ['scope', 'manhours', 'wage'] and confidence > 0:
                        required_types.add(mapped_type)
            
            # Repopulate treeview with updated column mappings
            if hasattr(sheet, 'column_mappings'):
                for mapping in sheet.column_mappings:
                    confidence = getattr(mapping, 'confidence', 0)
                    mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                    required = mapped_type in required_types
                    original_header = getattr(mapping, 'original_header', 'Unknown')
                    
                    # Determine if this mapping was user-edited
                    actions = "Manual Header" if getattr(mapping, 'user_edited', False) else "Auto-detected"
                    
                    tags = []
                    if required:
                        tags.append('required')
                    
                    tree.insert("", tk.END, values=(
                        original_header,
                        mapped_type,
                        f"{confidence:.1%}",
                        "Yes" if required else "No",
                        actions
                    ), tags=tags)
            
        except Exception as e:
            logger.error(f"Error refreshing sheet tab: {e}")
            # Don't show error dialog as this is a background refresh

    def _edit_column_mapping(self, tree, sheet):
        selection = tree.selection()
        if not selection:
            return
        item = tree.item(selection[0])
        values = item['values']
        if not values:
            return
        column_name = values[0]
        # Find the column mapping object
        col_mapping = None
        for cm in getattr(sheet, 'column_mappings', []):
            if getattr(cm, 'original_header', None) == column_name:
                col_mapping = cm
                break
        if not col_mapping:
            return
        # Radio button options for mapped type
        mapped_type_options = [
            "description", "quantity", "unit_price", "total_price", "unit", "code", "scope", "manhours", "wage", "ignore"
        ]
        # Base required types + new columns if successfully mapped
        base_required_types = {"description", "quantity", "unit_price", "total_price", "unit", "code"}
        required_types = base_required_types.copy()
        if hasattr(sheet, 'column_mappings'):
            for mapping in sheet.column_mappings:
                mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                confidence = getattr(mapping, 'confidence', 0)
                if mapped_type in ['scope', 'manhours', 'wage'] and confidence > 0:
                    required_types.add(mapped_type)
        # Dialog to edit mapped type
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Column: {column_name}")
        dialog.geometry("500x650")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main content frame
        main_frame = ttk.Frame(dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header
        header_label = tk.Label(main_frame, text=f"Column: {column_name}", font=("Arial", 12, "bold"))
        header_label.pack(pady=(0, 10))
        
        # Current mapping info
        current_type = getattr(col_mapping, 'mapped_type', 'unknown')
        current_confidence = getattr(col_mapping, 'confidence', 0)
        info_text = f"Current: {current_type} (Confidence: {current_confidence:.1%})"
        info_label = ttk.Label(main_frame, text=info_text, foreground="gray")
        info_label.pack(pady=(0, 15))
        
        # Selection label
        ttk.Label(main_frame, text="Select new mapped type:").pack(pady=(0, 10))
        
        # Radio buttons for each mapped type
        radio_frame = ttk.Frame(main_frame)
        radio_frame.pack(pady=5, fill=tk.X)
        type_var = tk.StringVar(value=current_type)
        
        for opt in mapped_type_options:
            radio_btn = ttk.Radiobutton(radio_frame, text=opt, variable=type_var, value=opt)
            radio_btn.pack(anchor=tk.W, padx=10, pady=2)
            
            # Note: ttk.Radiobutton styling is handled by the theme system
        
        # Learning info frame
        learning_frame = ttk.LabelFrame(main_frame, text="Learning Information")
        learning_frame.pack(fill=tk.X, pady=15)
        
        learning_text = "When you save a required type mapping, it will be added to the system's learning database for future use."
        learning_label = ttk.Label(learning_frame, text=learning_text, wraplength=450, justify=tk.LEFT)
        learning_label.pack(padx=10, pady=10)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        def save():
            new_type = type_var.get()
            # Check for duplicate required type
            if new_type in required_types:
                for cm in getattr(sheet, 'column_mappings', []):
                    if cm is not col_mapping and getattr(cm, 'mapped_type', None) == new_type:
                        other_col_name = getattr(cm, 'original_header', '')
                        proceed = messagebox.askyesno(
                            "Duplicate Required Type",
                            f"The required type '{new_type}' is already assigned to column '{other_col_name}'.\nIf you continue, column '{other_col_name}' will be set to 'unknown'.\nContinue?"
                        )
                        if not proceed:
                            return
                        # Set the other column to unknown and update the treeview
                        cm.mapped_type = "unknown"
                        cm.confidence = 0.0
                        cm.user_edited = True  # Mark as user-edited since user demoted it
                        # Find the row in the treeview for the other column
                        for row_id in tree.get_children():
                            row_vals = tree.item(row_id)['values']
                            if row_vals and row_vals[0] == other_col_name:
                                tree.set(row_id, column="Mapped Type", value="unknown")
                                tree.set(row_id, column="Required", value="No")
                                tree.set(row_id, column="Confidence", value="0.0%")
                                tree.set(row_id, column="Actions", value="Edited")
                                tree.item(row_id, tags=())
                                break
                        break
            # Update the column mapping
            col_mapping.mapped_type = new_type
            col_mapping.confidence = 1.0  # User is always right
            col_mapping.user_edited = True  # Mark as user-edited
            # Learn from user confirmation for required types
            learning_message = ""
            if new_type in required_types and self.column_mapper:
                try:
                    self.column_mapper.update_canonical_mapping(column_name, new_type)
                    learning_message = f"\n\n✓ Learning: '{column_name}' has been added to the system's mapping database."
                    logger.info(f"Learned new mapping: '{column_name}' -> '{new_type}'")
                except Exception as e:
                    learning_message = f"\n\n⚠ Warning: Failed to save mapping to database: {e}"
                    logger.warning(f"Failed to update canonical mapping: {e}")
            # Update the treeview row
            tree.set(selection[0], column="Mapped Type", value=new_type)
            tree.set(selection[0], column="Confidence", value="100.0%")
            # Update the 'Required' field in the treeview row
            required_val = "Yes" if new_type in required_types else "No"
            tree.set(selection[0], column="Required", value=required_val)
            tree.set(selection[0], column="Actions", value="Edited")
            # Update row highlighting for required
            if required_val == "Yes":
                tree.item(selection[0], tags=("required",))
            else:
                tree.item(selection[0], tags=())
            dialog.destroy()
        
        # Buttons
        save_btn = ttk.Button(button_frame, text="Save & Learn", command=save)
        save_btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        cancel_btn = ttk.Button(button_frame, text="Cancel", command=dialog.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Set focus to save button
        save_btn.focus_set()
        
        # Bind Enter key to save
        dialog.bind('<Return>', lambda e: save())
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def _apply_mappings_to_all_sheets(self, source_sheet):
        # Propagate user-edited column mappings to all other sheets with the same Original Header (case and whitespace insensitive)
        if not self.file_mapping or not hasattr(self.file_mapping, 'sheets'):
            self._update_status("No file loaded to propagate mappings.")
            return
        user_edited = [cm for cm in getattr(source_sheet, 'column_mappings', []) if getattr(cm, 'confidence', 0) == 1.0]
        if not user_edited:
            self._update_status("No user-edited columns to propagate.")
            return
        count = 0
        # Base required types + new columns if successfully mapped
        base_required_types = {"description", "quantity", "unit_price", "total_price", "unit", "code"}
        required_types = base_required_types.copy()
        if hasattr(source_sheet, 'column_mappings'):
            for mapping in source_sheet.column_mappings:
                mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                confidence = getattr(mapping, 'confidence', 0)
                if mapped_type in ['scope', 'manhours', 'wage'] and confidence > 0:
                    required_types.add(mapped_type)
        affected_sheets = set()
        for edited_cm in user_edited:
            edited_header = getattr(edited_cm, 'original_header', None)
            if edited_header is None:
                continue
            edited_header_key = edited_header.strip().lower()
            for target_sheet in self.file_mapping.sheets:
                if not hasattr(target_sheet, 'column_mappings'):
                    continue
                # Skip the same column object in the source sheet
                if target_sheet is source_sheet:
                    continue
                for target_cm in target_sheet.column_mappings:
                    target_header = getattr(target_cm, 'original_header', None)
                    if target_header is not None and target_header.strip().lower() == edited_header_key:
                        # If required, demote any other column with this type in the target sheet
                        if edited_cm.mapped_type in required_types:
                            for other_cm in target_sheet.column_mappings:
                                if other_cm is not target_cm and other_cm.mapped_type == edited_cm.mapped_type:
                                    other_cm.mapped_type = "unknown"
                                    other_cm.confidence = 0.0
                                    other_cm.user_edited = True  # Mark as user-edited since user demoted it
                        target_cm.mapped_type = edited_cm.mapped_type
                        target_cm.confidence = 1.0
                        target_cm.user_edited = True  # Mark as user-edited (propagated)
                        count += 1
                        affected_sheets.add(getattr(target_sheet, 'sheet_name', None))
        # Always refresh the current sheet as well
        affected_sheets.add(getattr(source_sheet, 'sheet_name', None))
        # Refresh Treeviews for affected sheets
        for sheet_name in affected_sheets:
            tree = self.sheet_treeviews.get(sheet_name)
            if tree:
                # Find the corresponding sheet object
                target_sheet = next((s for s in self.file_mapping.sheets if getattr(s, 'sheet_name', None) == sheet_name), None)
                if target_sheet:
                    # Clear and repopulate the treeview
                    tree.delete(*tree.get_children())
                    for mapping in target_sheet.column_mappings:
                        confidence = getattr(mapping, 'confidence', 0)
                        mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                        required = mapped_type in required_types
                        original_header = getattr(mapping, 'original_header', 'Unknown')
                        # Determine if this mapping was user-edited
                        if hasattr(mapping, 'is_user_edited'):
                            actions = "Edited" if getattr(mapping, 'is_user_edited', False) else "Auto-detected"
                        else:
                            actions = "Edited" if getattr(mapping, 'user_edited', False) else "Auto-detected"
                        tags = []
                        if required:
                            tags.append('required')
                        tree.insert("", tk.END, values=(
                            original_header,
                            mapped_type,
                            f"{confidence:.1%}",
                            "Yes" if required else "No",
                            actions
                        ), tags=tags)
        messagebox.showinfo("Propagation Complete", f"Propagated {count} column mappings to all other sheets.")
        self._update_status(f"Propagated {count} column mappings to all other sheets.")

    def _save_all_mappings_for_all_sheets(self):
        # Prevent saving mappings during comparison workflow
        if getattr(self, 'is_comparison_workflow', False):
            logger.info("Saving mappings prevented during comparison workflow")
            return
            
        if not self.file_mapping or not hasattr(self.file_mapping, 'sheets'):
            messagebox.showwarning("No File", "No file loaded.")
            return
        
        # Save column mappings first
        total_saved = 0
        total_failed = 0
        total_already = 0
        for sheet in self.file_mapping.sheets:
            saved, failed, already = self._save_all_mappings(sheet, show_dialogs=False)
            total_saved += saved
            total_failed += failed
            total_already += already
        
        # Trigger row mapping with updated column mappings
        self._trigger_row_mapping()

    def _trigger_row_mapping(self):
        """Trigger row mapping with the current column mappings"""
        # Prevent row mapping during comparison workflow
        if getattr(self, 'is_comparison_workflow', False):
            logger.info("Row mapping prevented during comparison workflow")
            return
            
        if not self.file_mapping or not hasattr(self.file_mapping, 'sheets'):
            self._update_status("No file loaded for row mapping.")
            return
            
        # CRITICAL VALIDATION: Check for missing required columns before proceeding
        validation_result = self._validate_required_columns()
        if not validation_result['is_valid']:
            self._show_missing_columns_warning(validation_result)
            return
        
        try:
            # Import row classifier here to avoid circular imports
            from core.row_classifier import RowClassifier
            from utils.config import ColumnType
            
            row_classifier = RowClassifier()
            
            # Get the original sheet data from the controller
            if hasattr(self, 'controller') and self.controller:
                # Try to get the original sheet data from the controller
                # Find the file key by matching the file mapping
                file_key = None
                for key, file_data in self.controller.current_files.items():
                    if file_data.get('file_mapping') == self.file_mapping:
                        file_key = key
                        break
                
                if file_key and hasattr(self.controller, 'current_files') and file_key in self.controller.current_files:
                    processor_results = self.controller.current_files[file_key].get('processor_results', {})
                    original_sheet_data = processor_results.get('sheet_data', {})
                else:
                    # Fallback: try to reconstruct from current file mapping
                    original_sheet_data = {}
                    for sheet in self.file_mapping.sheets:
                        # This is a fallback - we might not have the original data
                        original_sheet_data[sheet.sheet_name] = []
            else:
                # Fallback when controller is not available
                original_sheet_data = {}
                for sheet in self.file_mapping.sheets:
                    original_sheet_data[sheet.sheet_name] = []
            
            # Process each sheet for row mapping
            updated_sheets = []
            for sheet in self.file_mapping.sheets:
                sheet_data = original_sheet_data.get(sheet.sheet_name, [])
                if not sheet_data:
                    # Skip if we don't have the original data
                    continue
                
                # CRITICAL FIX: Check if header row was manually changed and adjust sheet data accordingly
                header_row_index = 0  # Default header row index
                if hasattr(self, 'controller') and self.controller:
                    # Find the file key by matching the file mapping
                    file_key = None
                    for key, file_data in self.controller.current_files.items():
                        if file_data.get('file_mapping') == self.file_mapping:
                            file_key = key
                            break
                    
                    if file_key and hasattr(self.controller, 'current_files') and file_key in self.controller.current_files:
                        processor_results = self.controller.current_files[file_key].get('processor_results', {})
                        header_row_indices = processor_results.get('header_row_indices', {})
                        
                        # Use the manually set header row index if available
                        if sheet.sheet_name in header_row_indices:
                            header_row_index = header_row_indices[sheet.sheet_name]
                            print(f"[DEBUG] Using manually set header row index {header_row_index} for sheet '{sheet.sheet_name}'")
                        elif hasattr(sheet, 'header_row_index'):
                            header_row_index = sheet.header_row_index
                
                # Remove the header row from sheet data before classification to prevent confusion
                # The row classifier should only classify data rows, not header rows
                if header_row_index < len(sheet_data):
                    # Create a copy of sheet data without the header row
                    data_rows = sheet_data[:header_row_index] + sheet_data[header_row_index + 1:]
                    print(f"[DEBUG] Removed header row {header_row_index} from sheet '{sheet.sheet_name}' data for classification")
                else:
                    data_rows = sheet_data
                
                # Convert column mappings to the format expected by row classifier
                # Use the current column mappings from the sheet (which reflect any manual changes)
                column_mapping_dict = {}
                
                # Use the sheet's current column mappings
                for col_mapping in sheet.column_mappings:
                    try:
                        # Convert string column type to ColumnType enum
                        col_type = ColumnType(col_mapping.mapped_type)
                        # Use the column index as-is (it's already 0-based from the mapping result)
                        column_mapping_dict[col_mapping.column_index] = col_type
                    except ValueError:
                        # Skip unknown column types
                        continue
                
                # DEBUG: Log the column mapping and first few rows to diagnose indexing
                # Only debug the "Miscellaneous" sheet where Bank Guarantee is located
                if sheet.sheet_name == "Miscellaneous":
                    # print(f"[DEBUG] Column mapping for {sheet.sheet_name}: {column_mapping_dict}")
                    if sheet_data:
                        # print(f"[DEBUG] First row data: {sheet_data[0] if len(sheet_data) > 0 else 'No data'}")
                        # print(f"[DEBUG] Row data length: {len(sheet_data[0]) if sheet_data else 0}")
                        
                        # Show headers to understand column mapping
                        if hasattr(sheet, 'column_mappings'):
                            # print(f"[DEBUG] Headers for {sheet.sheet_name}:")
                            for i, col_mapping in enumerate(sheet.column_mappings):
                                header = getattr(col_mapping, 'original_header', f'Column_{i}')
                                mapped_type = getattr(col_mapping, 'mapped_type', 'unknown')
                                confidence = getattr(col_mapping, 'confidence', 0)
                                # print(f"[DEBUG] Column {col_mapping.column_index}: '{header}' -> {mapped_type} (confidence: {confidence:.2f})")
                        
                        # Show first 10 rows to find Bank Guarantee
                        for i in range(min(10, len(sheet_data))):
                            if any('Bank Guarantee' in str(cell) for cell in sheet_data[i]):
                                # print(f"[DEBUG] FOUND BANK GUARANTEE at row {i}: {sheet_data[i]}")
                                pass
                            else:
                                # print(f"[DEBUG] Row {i}: {sheet_data[i]}")
                                pass
                
                # Perform row classification using the data rows (without header row)
                row_classification_result = row_classifier.classify_rows(data_rows, column_mapping_dict, sheet.sheet_name)
                
                # Update the sheet's row classifications
                sheet.row_classifications = []
                for row_class in row_classification_result.classifications:
                    from core.mapping_generator import RowClassificationInfo
                    # CRITICAL FIX: Adjust row index to account for removed header row
                    # The row classifier processed data WITHOUT the header row, so its indices
                    # are already shifted down. We need to shift them back up to match the original data.
                    adjusted_row_index = row_class.row_index
                    if row_class.row_index >= header_row_index:
                        adjusted_row_index = row_class.row_index + 1
                    
                    print(f"[DEBUG] Row classification: original_index={row_class.row_index}, header_row={header_row_index}, adjusted_index={adjusted_row_index}")
                    
                    row_info = RowClassificationInfo(
                        row_index=adjusted_row_index,
                        row_type=row_class.row_type.value,
                        confidence=row_class.confidence,
                        completeness_score=row_class.completeness_score,
                        hierarchical_level=row_class.hierarchical_level,
                        section_title=row_class.section_title,
                        validation_errors=row_class.validation_errors,
                        reasoning=row_class.reasoning,
                        position=row_class.position
                    )
                    sheet.row_classifications.append(row_info)
                
                # Update sheet statistics
                sheet.row_count = len(data_rows)  # Use data rows count (excluding header)
                sheet.overall_confidence = row_classification_result.overall_quality_score
                
                # Update processing status based on row classification quality
                if row_classification_result.overall_quality_score >= 0.8:
                    sheet.processing_status = "SUCCESS"
                elif row_classification_result.overall_quality_score >= 0.6:
                    sheet.processing_status = "PARTIAL"
                else:
                    sheet.processing_status = "NEEDS_REVIEW"
                
                updated_sheets.append(sheet.sheet_name)
            
            # Update the UI to reflect the new row mappings
            self._refresh_sheet_tabs()
            
            # Show success message
            if updated_sheets:
                self._update_status(f"Row mapping completed for {len(updated_sheets)} sheets.")
                
            # After row mapping is complete and data is available:
            # Show the Row Review section with correct data
                pass
            self._show_row_review(self.file_mapping, original_sheet_data)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error during row mapping: {e}", exc_info=True)
            messagebox.showerror(
                "Row Mapping Error",
                f"An error occurred during row mapping:\n{str(e)}\n\n"
                f"Please check the logs for more details."
            )
            self._update_status(f"Row mapping failed: {str(e)}")

    def _validate_required_columns(self):
        """
        Validate that all required columns are properly mapped across all sheets.
        Returns a dictionary with validation results.
        """
        # Define required columns
        base_required_types = {"description", "quantity", "unit_price", "total_price", "unit", "code"}
        
        validation_result = {
            'is_valid': True,
            'missing_columns': {},  # sheet_name -> [missing_column_types]
            'unmapped_sheets': [],
            'total_sheets': 0,
            'valid_sheets': 0
        }
        
        if not self.file_mapping or not hasattr(self.file_mapping, 'sheets'):
            validation_result['is_valid'] = False
            return validation_result
        
        for sheet in self.file_mapping.sheets:
            validation_result['total_sheets'] += 1
            sheet_name = sheet.sheet_name
            
            # Get mapped types for this sheet
            mapped_types = set()
            if hasattr(sheet, 'column_mappings'):
                for mapping in sheet.column_mappings:
                    mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                    confidence = getattr(mapping, 'confidence', 0)
                    # Only count columns with reasonable confidence or user-edited
                    if confidence > 0 or getattr(mapping, 'user_edited', False):
                        mapped_types.add(mapped_type)
            
            # Check for missing required columns
            missing_columns = []
            for required_type in base_required_types:
                if required_type not in mapped_types:
                    missing_columns.append(required_type)
            
            if missing_columns:
                validation_result['is_valid'] = False
                validation_result['missing_columns'][sheet_name] = missing_columns
            else:
                validation_result['valid_sheets'] += 1
        
        return validation_result
    
    def _show_missing_columns_warning(self, validation_result):
        """
        Show a detailed warning dialog about missing required columns.
        """
        from tkinter import messagebox
        
        missing_info = validation_result['missing_columns']
        total_sheets = validation_result['total_sheets']
        valid_sheets = validation_result['valid_sheets']
        
        # Build detailed message
        message_parts = [
            f"⚠️ COLUMN MAPPING INCOMPLETE ⚠️",
            f"",
            f"Cannot proceed to row mapping because some required columns are not mapped.",
            f"",
            f"Status: {valid_sheets}/{total_sheets} sheets have complete column mappings.",
            f""
        ]
        
        if missing_info:
            message_parts.append("Missing required columns by sheet:")
            message_parts.append("")
            
            for sheet_name, missing_columns in missing_info.items():
                message_parts.append(f"📋 {sheet_name}:")
                for col in missing_columns:
                    message_parts.append(f"   • {col}")
                message_parts.append("")
        
        message_parts.extend([
            "Required columns for BOQ processing:",
            "• description (item descriptions)",
            "• quantity (quantities)",  
            "• unit (units of measurement)",
            "• unit_price (unit prices)",
            "• total_price (total prices)",
            "• code (position codes)",
            "",
            "Please:",
            "1. Review the column mappings in each sheet tab",
            "2. Double-click columns to edit their mapping",
            "3. Ensure all required columns are properly mapped",
            "4. Try row mapping again"
        ])
        
        full_message = "\n".join(message_parts)
        
        messagebox.showerror(
            "Missing Required Columns",
            full_message
        )
        
        self._update_status(f"Column mapping incomplete: {total_sheets - valid_sheets} sheets need attention")

    def _refresh_sheet_tabs(self):
        """Refresh the sheet tabs to show updated row mapping information"""
        if not self.file_mapping or not hasattr(self.file_mapping, 'sheets'):
            return
        
        # Find the current file tab
        current_tab = None
        for tab in self.notebook.tabs():
            if hasattr(tab, 'file_mapping') and tab.file_mapping == self.file_mapping:
                current_tab = tab
                break
        
        if current_tab:
            # Clear and repopulate the file tab
            for widget in current_tab.winfo_children():
                widget.destroy()
            
            # Repopulate with updated data
            self._populate_file_tab(current_tab, self.file_mapping)

    def _save_all_mappings(self, sheet, show_dialogs=True):
        """Save all required field mappings from the given sheet to the JSON file if not already present. Returns (saved, failed, already_present)."""
        if not self.column_mapper:
            if show_dialogs:
                messagebox.showwarning("No Column Mapper", "Column mapper not available. Please reload the file.")
            return 0, 0, 0
        # Base required types + new columns if successfully mapped
        base_required_types = {"description", "quantity", "unit_price", "total_price", "unit", "code"}
        required_types = base_required_types.copy()
        if hasattr(sheet, 'column_mappings'):
            for mapping in sheet.column_mappings:
                mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                confidence = getattr(mapping, 'confidence', 0)
                if mapped_type in ['scope', 'manhours', 'wage'] and confidence > 0:
                    required_types.add(mapped_type)
        saved_count = 0
        failed_count = 0
        already_present_count = 0
        current_mappings = self.column_mapper.get_canonical_mappings()
        for mapping in getattr(sheet, 'column_mappings', []):
            mapped_type = getattr(mapping, 'mapped_type', '')
            if mapped_type in required_types:
                original_header = getattr(mapping, 'original_header', '')
                confidence = getattr(mapping, 'confidence', 0)
                if original_header and mapped_type:
                    is_already_present = False
                    if mapped_type in current_mappings:
                        normalized_header = original_header.strip()
                        for existing_header in current_mappings[mapped_type]:
                            if existing_header.strip() == normalized_header:
                                is_already_present = True
                                break
                    if is_already_present:
                        already_present_count += 1
                        logger.debug(f"Mapping already present: '{original_header}' -> '{mapped_type}'")
                    else:
                        try:
                            self.column_mapper.update_canonical_mapping(original_header, mapped_type)
                            saved_count += 1
                            action_type = "user-edited" if confidence == 1.0 else "auto-detected"
                            logger.info(f"Saved {action_type} mapping: '{original_header}' -> '{mapped_type}' (confidence: {confidence:.1%})")
                        except Exception as e:
                            failed_count += 1
                            logger.warning(f"Failed to save mapping '{original_header}' -> '{mapped_type}': {e}")
        if show_dialogs:
            if saved_count > 0:
                messagebox.showinfo(
                    "Mappings Saved", 
                    f"Successfully saved {saved_count} new mappings to the database.\n"
                    f"These mappings will be used for future file processing.\n\n"
                    f"Already present: {already_present_count} mappings"
                )
                self._update_status(f"Saved {saved_count} new mappings to database.")
            else:
                if already_present_count > 0:
                    messagebox.showinfo(
                        "No New Mappings to Save", 
                        f"All {already_present_count} required field mappings are already in the database.\n"
                        f"No new mappings were saved."
                    )
                    self._update_status("All mappings already present in database.")
                else:
                    messagebox.showinfo(
                        "No Required Mappings Found", 
                        "No required type mappings found to save.\n"
                        "Make sure columns are mapped to required types first."
                    )
                    self._update_status("No required mappings found.")
            if failed_count > 0:
                messagebox.showwarning(
                    "Some Mappings Failed", 
                    f"{failed_count} mappings failed to save. Check the logs for details."
                )
        return saved_count, failed_count, already_present_count

    def _on_row_review_click(self, event, sheet_name, tree, required_cols):
        region = tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        row_id = tree.identify_row(event.y)
        if not row_id:
            return
        idx = int(row_id)
        # Toggle validity
        is_valid = self.row_validity[sheet_name].get(idx, True)
        new_valid = not is_valid
        self.row_validity[sheet_name][idx] = new_valid
        
        # Update validity state (no longer need to track manual invalidations)
        
        # Update tag and status column
        tag = 'validrow' if new_valid else 'invalidrow'
        tree.item(row_id, tags=(tag,))
        vals = list(tree.item(row_id, 'values'))
        vals[-1] = "Valid" if new_valid else "Invalid"
        tree.item(row_id, values=vals)

    def _show_row_review(self, file_mapping, original_sheet_data=None):
        logger.info("_show_row_review called")
        logger.info(f"File mapping type: {type(file_mapping)}")
        logger.info(f"File mapping offer_info: {getattr(file_mapping, 'offer_info', 'No offer_info')}")
        logger.info(f"Is comparison workflow: {getattr(self, 'is_comparison_workflow', False)}")
        logger.info(f"Pending comparison export: {getattr(self, '_pending_comparison_export', False)}")
        
        # Prevent normal row review during comparison workflow
        if getattr(self, 'is_comparison_workflow', False):
            logger.info("Normal row review prevented during comparison workflow")
            return
            
        # Remove existing Row Review frame if present
        if self.row_review_frame:
            self.row_review_frame.destroy()
        
        # Hide all widgets/frames related to column mapping in the current tab
        tab = self.notebook.nametowidget(self.notebook.select())
        self._hidden_column_mapping_widgets = []
        for child in tab.winfo_children():
            # Hide everything except the Row Review frame (if present)
            if child != getattr(self, 'row_review_frame', None):
                child.grid_remove()
                self._hidden_column_mapping_widgets.append(child)
        
        # Add Row Review container
        self.row_review_frame = ttk.LabelFrame(tab, text="Row Review")
        self.row_review_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=5, pady=5)
        self.row_review_frame.grid_rowconfigure(0, weight=1)
        self.row_review_frame.grid_columnconfigure(0, weight=1)
        
        # Add a notebook for row review per sheet
        self.row_review_notebook = ttk.Notebook(self.row_review_frame)
        self.row_review_notebook.grid(row=0, column=0, sticky=tk.NSEW)
        self.row_review_treeviews = {}
        self.row_validity = {}
        
        # Add a "Back to Column Mapping" button
        back_frame = ttk.Frame(self.row_review_frame)
        back_frame.grid(row=1, column=0, sticky=tk.EW, padx=5, pady=5)
        back_frame.grid_columnconfigure(0, weight=1)
        back_btn = ttk.Button(back_frame, text="← Back to Column Mapping", 
                             command=lambda: self._show_column_mapping())
        back_btn.pack(side=tk.LEFT, padx=5)

        # Add a "Confirm Row Review" button below the Row Review window
        confirm_row_frame = ttk.Frame(tab)
        confirm_row_frame.grid(row=1, column=0, sticky=tk.EW, padx=5, pady=5)
        confirm_row_frame.grid_columnconfigure(0, weight=1)
        confirm_row_btn = ttk.Button(confirm_row_frame, text="Confirm Row Review", command=self._on_confirm_row_review)
        confirm_row_btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.confirm_row_frame = confirm_row_frame

        for sheet_idx, sheet in enumerate(file_mapping.sheets):
            frame = ttk.Frame(self.row_review_notebook)
            self.row_review_notebook.add(frame, text=sheet.sheet_name)
            # Use standard mapped column names for display and value extraction
            if self.column_mapper and hasattr(self.column_mapper, 'config'):
                base_required_types = [col_type.value for col_type in self.column_mapper.config.get_required_columns()]
            else:
                base_required_types = ["description", "quantity", "unit_price", "total_price", "unit", "code"]
            
            # Add new columns as "required" if they are successfully mapped (confidence > 0)
            required_types = base_required_types.copy()
            if hasattr(sheet, 'column_mappings'):
                for mapping in sheet.column_mappings:
                    mapped_type = getattr(mapping, 'mapped_type', 'unknown')
                    confidence = getattr(mapping, 'confidence', 0)
                    # Treat scope, manhours, wage as required if successfully mapped
                    if mapped_type in ['scope', 'manhours', 'wage'] and confidence > 0:
                        if mapped_type not in required_types:
                            required_types.append(mapped_type)
            
            # Define the correct column order for display
            display_column_order = ["code", "description", "unit", "quantity", "unit_price", "total_price", "scope", "manhours", "wage"]
            
            # Use the display order, but only include columns that are actually mapped
            available_columns = []
            for col in display_column_order:
                if col in required_types:
                    available_columns.append(col)
            
            # Add any remaining required columns that weren't in the display order
            for col in required_types:
                if col not in available_columns:
                    available_columns.append(col)
            
            mapped_type_to_index = {}
            if hasattr(sheet, 'column_mappings'):
                for cm in sheet.column_mappings:
                    mapped_type_to_index[getattr(cm, 'mapped_type', None)] = cm.column_index
            
            # Use mapped type keys as columns (not uppercased)
            columns = ["#"] + available_columns + ["status"]
            tree = ttk.Treeview(frame, columns=columns, show="headings", height=12, selectmode="none")
            for col in columns:
                tree.heading(col, text=col.capitalize() if col != '#' else '#')
                if col == "status":
                    tree.column(col, width=80, anchor=tk.CENTER)
                else:
                    tree.column(col, width=120 if col != "#" else 40, anchor=tk.W, minwidth=50, stretch=False)
            # Add scrollbars
            v_scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
            h_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
            self.row_review_treeviews[sheet.sheet_name] = tree
            # Set selection highlight to light blue (same as column mapping)
            style = ttk.Style(tree)
            style.map('Treeview', 
                     background=[('selected', '#B3E5FC')])  # Light blue background
            style.layout('Treeview.Item', [('Treeitem.padding', {'sticky': 'nswe', 'children': [('Treeitem.indicator', {'side': 'left', 'sticky': ''}), ('Treeitem.image', {'side': 'left', 'sticky': ''}), ('Treeitem.text', {'side': 'left', 'sticky': ''})]})])
            # Populate rows
            self.row_validity[sheet.sheet_name] = {}
            if hasattr(sheet, 'row_classifications'):
                for rc in sheet.row_classifications:
                    row_data = getattr(rc, 'row_data', None)
                    if row_data is None and hasattr(sheet, 'sheet_data'):
                        try:
                            row_data = sheet.sheet_data[rc.row_index]
                        except Exception:
                            row_data = None
                    if row_data is None:
                        row_data = []
                    row_values = [rc.row_index + 1]
                    for col in available_columns:
                        idx = mapped_type_to_index.get(col)
                        val = row_data[idx] if idx is not None and idx < len(row_data) else ""
                        
                        # Apply number formatting for specific columns
                        if col in ['unit_price', 'total_price', 'wage']:
                            val = format_number_eu(val)
                        elif col in ['quantity', 'manhours']:
                            val = format_number_eu(val)
                        
                        row_values.append(val)
                    # New: Use new row validity logic based on master vs comparison mode
                    from core.row_classifier import RowClassifier
                    from utils.config import ColumnType
                    row_classifier = RowClassifier()
                    
                    # Convert column mappings to format expected by row classifier
                    column_mapping = {}
                    for cm in sheet.column_mappings:
                        try:
                            col_type = ColumnType(cm.mapped_type)
                            column_mapping[cm.column_index] = col_type
                        except ValueError:
                            continue
                    
                    # Use master validation criteria for all rows
                        is_valid = row_classifier.validate_master_row_validity(row_data, column_mapping)
                    
                    self.row_validity[sheet.sheet_name][rc.row_index] = is_valid
                    status = "Valid" if is_valid else "Invalid"
                    tag = 'validrow' if is_valid else 'invalidrow'
                    tree.insert('', 'end', iid=str(rc.row_index), values=row_values + [status], tags=(tag,))
            tree.tag_configure('validrow', background='#E8F5E9')  # light green
            tree.tag_configure('invalidrow', background='#FFEBEE')  # light red
            tree.bind('<Button-1>', lambda e, s=sheet.sheet_name, t=tree: self._on_row_review_click(e, s, t, required_types))
        # Optionally, select the first sheet by default
        if file_mapping.sheets:
            self.row_review_notebook.select(0)

    def _show_column_mapping(self):
        """Show the Column Mapping section and hide Row Review"""
        # Hide Row Review
        if self.row_review_frame:
            self.row_review_frame.destroy()
            self.row_review_frame = None
        # Restore all previously hidden column mapping widgets
        tab = self.notebook.nametowidget(self.notebook.select())
        if hasattr(self, '_hidden_column_mapping_widgets'):
            for widget in self._hidden_column_mapping_widgets:
                widget.grid()
            self._hidden_column_mapping_widgets = []

    def _on_confirm_row_review(self):
        """Handle row review confirmation and start categorization or export for comparison workflow"""
        logger.info("_on_confirm_row_review called")
        logger.info(f"Current tab ID: {self.notebook.select()}")
        logger.info(f"Current file mapping: {self.file_mapping}")
        logger.info(f"Pending comparison export: {getattr(self, '_pending_comparison_export', False)}")
        logger.info(f"Is comparison workflow: {getattr(self, 'is_comparison_workflow', False)}")
        
        # If this was a comparison workflow, continue with comparison processing
        if getattr(self, '_pending_comparison_export', False):
            logger.info("Comparison workflow detected in _on_confirm_row_review")
            # Reset flag immediately to prevent re-entry
            self._pending_comparison_export = False
            offer_info = getattr(self, '_pending_comparison_offer_info', None)
            self._pending_comparison_offer_info = None
            
            # Get the current tab and continue with comparison processing
            current_tab_id = self.notebook.select()
            if current_tab_id:
                current_tab = self.notebook.nametowidget(current_tab_id)
                logger.info("Calling _compare_full to continue comparison processing")
                self._compare_full(current_tab)
            return
            
        # Additional safeguard: prevent normal workflow during comparison
        if getattr(self, 'is_comparison_workflow', False):
            logger.info("Normal workflow prevented during comparison workflow")
            return
        # --- Normal master BOQ workflow ---
        self._update_status("Row review confirmed. Filtering valid rows and starting categorization process...")
        current_tab_id = self.notebook.select()
        file_mapping = self.tab_id_to_file_mapping.get(current_tab_id)
        if not file_mapping:
            messagebox.showerror("Error", "Could not find file mapping for categorization")
            return
        
        # CRITICAL FIX: Store row validity in file_mapping so controller can access it
        file_mapping.row_validity = self.row_validity.copy()
        
        try:
            import pandas as pd
            rows = []
            sheet_count = 0
            
            # FILTER: Only process valid rows from each sheet
            for sheet in getattr(file_mapping, 'sheets', []):
                if getattr(sheet, 'sheet_type', 'BOQ') != 'BOQ':
                    logger.debug(f"Skipping sheet {sheet.sheet_name} with type {getattr(sheet, 'sheet_type', 'Unknown')}")
                    continue
                sheet_count += 1
                col_headers = [cm.mapped_type for cm in getattr(sheet, 'column_mappings', [])]
                sheet_name = sheet.sheet_name
                validity_dict = self.row_validity.get(sheet_name, {})
                
                # FILTER: Only include rows that are marked as valid
                valid_row_classifications = []
                for rc in getattr(sheet, 'row_classifications', []):
                    if validity_dict.get(rc.row_index, True):  # Only include valid rows
                        valid_row_classifications.append(rc)
                
                # Process only valid rows
                for rc in valid_row_classifications:
                    row_data = getattr(rc, 'row_data', None)
                    if row_data is None and hasattr(sheet, 'sheet_data'):
                        try:
                            row_data = sheet.sheet_data[rc.row_index]
                        except Exception:
                            row_data = None
                    if row_data is None:
                        row_data = []
                    row_dict = {}
                    for cm in sheet.column_mappings:
                        mapped_type = getattr(cm, 'mapped_type', None)
                        if not mapped_type:
                            continue
                        idx = cm.column_index
                        row_dict[mapped_type] = row_data[idx] if idx < len(row_data) else ''
                    for mt in col_headers:
                        if mt not in row_dict:
                            row_dict[mt] = ''
                    
                    # Ensure Source_Sheet column is always included
                    if 'Source_Sheet' not in row_dict:
                        row_dict['Source_Sheet'] = sheet_name
                    else:
                        # If it exists but is empty, populate it
                        if not row_dict['Source_Sheet'] or str(row_dict['Source_Sheet']).strip() == '':
                            row_dict['Source_Sheet'] = sheet_name
                    
                    rows.append(row_dict)
            
            if not rows:
                messagebox.showerror("Error", "No valid rows found for categorization.")
                return
            
            # Create filtered dataframe with only valid rows
            final_dataframe = pd.DataFrame(rows)
            logger.info(f"Filtered dataset created with {len(rows)} valid rows from {sheet_count} sheets")
            
            # Ensure Source_Sheet column exists (backup check)
            if 'Source_Sheet' not in final_dataframe.columns:
                logger.warning("Source_Sheet column missing from DataFrame, adding it with 'Unknown' values")
                final_dataframe['Source_Sheet'] = 'Unknown'
            else:
                # Fill any missing/empty Source_Sheet values
                mask = final_dataframe['Source_Sheet'].isna() | (final_dataframe['Source_Sheet'].astype(str).str.strip() == '')
                if mask.any():
                    # Try to infer from sheet information if available
                    if hasattr(file_mapping, 'sheets') and file_mapping.sheets:
                        # This is a fallback - we already added it per row above
                        final_dataframe.loc[mask, 'Source_Sheet'] = 'Unknown'
                    else:
                        final_dataframe.loc[mask, 'Source_Sheet'] = 'Unknown'
            
            # Store the filtered dataframe in file_mapping for later use
            file_mapping.filtered_dataframe = final_dataframe
            
            # Phase 1.1: Pass DataFrame directly to categorization instead of file_mapping with sheet structure
            self._start_categorization(final_dataframe, file_mapping)
        except Exception as e:
            logger.error(f"Error during row review confirmation: {e}", exc_info=True)
            messagebox.showerror("Error", f"An error occurred during row review confirmation: {str(e)}")

    def _start_categorization(self, dataframe, file_mapping):
        """
        Start the categorization process
        
        Args:
            dataframe: DataFrame to categorize (unified structure with all required columns)
            file_mapping: File mapping object (for metadata and context)
        """
        if not CATEGORIZATION_AVAILABLE:
            messagebox.showerror("Error", "Categorization components not available")
            return
        
        if dataframe is None or dataframe.empty:
            messagebox.showerror("Error", "No data available for categorization")
            return
        
        try:
            # Phase 1.2: Pass DataFrame directly to categorization dialog instead of file_mapping
            dialog = show_categorization_dialog(
                parent=self.root,
                controller=self.controller,
                dataframe=dataframe,
                file_mapping=file_mapping,  # Still pass file_mapping for metadata/context
                on_complete=self._on_categorization_complete
            )
            
        except Exception as e:
            logger.error(f"Failed to start categorization: {e}")
            messagebox.showerror("Error", f"Failed to start categorization: {str(e)}")
    
    def _on_categorization_complete(self, final_dataframe, categorization_result):
        """Handle categorization completion for both master BOQ and comparison workflows"""
        try:
            current_tab_path = self.notebook.select()
            
            # Get the actual tab widget using nametowidget
            current_tab = self.notebook.nametowidget(self.notebook.select())
            
            logger.debug(f"Current tab path: {current_tab_path}")
            logger.debug(f"Current tab widget type: {type(current_tab)}")
            logger.debug(f"Current tab widget: {current_tab}")
            
            # Get file_mapping to check if this is subset categorization (comparison rows)
            file_mapping = None
            file_data_found = False
            
            # First, try to find the matching file data in controller's current_files
            for file_key, file_data in self.controller.current_files.items():
                if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                    file_mapping = file_data['file_mapping']
                    file_data_found = True
                    break
            
            # If not found in current_files, try tab_id_to_file_mapping
            if not file_mapping:
                file_mapping = self.tab_id_to_file_mapping.get(current_tab_path)
            
            # Phase 1.4 & Phase 2: Handle subset categorization (comparison rows)
            if file_mapping and getattr(file_mapping, '_is_subset_categorization', False):
                logger.info("Merging categorized subset rows back into main dataset")
                
                # Get the original full DataFrame
                original_df = getattr(file_mapping, '_original_full_dataframe', None)
                if original_df is None or original_df.empty:
                    logger.error("Original DataFrame not found for merging categorized subset")
                    messagebox.showerror("Error", "Could not merge categorized rows: original dataset not found")
                    return
                
                # Merge categorized Category values back into the main dataset
                # Match by index if preserved, or by Description + code combination
                merged_df = original_df.copy()
                
                # Try matching by index first (if indices are preserved)
                for idx in final_dataframe.index:
                    if idx in merged_df.index:
                        # Direct index match - update Category
                        if 'Category' in final_dataframe.columns and 'Category' in merged_df.columns:
                            merged_df.at[idx, 'Category'] = final_dataframe.at[idx, 'Category']
                            logger.debug(f"Updated Category for row {idx} by index")
                    else:
                        # Index not preserved, match by Description + code
                        cat_row = final_dataframe.loc[idx]
                        desc = str(cat_row.get('Description', '')).strip()
                        code = str(cat_row.get('code', '')).strip()
                        category = cat_row.get('Category', '')
                        
                        if desc and 'Description' in merged_df.columns:
                            # Find matching rows
                            matches = merged_df[
                                (merged_df['Description'].astype(str).str.strip() == desc) &
                                (merged_df['code'].astype(str).str.strip() == code)
                            ]
                            
                            if len(matches) > 0:
                                # Update first match (should be unique)
                                match_idx = matches.index[0]
                                merged_df.at[match_idx, 'Category'] = category
                                logger.debug(f"Updated Category for row {match_idx} by Description+code match")
                
                # Update file_mapping with merged DataFrame
                file_mapping.dataframe = merged_df
                file_mapping.filtered_dataframe = merged_df  # Also update filtered_dataframe
                
                # Clear the subset categorization flag
                file_mapping._is_subset_categorization = False
                if hasattr(file_mapping, '_original_full_dataframe'):
                    delattr(file_mapping, '_original_full_dataframe')
                
                # Update final_dataframe to be the merged full dataset
                final_dataframe = merged_df
                
                logger.info("Successfully merged categorized subset rows back into main dataset")
            
            # Update file data stores with final categorized DataFrame
            if file_data_found:
                for file_key, file_data in self.controller.current_files.items():
                    if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                        file_data['final_dataframe'] = final_dataframe
                        file_data['categorization_result'] = categorization_result
                        break
            
            # Update file_mapping if found
            if file_mapping:
                file_mapping.final_dataframe = final_dataframe
                file_mapping.categorization_result = categorization_result
                file_mapping.dataframe = final_dataframe  # Ensure dataframe is updated
                
                # Also update filtered_dataframe if it exists
                if not hasattr(file_mapping, 'filtered_dataframe') or file_mapping.filtered_dataframe is None:
                    file_mapping.filtered_dataframe = final_dataframe
            
            # Update the tab with the final categorized data
            logger.debug(f"Calling _show_final_categorized_data with tab type: {type(current_tab)}")
            self._show_final_categorized_data(current_tab, final_dataframe, categorization_result)
            
            # Refresh summary grid
            logger.debug("Calling centralized summary grid refresh after categorization")
            self._refresh_summary_grid_centralized()
            
            # Update status
            self._update_status("Categorization complete. All rows are now categorized.")
                
        except Exception as e:
            logger.error(f"Error handling categorization completion: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Error handling categorization completion: {str(e)}")

    def _remap_comparison_columns(self, df):
        """
        Remap the columns of the comparison DataFrame to canonical names using ColumnMapper.
        """
        from core.column_mapper import ColumnMapper
        import pandas as pd
        
        mapper = ColumnMapper()
        headers = list(df.columns)
        
        # Log original headers for debugging
        logger.info(f"Original headers: {headers}")
        
        # First, try to map using ColumnMapper
        mapping = mapper.map_columns_to_types(headers)
        
        # Build a mapping from original header to canonical name
        header_map = {}
        for m in mapping:
            if m.mapped_type.value not in header_map.values():
                header_map[m.original_header] = m.mapped_type.value
        
        # Log the mapping for debugging
        logger.info(f"ColumnMapper mapping: {header_map}")
        
        # If ColumnMapper didn't find much, try manual mapping based on common patterns
        if len(header_map) < 3:  # If we found less than 3 columns
            logger.info("ColumnMapper found few matches, trying manual mapping...")
            
            # Manual mapping based on common patterns
            manual_mapping = {}
            for header in headers:
                header_lower = str(header).lower().strip()
                
                # Quantity patterns
                if any(keyword in header_lower for keyword in ['qty', 'quantity', 'no', 'number', 'count']):
                    manual_mapping[header] = 'quantity'
                
                # Description patterns
                elif any(keyword in header_lower for keyword in ['description', 'item', 'work', 'activity', 'task', 'detail']):
                    manual_mapping[header] = 'Description'
                
                # Unit price patterns
                elif any(keyword in header_lower for keyword in ['unit price', 'rate', 'price per unit', 'unit cost']):
                    manual_mapping[header] = 'unit_price'
                
                # Total price patterns
                elif any(keyword in header_lower for keyword in ['total', 'total price', 'total cost', 'value', 'amount']):
                    manual_mapping[header] = 'total_price'
                
                # Unit patterns
                elif any(keyword in header_lower for keyword in ['unit', 'measurement', 'uom', 'unit of measure']):
                    manual_mapping[header] = 'unit'
                
                # Code patterns
                elif any(keyword in header_lower for keyword in ['code', 'item code', 'reference', 'ref', 'item no']):
                    manual_mapping[header] = 'code'
            
            # Update header_map with manual mappings
            header_map.update(manual_mapping)
            logger.info(f"Manual mapping additions: {manual_mapping}")
        
        # Rename columns
        df_renamed = df.rename(columns=header_map)
        
        # Ensure we have all required columns with proper casing
        required_columns = {
            'quantity': 'quantity',
            'unit_price': 'unit_price', 
            'total_price': 'total_price',
            'Description': 'Description',
            'code': 'code',
            'unit': 'unit'
        }
        
        # Check if we have the required columns (case-insensitive)
        missing_columns = []
        for required_col, canonical_name in required_columns.items():
            found = False
            for col in df_renamed.columns:
                if col.lower() == required_col.lower():
                    # Rename to canonical name if needed
                    if col != canonical_name:
                        df_renamed = df_renamed.rename(columns={col: canonical_name})
                    found = True
                    break
            if not found:
                missing_columns.append(canonical_name)
        
        if missing_columns:
            logger.warning(f"Missing required columns after mapping: {missing_columns}")
            # Add empty columns for missing required columns
            for col in missing_columns:
                df_renamed[col] = None
        
        logger.info(f"Final columns: {list(df_renamed.columns)}")
        return df_renamed

    def _compare_full(self, tab):
        """
        Optimized comparison workflow using master BOQ structure
        
        Args:
            tab: The tab containing the master BoQ data
        """
        try:
            from core.comparison_engine import ComparisonProcessor
            import pandas as pd
            
            # Get the current tab ID and file mapping
            current_tab_id = self.notebook.select()
            master_file_mapping = self.tab_id_to_file_mapping.get(current_tab_id)
            
            if not master_file_mapping:
                messagebox.showerror("Error", "Could not find master file mapping")
                return
            
            # Create unified master dataset with consistent structure
            master_df = self._create_unified_dataframe(master_file_mapping, is_master=True)
            if master_df is None or master_df.empty:
                messagebox.showerror("Error", "No data available for comparison")
                return
            
            # Prompt for comparison offer information FIRST (same as master workflow)
            comparison_offer_info = self._prompt_offer_info(is_first_boq=False)
            if comparison_offer_info is None:
                self._update_status("Comparison cancelled (no offer information provided)")
                return
            
            # Prompt for comparison file SECOND (same as master workflow)
            filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            comparison_file = filedialog.askopenfilename(
                title="Select Comparison BoQ File",
                filetypes=filetypes
            )
            
            if not comparison_file:
                self._update_status("Comparison cancelled")
                return
            
            # Process comparison file using optimized method
            self._update_status("Processing comparison file...")
            
            # Process comparison file using the same logic as "Use Mapping"
            self._update_status("Processing comparison file...")
            
            # Use the same file processing logic as "Use Mapping" workflow but adapted for comparison
            comparison_file_mapping = self._process_comparison_file_with_master_mappings(
                comparison_file, 
                master_file_mapping,  # Use master BOQ mappings
                comparison_offer_info
            )
            
            if not comparison_file_mapping:
                messagebox.showerror("Error", "Failed to process comparison file")
                return
            
            # Set comparison workflow state
            self.is_comparison_workflow = True
            self.master_file_mapping = master_file_mapping
            self.comparison_processor = ComparisonProcessor()
            
            # Load master dataset
            self.comparison_processor.load_master_dataset(master_df)
            
            # Create unified comparison dataset with consistent structure
            comparison_df = self._create_unified_dataframe(comparison_file_mapping, is_master=False)
            if comparison_df is None or comparison_df.empty:
                messagebox.showerror("Error", "No comparison data available")
                return
            
            # Debug: Log both datasets for verification
            logger.info(f"Master DataFrame shape: {master_df.shape}, columns: {list(master_df.columns)}")
            logger.info(f"Comparison DataFrame shape: {comparison_df.shape}, columns: {list(comparison_df.columns)}")
            
            # Verify both datasets have the same required columns
            master_cols = set(master_df.columns)
            comparison_cols = set(comparison_df.columns)
            missing_in_comparison = master_cols - comparison_cols
            missing_in_master = comparison_cols - master_cols
            
            if missing_in_comparison:
                logger.warning(f"Comparison dataset missing columns: {missing_in_comparison}")
                # Add missing columns to comparison dataset
                for col in missing_in_comparison:
                    comparison_df[col] = ''
            
            if missing_in_master:
                logger.warning(f"Master dataset missing columns: {missing_in_master}")
                # Add missing columns to master dataset
                for col in missing_in_master:
                    master_df[col] = ''
            
            # Ensure both datasets have the same column order
            all_columns = list(set(master_df.columns) | set(comparison_df.columns))
            master_df = master_df.reindex(columns=all_columns, fill_value='')
            comparison_df = comparison_df.reindex(columns=all_columns, fill_value='')
            
            self.comparison_processor.load_comparison_data(comparison_df)
            
            # Validate comparison data with enhanced error handling
            is_valid, message = self.comparison_processor.validate_comparison_data()
            if not is_valid:
                # Try to fix common validation issues
                logger.warning(f"Initial validation failed: {message}")
                
                # Check if it's a column mismatch issue
                if "missing columns" in message.lower():
                    # We've already handled column alignment above, so this shouldn't happen
                    logger.error(f"Column alignment failed despite our attempts: {message}")
                    messagebox.showerror("Validation Error", 
                                       f"Comparison data validation failed: {message}\n\n"
                                       f"Master columns: {list(master_df.columns)}\n"
                                       f"Comparison columns: {list(comparison_df.columns)}")
                    return
                else:
                    messagebox.showerror("Validation Error", f"Comparison data validation failed: {message}")
                    return
            
            # Process rows for validity
            self._update_status("Processing row validity...")
            row_results = self.comparison_processor.process_comparison_rows()
            
            # Show comparison row review dialog
            if COMPARISON_ROW_REVIEW_AVAILABLE:
                logger.info("Using comparison row review dialog")
                logger.info(f"Comparison file mapping: {comparison_file_mapping}")
                logger.info(f"Comparison offer info: {comparison_offer_info}")
                # Set flag to indicate this is a comparison workflow
                self._pending_comparison_export = True
                self._pending_comparison_offer_info = comparison_offer_info
                
                confirmed, updated_results = show_comparison_row_review(
                    self.root, 
                    comparison_file_mapping,  # Pass file mapping instead of DataFrame
                    row_results, 
                    comparison_offer_info.get('offer_name', 'Comparison')
                )
                
                logger.info(f"Comparison row review dialog returned: confirmed={confirmed}")
                logger.info(f"Updated results count: {len(updated_results) if updated_results else 0}")
                
                if not confirmed:
                    self._update_status("Comparison cancelled by user")
                    return
                
                # DEBUG: Export both datasets when user confirms row review
                # Create filtered comparison dataset with only valid rows
                valid_comparison_rows = [r for r in updated_results if r['is_valid']]
                valid_comparison_indices = [r['row_index'] for r in valid_comparison_rows]
                filtered_comparison_df = comparison_df.iloc[valid_comparison_indices].copy()
                
                # Store the filtered comparison dataset for later use
                comparison_file_mapping.filtered_dataframe = filtered_comparison_df
                
                # Reload comparison processor with filtered dataset (only valid rows)
                self.comparison_processor.load_comparison_data(filtered_comparison_df)
                
                # DEBUG EXPORT (COMMENTED OUT)
                # self._debug_export_datasets_before_merge(master_df, filtered_comparison_df, comparison_offer_info)
                
                # Update processor with user modifications
                self.comparison_processor.row_results = updated_results
            else:
                # Fallback: use original results
                updated_results = row_results
            
            # Process valid rows with MERGE/ADD logic
            self._update_status("Processing valid rows...")
            offer_name = comparison_offer_info.get('offer_name', 'Comparison')
            instance_results = self.comparison_processor.process_valid_rows(offer_name=offer_name)
            
            # Collect and display warnings
            all_warnings = []
            if self.comparison_processor.comparison_warnings:
                all_warnings.extend(self.comparison_processor.comparison_warnings)
                # Clear warnings after collecting
                self.comparison_processor.comparison_warnings.clear()
            
            if all_warnings:
                warning_messages = []
                for warning in all_warnings:
                    # Extract sheet name and row index from the warning object
                    # The row_index in ValidationIssue is 0-based, convert to 1-based for display
                    # The suggestion field might contain the sheet name for unit mismatches
                    sheet_name = "Unknown Sheet"
                    row_display_index = warning.row_index + 2 # Convert to Excel row number
                    
                    if warning.suggestion and "sheet" in warning.suggestion:
                        try:
                            # Extract sheet name from suggestion string
                            match = re.search(r"sheet '([^']+)'", warning.suggestion)
                            if match:
                                sheet_name = match.group(1)
                        except Exception:
                            pass # Fallback to "Unknown Sheet"
                    
                    # Format message based on warning type
                    if warning.validation_type == ValidationType.CONSISTENCY: # Unit mismatch
                        message = f"Unit Mismatch: Sheet '{sheet_name}', Row {row_display_index}. Master unit '{warning.expected_value}' vs. Comparison unit '{warning.actual_value}'."
                    elif warning.validation_type == ValidationType.DATA_TYPE: # Invalid data type
                        message = f"Invalid Data: Sheet '{sheet_name}', Row {row_display_index}, Column '{warning.column_index}'. Value '{warning.actual_value}' is not valid. Suggestion: {warning.suggestion}"
                    else:
                        message = f"Warning: Sheet '{sheet_name}', Row {row_display_index}. {warning.message}"
                    
                    warning_messages.append(message)
                
                full_warning_message = "The comparison completed with the following warnings:\n\n" + "\n".join(warning_messages)
                messagebox.showwarning("Comparison Warnings", full_warning_message)
            
            # Clean up data
            self._update_status("Cleaning up data...")
            cleanup_results = self.comparison_processor.cleanup_comparison_data()
            
            # Update the main dataset in place instead of showing results
            self._update_main_dataset_with_comparison_results(self.comparison_processor, comparison_offer_info)
            
            # Phase 2.1: Check for uncategorized rows and categorize them
            self._categorize_comparison_rows(comparison_offer_info)
            
            # Reset comparison workflow flags
            self.is_comparison_workflow = False
            self._pending_comparison_export = False
            self._pending_comparison_offer_info = None
            
            self._update_status("Comparison completed successfully")
            
        except Exception as e:
            logger.error(f"Error in _compare_full: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Comparison failed: {str(e)}")
            
            # Reset comparison workflow flags even on error
            self.is_comparison_workflow = False
            self._pending_comparison_export = False
            self._pending_comparison_offer_info = None

    def _process_comparison_file(self, filepath, offer_info):
        """
        Process comparison file using the SAME logic as master file processing
        
        Args:
            filepath: Path to comparison file
            offer_info: Offer information dictionary
            
        Returns:
            FileMapping object or None if failed
        """
        try:
            # Use the EXACT SAME processing logic as master file
            # This ensures identical data handling and calculations
            
            # Get visible sheets first
            from core.file_processor import ExcelProcessor
            processor = ExcelProcessor()
            processor.load_file(Path(filepath))
            visible_sheets = processor.get_visible_sheets()
            if not visible_sheets:
                return None
            
            # Use the same sheet categorization logic as master
            if SHEET_CATEGORIZATION_AVAILABLE:
                # For comparison, we'll use the same sheet structure as master
                # but we need to get the sheet categories from the master
                boq_sheets = visible_sheets  # Use all visible sheets for comparison
                categories = {sheet: "BOQ" for sheet in visible_sheets}
            else:
                # Fallback: treat all sheets as BOQ (same as master)
                boq_sheets = visible_sheets
                categories = {sheet: "BOQ" for sheet in visible_sheets}
            
            # Process using the SAME controller.process_file method as master
            file_mapping = self.controller.process_file(
                Path(filepath),
                progress_callback=lambda p, m: None,  # No progress callback for comparison
                sheet_filter=boq_sheets,
                sheet_types=categories
            )
            
            if not file_mapping:
                return None
            
            # Store offer information (same as master)
            file_mapping.offer_info = offer_info
            
            return file_mapping
            
        except Exception as e:
            logger.error(f"Error processing comparison file: {e}")
            return None

    def _validate_mapping_compatibility(self, master_mapping, comparison_mapping):
        """
        Validate that master and comparison mappings are compatible
        
        Args:
            master_mapping: FileMapping for master BoQ
            comparison_mapping: FileMapping for comparison BoQ
            
        Returns:
            dict: Validation result
        """
        try:
            # Check if both have sheets
            if not master_mapping.sheets or not comparison_mapping.sheets:
                return {
                    'is_valid': False,
                    'summary': 'One or both files have no sheets',
                    'errors': ['Missing sheets in one or both files']
                }
            
            # Check if sheet names match
            master_sheet_names = {sheet.sheet_name for sheet in master_mapping.sheets}
            comparison_sheet_names = {sheet.sheet_name for sheet in comparison_mapping.sheets}
            
            if master_sheet_names != comparison_sheet_names:
                missing_in_comparison = master_sheet_names - comparison_sheet_names
                extra_in_comparison = comparison_sheet_names - master_sheet_names
                
                errors = []
                if missing_in_comparison:
                    errors.append(f"Missing sheets in comparison: {missing_in_comparison}")
                if extra_in_comparison:
                    errors.append(f"Extra sheets in comparison: {extra_in_comparison}")
                
                return {
                    'is_valid': False,
                    'summary': 'Sheet names do not match between master and comparison files',
                    'errors': errors
                }
            
            # Check column mappings for each sheet
            errors = []
            for master_sheet in master_mapping.sheets:
                comparison_sheet = next(
                    (s for s in comparison_mapping.sheets if s.sheet_name == master_sheet.sheet_name),
                    None
                )
                
                if not comparison_sheet:
                    continue
                
                # Check column mappings
                master_columns = {cm.original_header: cm.mapped_type for cm in master_sheet.column_mappings}
                comparison_columns = {cm.original_header: cm.mapped_type for cm in comparison_sheet.column_mappings}
                
                if master_columns != comparison_columns:
                    missing_cols = set(master_columns.keys()) - set(comparison_columns.keys())
                    extra_cols = set(comparison_columns.keys()) - set(master_columns.keys())
                    different_mappings = set(master_columns.keys()) & set(comparison_columns.keys())
                    different_mappings = {col for col in different_mappings if master_columns[col] != comparison_columns[col]}
                    
                    if missing_cols:
                        errors.append(f"Sheet '{master_sheet.sheet_name}': Missing columns: {missing_cols}")
                    if extra_cols:
                        errors.append(f"Sheet '{master_sheet.sheet_name}': Extra columns: {extra_cols}")
                    if different_mappings:
                        errors.append(f"Sheet '{master_sheet.sheet_name}': Different column mappings: {different_mappings}")
            
            if errors:
                return {
                    'is_valid': False,
                    'summary': 'Column mappings do not match between master and comparison files',
                    'errors': errors
                }
            
            return {
                'is_valid': True,
                'summary': 'Mappings are compatible'
            }
            
        except Exception as e:
            return {
                'is_valid': False,
                'summary': f'Validation error: {str(e)}',
                'errors': [str(e)]
            }

    def _update_main_dataset_with_comparison_results(self, processor, offer_info):
        """
        Update the main dataset in place with comparison results
        
        Args:
            processor: ComparisonProcessor instance
            offer_info: Offer information dictionary
        """
        try:
            logger.info("=== STARTING _update_main_dataset_with_comparison_results ===")
            
            # Get the current tab (master dataset)
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                logger.warning("No current tab found for updating main dataset")
                return
            
            logger.info(f"Current tab ID: {current_tab_id}")
            
            # Get the current file mapping from the tab_id_to_file_mapping dictionary
            file_mapping = self.tab_id_to_file_mapping.get(current_tab_id)
            
            if not file_mapping:
                logger.warning("No file mapping found for current tab")
                return
            
            # Get the updated dataframe from the processor's master_dataset
            updated_df = processor.master_dataset.copy()
            
            if updated_df is None or updated_df.empty:
                logger.warning("No updated dataframe available from processor")
                return
            
            logger.info(f"Updating master dataset with {len(updated_df)} rows and {len(updated_df.columns)} columns")
            
            # Update the file mapping's dataframe with the merged data
            file_mapping.dataframe = updated_df.copy()
            
            # CRITICAL FIX: Store comparison offer info in controller's current_files
            # This ensures the comparison offer info is available for summary display
            if offer_info:
                offer_name = offer_info.get('offer_name', 'Comparison')
                logger.info(f"Storing comparison offer info for offer: {offer_name}")
                
                # Find the file key for the current tab
                file_key = None
                for key, file_data in self.controller.current_files.items():
                    if hasattr(file_data.get('file_mapping'), 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_id):
                        file_key = key
                        break
                
                if file_key:
                    # Ensure offers dictionary exists
                    if 'offers' not in self.controller.current_files[file_key]:
                        self.controller.current_files[file_key]['offers'] = {}
                    
                    # Store the comparison offer info
                    self.controller.current_files[file_key]['offers'][offer_name] = offer_info
                    logger.info(f"Stored comparison offer info: {offer_info}")
                else:
                    logger.warning("Could not find file key to store comparison offer info")
            
            # Reset comparison workflow flags to allow normal UI refresh
            self.is_comparison_workflow = False
            self.comparison_processor = None
            
            # Show success message
            merge_count = len(processor.merge_results) if hasattr(processor, 'merge_results') else 0
            add_count = len(processor.add_results) if hasattr(processor, 'add_results') else 0
            status_msg = f"Comparison completed: {merge_count} merges, {add_count} adds"
            self._update_status(status_msg)
            logger.info(f"Status message: {status_msg}")
            print("=== AFTER STATUS, BEFORE TRY ===")
            
            # Get the current tab widget
            current_tab = self.notebook.nametowidget(current_tab_id)
            logger.info("DEBUG: Got current tab widget successfully")
            
            # Debug: Log what we're about to do
            logger.info(f"About to update tab with comparison data")
            logger.info(f"Current tab type: {type(current_tab)}")
            logger.info(f"Updated dataframe shape: {updated_df.shape}")
            logger.info(f"Updated dataframe columns: {list(updated_df.columns)}")
            
            # Update the existing tab with comparison data, preserving formatting and column order
            logger.info("Calling _update_tab_with_comparison_data...")
            logger.info(f"DEBUG: current_tab type: {type(current_tab)}")
            logger.info(f"DEBUG: updated_df shape: {updated_df.shape}")
            logger.info(f"DEBUG: updated_df columns: {list(updated_df.columns)}")
            logger.info(f"DEBUG: About to call _update_tab_with_comparison_data method")
            logger.info(f"DEBUG: Method exists: {hasattr(self, '_update_tab_with_comparison_data')}")
            try:
                self._update_tab_with_comparison_data(current_tab, updated_df)
                logger.info("_update_tab_with_comparison_data completed successfully")
            except Exception as e:
                logger.error(f"Exception in _update_tab_with_comparison_data: {e}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
                raise
            
            logger.info("=== COMPLETED _update_main_dataset_with_comparison_results ===")
            print("=== END OF _update_main_dataset_with_comparison_results ===")
            
        except Exception as e:
            logger.error(f"Error updating main dataset with comparison results: {e}")
            self._update_status(f"Error updating dataset: {str(e)}")
    
    def _categorize_comparison_rows(self, offer_info):
        """
        Phase 2.1 & 2.2: Identify and categorize uncategorized rows from comparison results
        
        Args:
            offer_info: Offer information dictionary
        """
        try:
            if not CATEGORIZATION_AVAILABLE:
                logger.info("Categorization not available, skipping uncategorized row categorization")
                return
            
            # Get the current tab and file mapping
            current_tab_id = self.notebook.select()
            if not current_tab_id:
                logger.warning("No current tab found for categorizing comparison rows")
                return
            
            file_mapping = self.tab_id_to_file_mapping.get(current_tab_id)
            if not file_mapping:
                logger.warning("No file mapping found for categorizing comparison rows")
                return
            
            # Get the updated dataframe
            updated_df = file_mapping.dataframe if hasattr(file_mapping, 'dataframe') and file_mapping.dataframe is not None else None
            if updated_df is None or updated_df.empty:
                logger.info("No data available for categorizing comparison rows")
                return
            
            # Phase 2.2: Identify uncategorized rows
            # Check if Category column exists
            if 'Category' not in updated_df.columns:
                logger.info("No Category column found, adding it")
                updated_df['Category'] = ''
            
            # Identify uncategorized rows
            uncategorized_mask = (
                updated_df['Category'].isna() | 
                (updated_df['Category'] == '') | 
                (updated_df['Category'].astype(str).str.strip() == '')
            )
            uncategorized_rows = updated_df[uncategorized_mask].copy()
            
            if len(uncategorized_rows) == 0:
                logger.info("No uncategorized rows found after comparison")
                self._update_status("Comparison completed. All rows are categorized.")
                return
            
            logger.info(f"Found {len(uncategorized_rows)} uncategorized rows after comparison")
            
            # Ask user if they want to categorize the uncategorized rows
            response = messagebox.askyesno(
                "Categorize New Rows",
                f"Comparison completed. {len(uncategorized_rows)} new rows need categorization.\n\n"
                "These rows will be excluded from category totals in Summarize until categorized.\n\n"
                "Would you like to categorize them now?",
                icon='question'
            )
            
            if not response:
                logger.info("User chose not to categorize comparison rows now")
                self._update_status(f"Comparison completed. {len(uncategorized_rows)} rows are uncategorized.")
                return
            
            # Start categorization for uncategorized rows only
            self._update_status(f"Categorizing {len(uncategorized_rows)} new rows from comparison...")
            logger.info(f"Starting categorization for {len(uncategorized_rows)} uncategorized rows")
            
            # Store original DataFrame and indices for merging categorized results back
            # Preserve original indices in uncategorized_rows so we can merge back by index
            uncategorized_rows = uncategorized_rows.copy()  # Ensure we have a copy
            file_mapping._original_full_dataframe = updated_df.copy()  # Store for merging
            file_mapping._is_subset_categorization = True  # Flag to indicate subset categorization
            
            # Phase 2.2: Pass only uncategorized rows DataFrame to categorization
            self._start_categorization(uncategorized_rows, file_mapping)
            
        except Exception as e:
            logger.error(f"Error categorizing comparison rows: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to categorize comparison rows: {str(e)}")

    def _refresh_tab_with_updated_data(self, tab, file_mapping):
        """
        Simple UI refresh that shows the updated data without triggering column mapping workflow
        
        Args:
            tab: The tab widget to refresh
            file_mapping: The updated file mapping with new data
        """
        try:
            logger.info("Starting simple UI refresh for updated data")
            
            # Clear existing content
            for widget in tab.winfo_children():
                widget.destroy()
            
            # Create main container frame
            main_frame = ttk.Frame(tab)
            main_frame.grid(row=0, column=0, sticky=tk.NSEW)
            main_frame.grid_rowconfigure(0, weight=1)
            main_frame.grid_columnconfigure(0, weight=1)
            
            # Get the updated dataframe
            df = file_mapping.dataframe
            
            # Filter out ignore columns and get meaningful columns
            meaningful_columns = [col for col in df.columns if not (col.startswith('ignore') or col == 'ignore')]
            
            # Define the exact column order from row review (same as original)
            display_column_order = ["code", "description", "unit", "quantity", "unit_price", "total_price", "scope", "manhours", "wage"]
            
            # Add any offer-specific columns (those with [offer_name] pattern)
            offer_columns = [col for col in meaningful_columns if '[' in col and ']' in col]
            
            # Create final column order: standard columns first, then offer-specific columns
            final_columns = []
            for col in display_column_order:
                if col in meaningful_columns:
                    final_columns.append(col)
            
            # Add offer-specific columns at the end
            for col in offer_columns:
                if col not in final_columns:
                    final_columns.append(col)
            
            # Add any remaining meaningful columns
            for col in meaningful_columns:
                if col not in final_columns:
                    final_columns.append(col)
            
            # Create treeview for data display
            tree = ttk.Treeview(main_frame, columns=final_columns, show="headings", height=20)
            
            # Configure columns with proper formatting (same as row review)
            for col in final_columns:
                tree.heading(col, text=col)
                
                # Set column widths based on content type (same as row review)
                if col == "status":
                    tree.column(col, width=80, anchor=tk.CENTER)
                else:
                    tree.column(col, width=120 if col != "#" else 40, anchor=tk.W, minwidth=50, stretch=False)
            
            # Add scrollbars
            v_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
            h_scrollbar = ttk.Scrollbar(main_frame, orient=tk.HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
            
            # Pack treeview and scrollbars
            tree.grid(row=0, column=0, sticky=tk.NSEW, padx=5, pady=5)
            v_scrollbar.grid(row=0, column=1, sticky=tk.NS)
            h_scrollbar.grid(row=1, column=0, sticky=tk.EW)
            
            # Populate with data using exact same formatting as row review
            for idx, row in df.iterrows():
                values = []
                for col in final_columns:
                    val = row.get(col, '')
                    
                    # Apply exact same formatting as row review
                    if pd.notna(val) and val != '':
                        if col in ['unit_price', 'total_price', 'wage']:
                            # Currency formatting (same as row review)
                            val = format_number_eu(val)
                        elif col in ['quantity', 'manhours']:
                            # Number formatting (same as row review)
                            # Use standard European number formatting for all numeric columns
                            val = format_number_eu(val)
                        else:
                            val = str(val)
                    else:
                        val = ''
                    
                    values.append(val)
                
                tree.insert('', 'end', values=values)
            
            # Add subtle success indicator at the bottom
            status_frame = ttk.Frame(main_frame)
            status_frame.grid(row=2, column=0, sticky=tk.EW, pady=(5, 0))
            
            status_label = ttk.Label(status_frame, text="✓ Dataset updated with comparison results", 
                                   foreground="green", font=("Arial", 9))
            status_label.pack(side=tk.LEFT)
            
            logger.info("Simple UI refresh completed successfully")
            
        except Exception as e:
            logger.error(f"Error in simple UI refresh: {e}")
            # Fallback: just show a simple message
            for widget in tab.winfo_children():
                widget.destroy()
            
            message_frame = ttk.Frame(tab)
            message_frame.grid(row=0, column=0, sticky=tk.NSEW)
            message_frame.grid_rowconfigure(0, weight=1)
            message_frame.grid_columnconfigure(0, weight=1)
            
            message_label = ttk.Label(message_frame, 
                                    text="✓ Dataset updated successfully with comparison results\n\n"
                                         "The master dataset has been updated with the comparison data.",
                                    font=("Arial", 12), justify=tk.CENTER)
            message_label.grid(row=0, column=0, padx=50, pady=50)

    def _show_comparison_results(self, processor, offer_info):
        """
        Show comparison results to user
        
        Args:
            processor: ComparisonProcessor instance
            offer_info: Offer information dictionary
        """
        try:
            # Create results summary
            total_rows = len(processor.row_results)
            valid_rows = sum(1 for r in processor.row_results if r['is_valid'])
            invalid_rows = total_rows - valid_rows
            
            merge_count = len(processor.merge_results)
            add_count = len(processor.add_results)
            
            summary = f"""
Comparison Results for {offer_info.get('offer_name', 'Comparison')}:

Total Rows Processed: {total_rows}
Valid Rows: {valid_rows}
Invalid Rows: {invalid_rows}

Operations Performed:
- Merge Operations: {merge_count}
- Add Operations: {add_count}

Master Dataset Updated: {len(processor.master_dataset)} rows
            """
            
            messagebox.showinfo("Comparison Complete", summary)
            
            # Show detailed results in a new tab
            self._show_detailed_comparison_results(processor, offer_info)
            
        except Exception as e:
            logger.error(f"Error showing comparison results: {e}")
            messagebox.showerror("Error", f"Failed to show comparison results: {str(e)}")

    def _show_detailed_comparison_results(self, processor, offer_info):
        """
        Show detailed comparison results in a new tab
        
        Args:
            processor: ComparisonProcessor instance
            offer_info: Offer information dictionary
        """
        try:
            # Create a new tab for comparison results
            tab_title = f"Comparison Results - {offer_info.get('offer_name', 'Comparison')}"
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=tab_title)
            
            # Create the tab content
            self._populate_comparison_results_tab(tab, processor, offer_info)
            
            # Switch to the new tab
            self.notebook.select(tab)
            
        except Exception as e:
            logger.error(f"Error showing detailed comparison results: {e}")
            messagebox.showerror("Error", f"Failed to show detailed comparison results: {str(e)}")

    def _populate_comparison_results_tab(self, tab, processor, offer_info):
        """
        Populate the comparison results tab with detailed information
        
        Args:
            tab: The tab to populate
            processor: ComparisonProcessor instance
            offer_info: Offer information dictionary
        """
        try:
            # Create main frame
            main_frame = ttk.Frame(tab, padding="10")
            main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # Configure grid weights
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(0, weight=1)
            main_frame.columnconfigure(0, weight=1)
            main_frame.rowconfigure(1, weight=1)
            
            # Title
            title_label = ttk.Label(main_frame, text=f"Comparison Results - {offer_info.get('offer_name', 'Comparison')}", 
                                   font=("Arial", 14, "bold"))
            title_label.grid(row=0, column=0, pady=(0, 10), sticky=tk.W)
            
            # Create notebook for different views
            results_notebook = ttk.Notebook(main_frame)
            results_notebook.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
            
            # Summary tab
            summary_frame = ttk.Frame(results_notebook)
            results_notebook.add(summary_frame, text="Summary")
            self._populate_summary_frame(summary_frame, processor, offer_info)
            
            # Updated dataset tab
            dataset_frame = ttk.Frame(results_notebook)
            results_notebook.add(dataset_frame, text="Updated Dataset")
            self._populate_dataset_frame(dataset_frame, processor, offer_info)
            
            # Button frame
            button_frame = ttk.Frame(main_frame)
            button_frame.grid(row=2, column=0, pady=(10, 0), sticky=(tk.E, tk.W))
            button_frame.columnconfigure(1, weight=1)
            
            # Export button
            export_btn = ttk.Button(button_frame, text="Export Results", 
                                   command=lambda: self._export_comparison_results(processor, offer_info))
            export_btn.grid(row=0, column=0, padx=(0, 10))
            
            # Close button
            close_btn = ttk.Button(button_frame, text="Close", command=lambda: self.notebook.forget(tab))
            close_btn.grid(row=0, column=2)
            
        except Exception as e:
            logger.error(f"Error populating comparison results tab: {e}")
            raise

    def _populate_summary_frame(self, frame, processor, offer_info):
        """Populate the summary frame with comparison statistics"""
        try:
            # Statistics
            total_rows = len(processor.row_results)
            valid_rows = sum(1 for r in processor.row_results if r['is_valid'])
            invalid_rows = total_rows - valid_rows
            merge_count = len(processor.merge_results)
            add_count = len(processor.add_results)
            
            # Create statistics text
            stats_text = f"""
Comparison Statistics for {offer_info.get('offer_name', 'Comparison')}:

Processing Results:
• Total Rows Processed: {total_rows}
• Valid Rows: {valid_rows}
• Invalid Rows: {invalid_rows}

Operations Performed:
• Merge Operations: {merge_count}
• Add Operations: {add_count}

Dataset Information:
• Master Dataset Rows: {len(processor.master_dataset)}
• New Columns Created: {len([col for col in processor.master_dataset.columns if '[' in col])}

Offer Information:
• Offer Name: {offer_info.get('offer_name', 'N/A')}
• Client: {offer_info.get('client_name', 'N/A')}
• Project: {offer_info.get('project_name', 'N/A')}
• Date: {offer_info.get('date', 'N/A')}
            """
            
            # Create text widget
            text_widget = tk.Text(frame, wrap=tk.WORD, height=20, width=60)
            text_widget.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
            
            # Add scrollbar
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=text_widget.yview)
            scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
            text_widget.configure(yscrollcommand=scrollbar.set)
            
            # Insert text
            text_widget.insert(tk.END, stats_text)
            text_widget.config(state=tk.DISABLED)
            
            # Configure grid weights
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)
            
        except Exception as e:
            logger.error(f"Error populating summary frame: {e}")
            raise

    def _populate_dataset_frame(self, frame, processor, offer_info):
        """Populate the dataset frame with the updated master dataset"""
        try:
            # Create treeview
            columns = list(processor.master_dataset.columns)
            tree = ttk.Treeview(frame, columns=columns, show='headings', height=15)
            
            # Configure columns
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=100, minwidth=80)
            
            # Add scrollbars
            vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Grid layout
            tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
            vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
            hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
            
            # Populate with data (limit to first 100 rows for performance)
            for i, row in processor.master_dataset.head(100).iterrows():
                values = [str(val) if pd.notna(val) else "" for val in row.values]
                tree.insert('', 'end', values=values)
            
            # Configure grid weights
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)
            
        except Exception as e:
            logger.error(f"Error populating dataset frame: {e}")
            raise

    def _export_comparison_results(self, processor, offer_info):
        """Export comparison results to Excel with proper formatting and data validation"""
        try:
            filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
            filename = filedialog.asksaveasfilename(
                title="Export Comparison Results",
                filetypes=filetypes,
                defaultextension=".xlsx"
            )
            
            if filename:
                # Export the updated master dataset with proper formatting
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    processor.master_dataset.to_excel(writer, index=False, sheet_name='Comparison Results')
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Comparison Results']
                    
                    # Apply European number formatting to numeric columns
                    from openpyxl.styles import NamedStyle
                    from openpyxl.utils import get_column_letter
                    
                    # Identify numeric columns (those containing price, quantity, manhours, wage)
                    numeric_columns = []
                    for col_idx, col_name in enumerate(processor.master_dataset.columns, 1):
                        col_lower = col_name.lower()
                        if any(keyword in col_lower for keyword in ['price', 'quantity', 'manhours', 'wage']):
                            numeric_columns.append((col_idx, col_name))
                    
                    # Apply formatting to numeric columns
                    for col_idx, col_name in numeric_columns:
                        # Create European number style
                        euro_style = NamedStyle(name=f"euro_number_{col_name}")
                        euro_style.number_format = '#,##0.00'
                        
                        # Apply to the column
                        col_letter = get_column_letter(col_idx)
                        for row in range(2, len(processor.master_dataset) + 2):  # Skip header row
                            cell = worksheet[f'{col_letter}{row}']
                            cell.style = euro_style
                    
                    # Add data validation for Category column if it exists
                    if 'Category' in processor.master_dataset.columns:
                        category_col_idx = list(processor.master_dataset.columns).index('Category') + 1
                        category_col_letter = get_column_letter(category_col_idx)
                        
                        # Get available categories in the correct order (from category_order)
                        try:
                            category_order = ['General Costs', 'Site Costs', 'Civil Works', 'Earth Movement', 'Roads', 'OEM Building', 'Electrical Works', 'Solar Cables', 'LV Cables', 'MV Cables', 'Trenching', 'PV Mod. Installation', 'Cleaning and Cabling of PV Mod.', 'Tracker Inst.', 'Other']
                            available_categories = category_order
                            
                            # Create data validation for Category column
                            from openpyxl.worksheet.datavalidation import DataValidation
                            
                            # Create validation rule
                            dv = DataValidation(
                                type="list",
                                formula1=f'"{",".join(available_categories)}"',
                                allow_blank=True,
                                showErrorMessage=True,
                                errorTitle="Invalid Category",
                                error="Please select a category from the dropdown list.",
                                showInputMessage=True,
                                promptTitle="Category Selection",
                                prompt="Select a category from the dropdown list."
                            )
                            
                            # Add validation to worksheet
                            worksheet.add_data_validation(dv)
                            
                            # Apply validation to Category column (skip header row)
                            for row in range(2, len(processor.master_dataset) + 2):
                                dv.add(f'{category_col_letter}{row}')
                            
                        except Exception as e:
                            logger.warning(f"Could not add category validation: {e}")
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Format header row
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Add summary sheet with formulas
                    self._add_summary_sheet_with_formulas_comparison(workbook, processor.master_dataset, offer_info)
                
                messagebox.showinfo("Export Complete", f"Comparison results exported to {filename}")
                
        except Exception as e:
            logger.error(f"Error exporting comparison results: {e}")
            messagebox.showerror("Export Error", f"Failed to export results: {str(e)}")

    def _prompt_offer_info(self, is_first_boq=True):
        """
        Prompt user for offer information
        
        Args:
            is_first_boq: True if this is the first BoQ, False for comparison
            
        Returns:
            dict: Offer information or None if cancelled
        """
        try:
            if OFFER_INFO_AVAILABLE:
                # Get previous offer info for comparison BOQs
                previous_offer_info = None
                if not is_first_boq and hasattr(self, 'current_offer_info') and self.current_offer_info:
                    previous_offer_info = self.current_offer_info
                    logger.debug(f"Using previous offer info for comparison: {previous_offer_info}")
                
                result = show_offer_info_dialog(self.root, is_first_boq, previous_offer_info)
                logger.debug(f"Offer info dialog returned: {result}")
                return result
            else:
                # Fallback: return default info
                fallback_info = {
                    'offer_name': 'Default Offer' if is_first_boq else 'Comparison Offer',
                    'client_name': 'Default Client',
                    'project_name': 'Default Project',
                    'date': datetime.now().strftime('%Y-%m-%d')
                }
                logger.debug(f"Using fallback offer info: {fallback_info}")
                return fallback_info
        except Exception as e:
            logger.error(f"Error prompting for offer info: {e}")
            return None

    def _clear_all_files(self):
        """Clear all loaded files and reset the interface"""
        try:
            # Clear all tabs
            for tab in self.notebook.tabs():
                self.notebook.forget(tab)
            
            # Clear controller data
            if hasattr(self.controller, 'current_files'):
                self.controller.current_files.clear()
            
            # Reset comparison workflow state
            self.is_comparison_workflow = False
            self.master_file_mapping = None
            self.comparison_processor = None
            
            # Update status
            self._update_status("All files cleared")
            
            # Refresh summary grid
            self._refresh_summary_grid_centralized()
            
        except Exception as e:
            logger.error(f"Error clearing all files: {e}")
            messagebox.showerror("Error", f"Failed to clear files: {str(e)}")

    def _load_analysis(self):
        """Load a previously saved analysis file"""
        try:
            # Clear all previous data when loading analysis
            self._clear_all_files()
            
            filetypes = [("Pickle files", "*.pkl"), ("All files", "*.*")]
            filename = filedialog.askopenfilename(
                title="Load Analysis",
                filetypes=filetypes
            )
            
            if filename:
                with open(filename, 'rb') as f:
                    analysis_data = pickle.load(f)
                
                # Process the loaded analysis
                if 'dataframe' in analysis_data and 'categorization_result' in analysis_data:
                    # Create a new tab for the loaded analysis
                    tab = ttk.Frame(self.notebook)
                    self.notebook.add(tab, text=f"Loaded Analysis - {os.path.basename(filename)}")
                    self.notebook.select(tab)
                    
                    # Store the analysis data in controller for summary display
                    file_key = f"loaded_analysis_{int(time.time())}"
                    self.controller.current_files[file_key] = {
                        'file_mapping': type('MockFileMapping', (), {
                            'tab': tab,
                            'offer_info': analysis_data.get('offer_info', {
                                'offer_name': 'Loaded Analysis',
                                'project_name': 'Loaded Project',
                                'project_size': 'N/A',
                                'date': datetime.now().strftime('%Y-%m-%d')
                            })
                        })(),
                        'final_dataframe': analysis_data['dataframe'],
                        'offer_info': analysis_data.get('offer_info', {
                            'offer_name': 'Loaded Analysis',
                            'project_name': 'Loaded Project',
                            'project_size': 'N/A',
                            'date': datetime.now().strftime('%Y-%m-%d')
                        })
                    }
                    
                    # Show the final categorized data directly
                    final_dataframe = analysis_data['dataframe']
                    categorization_result = analysis_data['categorization_result']
                    
                    self._show_final_categorized_data(tab, final_dataframe, categorization_result)
                    
                    self._update_status(f"Analysis loaded from {filename}")
                else:
                    messagebox.showerror("Error", "Invalid analysis file format")
                
        except Exception as e:
            logger.error(f"Error loading analysis: {e}")
            messagebox.showerror("Error", f"Failed to load analysis: {str(e)}")

    def _use_mapping(self):
        """Load and apply a saved mapping to a new file"""
        try:
            # Clear all previous data when using mapping
            self._clear_all_files()
            
            filetypes = [("Pickle files", "*.pkl"), ("All files", "*.*")]
            filename = filedialog.askopenfilename(
                title="Load Mapping",
                filetypes=filetypes
            )
            
            if filename:
                with open(filename, 'rb') as f:
                    mapping_data = pickle.load(f)
                
                # Apply the mapping to a new file
                self._open_excel_file_with_mapping(mapping_data)
                
                self._update_status(f"Mapping loaded from {filename}")
                
        except Exception as e:
            logger.error(f"Error loading mapping: {e}")
            messagebox.showerror("Error", f"Failed to load mapping: {str(e)}")

    def _open_excel_file_with_mapping(self, mapping_data):
        """Open an Excel file and apply a saved mapping - streamlined workflow"""
        try:
            filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            filepath = filedialog.askopenfilename(
                title="Select Excel File to Apply Mapping",
                filetypes=filetypes
            )
            
            if filepath:
                # Prompt for offer info first
                offer_info = self._prompt_offer_info(is_first_boq=True)
                if not offer_info:
                    self._update_status("Mapping application cancelled (no offer information provided).")
                    return
                
                # Process file with mapping in a streamlined way
                self._process_file_with_mapping(filepath, mapping_data, offer_info)
                
        except Exception as e:
            logger.error(f"Error applying mapping: {e}")
            messagebox.showerror("Error", f"Failed to apply mapping: {str(e)}")

    def _process_file_with_mapping(self, filepath, mapping_data, offer_info):
        """Process file with pre-existing mapping - skip categorization and column mapping"""
        try:
            # Clear previous results
            self.sheet_treeviews.clear()
            if self.notebook:
                for tab in self.notebook.tabs():
                    self.notebook.forget(tab)

            self._update_status(f"Processing {os.path.basename(filepath)} with saved mapping...")
            self.progress_var.set(0)

            # Create a new tab for the file
            filename = os.path.basename(filepath)
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=filename)
            self.notebook.select(tab)

            # Configure grid for the tab frame
            tab.grid_rowconfigure(0, weight=1)
            tab.grid_columnconfigure(0, weight=1)

            loading_label = ttk.Label(tab, text="Processing with saved mapping...")
            loading_label.grid(row=0, column=0, pady=40, padx=100)
            self.root.update_idletasks()

            def process_in_thread():
                try:
                    # Process file with mapping
                    file_mapping = self.controller.process_file(
                        Path(filepath),
                        progress_callback=lambda p, m: self.root.after(0, self.update_progress, p, m)
                    )
                    
                    # Apply the saved mapping to the file_mapping
                    self._apply_saved_mapping(file_mapping, mapping_data)
                    
                    # Store the file mapping
                    self.file_mapping = file_mapping
                    self.column_mapper = file_mapping.column_mapper if hasattr(file_mapping, 'column_mapper') else None
                    
                    # Schedule completion on main thread
                    self.root.after(0, self._on_mapping_processing_complete, tab, filepath, file_mapping, loading_label, offer_info)
                    
                except Exception as e:
                    logger.error(f"Failed to process file with mapping: {e}", exc_info=True)
                    self.root.after(0, self._on_processing_error, tab, filename, loading_label)

            # Start processing thread
            threading.Thread(target=process_in_thread, daemon=True).start()
            
        except Exception as e:
            logger.error(f"Error in _process_file_with_mapping: {e}")
            messagebox.showerror("Error", f"Failed to process file with mapping: {str(e)}")

    def _apply_saved_mapping(self, file_mapping, mapping_data):
        """Apply saved mapping to file_mapping"""
        try:
            # Apply sheet categories for filtering
            if 'sheet_categories' in mapping_data:
                self.current_sheet_categories = mapping_data['sheet_categories']
                logger.debug(f"Applied sheet categories: {self.current_sheet_categories}")
                
                # Filter sheets based on saved categories
                boq_sheets = [sheet for sheet, cat in self.current_sheet_categories.items() if cat == "BOQ"]
                logger.debug(f"Filtered to BOQ sheets: {boq_sheets}")
                
                # Only keep BOQ sheets in the file mapping
                if boq_sheets:
                    file_mapping.sheets = [sheet for sheet in file_mapping.sheets if sheet.sheet_name in boq_sheets]
                    logger.debug(f"Filtered file_mapping to {len(file_mapping.sheets)} BOQ sheets")
            
            # Apply column mappings from saved mapping
            if 'column_mappings' in mapping_data:
                for sheet in file_mapping.sheets:
                    if sheet.sheet_name in mapping_data['column_mappings']:
                        sheet.column_mappings = mapping_data['column_mappings'][sheet.sheet_name]
                        logger.debug(f"Applied column mappings to sheet {sheet.sheet_name}")
            
            # Apply any other mapping data
            if 'header_row_index' in mapping_data:
                for sheet in file_mapping.sheets:
                    sheet.header_row_index = mapping_data['header_row_index']
                    
            logger.debug("Applied saved mapping to file_mapping")
            
        except Exception as e:
            logger.error(f"Error applying saved mapping: {e}")

    def _on_mapping_processing_complete(self, tab, filepath, file_mapping, loading_widget, offer_info):
        """Handle completion of file processing with mapping"""
        try:
            logger.info(f"Mapping processing complete for file: {filepath}")
            
            # Prevent setting file_mapping during comparison workflow
            if getattr(self, 'is_comparison_workflow', False):
                logger.info("Skipping file_mapping assignment during comparison workflow in mapping processing")
                loading_widget.destroy()
                return
            
            # Store the file mapping and column mapper
            self.file_mapping = file_mapping
            self.column_mapper = file_mapping.column_mapper if hasattr(file_mapping, 'column_mapper') else None
            
            # Store offer info
            self.current_offer_info = offer_info
            
            # Store offer info in the controller's current_files
            file_key = str(Path(filepath).resolve())
            if file_key in self.controller.current_files:
                offer_info_enhanced = {
                    'offer_name': offer_info.get('offer_name', 'Unknown'),
                    'project_name': offer_info.get('project_name', 'Unknown'),
                    'project_size': offer_info.get('project_size', 'N/A'),
                    'date': offer_info.get('date', datetime.now().strftime('%Y-%m-%d'))
                }
                
                offer_name = offer_info_enhanced['offer_name']
                if 'offers' not in self.controller.current_files[file_key]:
                    self.controller.current_files[file_key]['offers'] = {}
                self.controller.current_files[file_key]['offers'][offer_name] = offer_info_enhanced
                self.controller.current_files[file_key]['offer_info'] = offer_info_enhanced
                
                # Store offer info directly in file_mapping
                file_mapping.offer_info = offer_info_enhanced
                
                logger.debug(f"Stored offer info for mapping workflow: {offer_info_enhanced}")
            
            # Store file mapping in tab_id_to_file_mapping for row review
            current_tab_id = self.notebook.select()
            self.tab_id_to_file_mapping[current_tab_id] = file_mapping
            
            # Remove loading widget and go straight to row review
            loading_widget.destroy()
            self._show_row_review_with_mapping(tab, file_mapping)
            
            # Update status
            self._update_status(f"Mapping applied successfully: {os.path.basename(filepath)}")
            
        except Exception as e:
            logger.error(f"Error in mapping processing completion: {e}")
            messagebox.showerror("Error", f"Failed to complete mapping processing: {str(e)}")

    def _show_row_review_with_mapping(self, tab, file_mapping):
        """Show row review directly when using saved mapping"""
        try:
            # Prevent row review during comparison workflow
            if getattr(self, 'is_comparison_workflow', False):
                logger.info("Row review with mapping prevented during comparison workflow")
                return
            
            # Clear existing content
            for widget in tab.winfo_children():
                widget.destroy()
            
            # Create main frame
            main_frame = ttk.Frame(tab)
            main_frame.grid(row=0, column=0, sticky=tk.NSEW)
            main_frame.grid_rowconfigure(0, weight=1)
            main_frame.grid_columnconfigure(0, weight=1)
            
            # Title
            title_label = ttk.Label(main_frame, text="Row Review (Using Saved Mapping)", 
                                   font=("Arial", 14, "bold"))
            title_label.grid(row=0, column=0, pady=(0, 10), sticky=tk.W)
            
            # Show row review directly
            self._show_row_review(file_mapping)
            
        except Exception as e:
            logger.error(f"Error showing row review with mapping: {e}")
            messagebox.showerror("Error", f"Failed to show row review: {str(e)}")

    def _refresh_summary_grid_centralized(self):
        """Centralized method to refresh the summary grid"""
        try:
            # This method should refresh any summary or overview displays
            # For now, just log that it was called
            logger.debug("Summary grid refresh called")
            
            # Update status if needed
            self._update_status("Summary updated")
            
        except Exception as e:
            logger.error(f"Error refreshing summary grid: {e}")

    def _show_final_categorized_data(self, tab, final_dataframe, categorization_result):
        # Use the dataframe with comparison columns if available
        display_df = getattr(self, '_current_dataframe_with_comparison', final_dataframe)
        """Show the final categorized data in the tab"""
        try:
            # Validate tab parameter
            if not hasattr(tab, 'winfo_children'):
                logger.error(f"Invalid tab parameter: {type(tab)} - {tab}")
                messagebox.showerror("Error", f"Invalid tab parameter: {type(tab)}")
                return
            
            # Clear existing content
            for widget in tab.winfo_children():
                widget.destroy()
            
            # Create main frame
            main_frame = ttk.Frame(tab, padding="10")
            main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # Configure grid weights
            tab.columnconfigure(0, weight=1)
            tab.rowconfigure(0, weight=1)
            main_frame.columnconfigure(0, weight=1)
            main_frame.rowconfigure(1, weight=1)
            main_frame.rowconfigure(2, weight=0)  # For summary panel
            main_frame.rowconfigure(3, weight=0)  # For category summary (when shown)
            main_frame.rowconfigure(4, weight=0)  # For buttons
            
            # Title
            title_label = ttk.Label(main_frame, text="Final Categorized Data", 
                                   font=("Arial", 14, "bold"))
            title_label.grid(row=0, column=0, pady=(0, 10), sticky=tk.W)
            
            # Prepare data for display
            display_df = final_dataframe.copy()
            
            # Remove unwanted columns - expanded list of internal processing columns
            columns_to_remove = [
                'ignore', 'Position', 'Labour', 
                'ignore_14', 'ignore_15', '_3'
            ]
            
            # Also remove any columns that start with 'ignore_' or are just numbers
            for col in list(display_df.columns):
                if col.startswith('ignore_') or col.isdigit() or col in columns_to_remove:
                    display_df = display_df.drop(columns=[col])
            
            # Keep only essential columns for comparison results
            essential_columns = [
                'code', 'Category', 'Description', 'unit', 'quantity', 'unit_price', 'total_price',
                'manhours', 'wage'
            ]
            
            # Add any offer-specific columns (quantity[OfferName], etc.)
            offer_columns = [col for col in display_df.columns if '[' in col and ']' in col]
            
            # Final column order: essential columns first, then offer-specific columns
            final_columns = [col for col in essential_columns if col in display_df.columns]
            final_columns.extend(offer_columns)
            
            # Add any remaining columns that aren't offer-specific
            remaining_columns = [col for col in display_df.columns if col not in final_columns and col not in offer_columns]
            final_columns.extend(remaining_columns)
            
            display_df = display_df[final_columns]
            
            # Use the same column order as comparison window
            desired_order = ['Source_Sheet', 'code', 'Category', 'Description', 'scope', 'unit', 'quantity', 'unit_price', 'total_price', 'manhours', 'wage']
            
            # Build final column order - same logic as comparison window
            final_columns = []
            for col in desired_order:
                if col in display_df.columns:
                    final_columns.append(col)
            
            # Add any offer-specific columns (those with [offer_name] pattern)
            offer_columns = [col for col in display_df.columns if '[' in col and ']' in col]
            for col in offer_columns:
                if col not in final_columns:
                    final_columns.append(col)
            
            # Add any remaining columns
            for col in display_df.columns:
                if col not in final_columns:
                    final_columns.append(col)
            
            # Source_Sheet should now be present from the initial data processing
            if 'Source_Sheet' not in final_columns:
                logger.warning("Source_Sheet column is still missing - this indicates a processing issue")
            
            display_df = display_df[final_columns]
            
            # Create treeview
            columns = list(display_df.columns)
            tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=15)
            
            # Apply light blue selection style to the treeview (same as column mapping)
            style = ttk.Style(tree)
            style.map('Treeview', 
                     background=[('selected', '#B3E5FC')])  # Light blue background
            
            # Configure columns with appropriate widths
            column_widths = {
                'Description': 400,  # Wider for descriptions
                'code': 100,
                'unit': 80,
                'quantity': 100,
                'unit_price': 120,
                'total_price': 120,
                'Category': 150,
                'manhours': 100,
                'wage': 80
            }
            
            # Set default width for offer-specific columns
            default_offer_width = 120
            
            for col in columns:
                # Display pretty names in headers
                pretty_name = self._get_pretty_column_name(col)
                tree.heading(col, text=pretty_name)
                # Use specific width if defined, otherwise use default for offer columns
                if col in column_widths:
                    width = column_widths[col]
                elif '[' in col and ']' in col:
                    width = default_offer_width
                else:
                    width = 100
                tree.column(col, width=width, minwidth=80)
            
            # Add scrollbars
            vsb = ttk.Scrollbar(main_frame, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(main_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Grid layout with proper scrollbar configuration
            tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
            vsb.grid(row=1, column=1, sticky=(tk.N, tk.S))
            hsb.grid(row=2, column=0, sticky=(tk.W, tk.E))
            
            # Configure grid weights to ensure scrollbars work properly
            main_frame.columnconfigure(0, weight=1)
            main_frame.rowconfigure(1, weight=1)
            
            # Create summary panel (treeview with offers)
            summary_frame = ttk.LabelFrame(main_frame, text="Summary", padding="10")
            summary_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)
            
            # Populate with data using exact same formatting as comparison window
            for i, row in display_df.iterrows():
                formatted_values = []
                for col in columns:
                    val = row[col]
                    
                    # Apply exact same formatting as comparison window
                    if pd.notna(val) and val != '':
                        # Check if this is a comparison column (has [offer_name] pattern)
                        is_comparison_col = '[' in col and ']' in col
                        
                        # Extract base column name for comparison columns
                        base_col = col.split('[')[0] if is_comparison_col else col
                        
                        if base_col in ['unit_price', 'total_price', 'wage']:
                            # Currency formatting (same as comparison window)
                            val = format_number_eu(val)
                        elif base_col in ['quantity', 'manhours']:
                            # Number formatting (same as comparison window)
                            val = format_number_eu(val)
                        else:
                            val = str(val)
                    else:
                        val = ''
                    
                    formatted_values.append(val)
                
                tree.insert('', 'end', values=formatted_values)
            
            # Configure summary frame
            summary_frame.columnconfigure(0, weight=1)
            summary_frame.rowconfigure(0, weight=1)
            
            # Create treeview for summary
            summary_columns = ['Offer Name', 'Project Name', 'Project Size', 'Date', 'Total Price']
            summary_tree = ttk.Treeview(summary_frame, columns=summary_columns, show='headings', height=3)
            
            # Apply light blue selection style to the summary treeview (same as column mapping)
            summary_style = ttk.Style(summary_tree)
            summary_style.map('Treeview', 
                             background=[('selected', '#B3E5FC')])  # Light blue background
            
            # Configure columns
            for col in summary_columns:
                summary_tree.heading(col, text=col)
                summary_tree.column(col, width=120, minwidth=80)
            
            # Add scrollbars for summary
            summary_vsb = ttk.Scrollbar(summary_frame, orient="vertical", command=summary_tree.yview)
            summary_hsb = ttk.Scrollbar(summary_frame, orient="horizontal", command=summary_tree.xview)
            summary_tree.configure(yscrollcommand=summary_vsb.set, xscrollcommand=summary_hsb.set)
            
            # Grid layout for summary
            summary_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
            summary_vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
            summary_hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
            
            # Calculate summary data
            total_cost = 0.0
            if 'total_price' in display_df.columns:
                try:
                    numeric_prices = pd.to_numeric(display_df['total_price'], errors='coerce')
                    total_cost = numeric_prices.sum()
                except Exception as e:
                    logger.warning(f"Error calculating total cost: {e}")
                    total_cost = 0.0
            
            # Format total cost with European formatting (no currency symbol, dot for thousands, comma for decimals)
            if total_cost > 0:
                formatted_total_cost = format_number_eu(total_cost)
            else:
                formatted_total_cost = "0,00"
            
            # Get offer information from file mapping or use defaults
            offer_name = "Current Offer"
            project_name = "Current Project"
            project_size = "N/A"
            date = datetime.now().strftime('%Y-%m-%d')
            
            # Try to get actual offer info from file mapping
            current_tab_path = self.notebook.select()
            logger.debug(f"Looking for offer info for tab path: {current_tab_path}")
            
            # First, try to get offer info directly from the current file_mapping if available
            if hasattr(self, 'current_file_mapping') and self.current_file_mapping and hasattr(self.current_file_mapping, 'offer_info'):
                offer_info = self.current_file_mapping.offer_info
                offer_name = offer_info.get('offer_name', offer_name)
                project_name = offer_info.get('project_name', project_name)
                project_size = offer_info.get('project_size', project_size)
                date = offer_info.get('date', date)
                logger.debug(f"Found offer info directly from current_file_mapping: {offer_info}")
            elif hasattr(self, 'file_mapping') and self.file_mapping and hasattr(self.file_mapping, 'offer_info'):
                offer_info = self.file_mapping.offer_info
                offer_name = offer_info.get('offer_name', offer_name)
                project_name = offer_info.get('project_name', project_name)
                project_size = offer_info.get('project_size', project_size)
                date = offer_info.get('date', date)
                logger.debug(f"Found offer info directly from file_mapping: {offer_info}")
            elif hasattr(self, 'current_offer_info') and self.current_offer_info:
                # For Use Mapping workflow, check current_offer_info
                offer_info = self.current_offer_info
                offer_name = offer_info.get('offer_name', offer_name)
                project_name = offer_info.get('project_name', project_name)
                project_size = offer_info.get('project_size', project_size)
                date = offer_info.get('date', date)
                logger.debug(f"Found offer info from current_offer_info: {offer_info}")
            else:
                # Fallback: search through controller's current_files
                logger.debug("No direct file_mapping offer_info, searching through controller files")
                for file_key, file_data in self.controller.current_files.items():
                    logger.debug(f"Checking file_key: {file_key}")
                    logger.debug(f"File data keys: {list(file_data.keys())}")
                    
                    if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                        logger.debug(f"Found matching file data for tab path")
                        # Try to get offer info from file mapping
                        if hasattr(file_data['file_mapping'], 'offer_info') and file_data['file_mapping'].offer_info:
                            offer_info = file_data['file_mapping'].offer_info
                            offer_name = offer_info.get('offer_name', offer_name)
                            project_name = offer_info.get('project_name', project_name)
                            project_size = offer_info.get('project_size', project_size)
                            date = offer_info.get('date', date)
                            logger.debug(f"Found offer info in file_mapping: {offer_info}")
                        # Also check in the controller's current_files
                        elif 'offer_info' in file_data:
                            offer_info = file_data['offer_info']
                            offer_name = offer_info.get('offer_name', offer_name)
                            project_name = offer_info.get('project_name', project_name)
                            project_size = offer_info.get('project_size', project_size)
                            date = offer_info.get('date', date)
                            logger.debug(f"Found offer info in current_files: {offer_info}")
                        else:
                            logger.debug("No offer info found in either location")
                        break
                else:
                    logger.debug("No matching file data found for current tab")
            
            # Clear existing summary data
            for item in summary_tree.get_children():
                summary_tree.delete(item)
            
            # Collect all offers and their total costs
            offers_data = []
            
            # Add current offer
            current_offer_data = {
                'offer_name': offer_name,
                'project_name': project_name,
                'project_size': project_size,
                'date': date,
                'total_cost': total_cost,
                'formatted_total_cost': formatted_total_cost
            }
            offers_data.append(current_offer_data)
            
            # Check for comparison offers in the current dataframe (regardless of workflow flag)
            comparison_columns = [col for col in display_df.columns if '[' in col and ']' in col and 'total_price' in col]
            logger.info(f"DEBUG: display_df columns: {list(display_df.columns)}")
            logger.info(f"DEBUG: comparison_columns found: {comparison_columns}")
            
            if comparison_columns:
                # Look for comparison offers in the current dataframe
                
                for col in comparison_columns:
                    try:
                        # Extract offer name from column name (e.g., "total_price[OfferName]" -> "OfferName")
                        offer_name_from_col = col.split('[')[1].split(']')[0]
                        
                        # Calculate total cost for this offer
                        offer_total_cost = 0.0
                        if col in display_df.columns:
                            numeric_prices = pd.to_numeric(display_df[col], errors='coerce')
                            offer_total_cost = numeric_prices.sum()
                        
                        # Format total cost
                        if offer_total_cost > 0:
                            formatted_offer_cost = format_number_eu(offer_total_cost)
                        else:
                            formatted_offer_cost = "0,00"
                        
                        # Add comparison offer data - use the user's entered project info
                        # Get the comparison offer info from the controller's current_files
                        comparison_offer_info = None
                        for file_key, file_data in self.controller.current_files.items():
                            if 'offers' in file_data and offer_name_from_col in file_data['offers']:
                                comparison_offer_info = file_data['offers'][offer_name_from_col]
                                break
                        
                        # Use user's entered project info if available, otherwise use defaults
                        if comparison_offer_info:
                            project_name = comparison_offer_info.get('project_name', f"Project {offer_name_from_col}")
                            project_size = comparison_offer_info.get('project_size', 'N/A')
                            date = comparison_offer_info.get('date', date)
                        else:
                            project_name = f"Project {offer_name_from_col}"
                            project_size = "N/A"
                        
                        comparison_offer_data = {
                            'offer_name': offer_name_from_col,
                            'project_name': project_name,
                            'project_size': project_size,
                            'date': date,
                            'total_cost': offer_total_cost,
                            'formatted_total_cost': formatted_offer_cost
                        }
                        offers_data.append(comparison_offer_data)
                        
                    except Exception as e:
                        logger.warning(f"Error processing comparison offer column {col}: {e}")
            
            # Sort offers by total cost (lowest first)
            offers_data.sort(key=lambda x: x['total_cost'])
            
            # Insert all offers into summary treeview
            for offer_data in offers_data:
                summary_tree.insert('', 'end', values=(
                    offer_data['offer_name'],
                    offer_data['project_name'],
                    offer_data['project_size'],
                    offer_data['date'],
                    offer_data['formatted_total_cost']
                ))
            
            # Configure grid weights for summary frame
            summary_frame.columnconfigure(0, weight=1)
            summary_frame.rowconfigure(0, weight=1)
            
            # Create button frame (centered)
            button_frame = ttk.Frame(main_frame)
            button_frame.grid(row=5, column=0, sticky=(tk.W, tk.E), padx=5, pady=10)
            
            # Create inner frame for centering buttons
            inner_button_frame = ttk.Frame(button_frame)
            inner_button_frame.pack(expand=True)
            
            # Action buttons (centered)
            summarize_btn = ttk.Button(inner_button_frame, text="Summarize", 
                                     command=lambda: self._summarize_with_uncategorized_check(final_dataframe))
            summarize_btn.pack(side=tk.LEFT, padx=5)
            
            save_mapping_btn = ttk.Button(inner_button_frame, text="Save Mapping", 
                                        command=lambda: self._save_mapping_for_categorized_data(final_dataframe))
            save_mapping_btn.pack(side=tk.LEFT, padx=5)
            
            save_analysis_btn = ttk.Button(inner_button_frame, text="Save Analysis", 
                                         command=lambda: self._save_analysis_for_categorized_data(final_dataframe, categorization_result))
            save_analysis_btn.pack(side=tk.LEFT, padx=5)
            
            compare_full_btn = ttk.Button(inner_button_frame, text="Compare Full", 
                                        command=lambda: self._compare_full_from_categorized_data(tab))
            compare_full_btn.pack(side=tk.LEFT, padx=5)
            
            export_btn = ttk.Button(inner_button_frame, text="Export Data", 
                                   command=lambda: self._export_categorized_data(final_dataframe))
            export_btn.pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            logger.error(f"Error showing final categorized data: {e}")
            messagebox.showerror("Error", f"Failed to show categorized data: {str(e)}")
    
    def _summarize_with_uncategorized_check(self, dataframe):
        """
        Phase 2.3: Check for uncategorized rows before summarizing and warn user
        
        Args:
            dataframe: DataFrame to summarize
        """
        try:
            # Check if Category column exists
            if 'Category' not in dataframe.columns:
                logger.warning("No Category column found in DataFrame")
                # Proceed with summarize anyway
                self._summarize_categorized_data(dataframe)
                return
            
            # Identify uncategorized rows
            uncategorized_mask = (
                dataframe['Category'].isna() | 
                (dataframe['Category'] == '') | 
                (dataframe['Category'].astype(str).str.strip() == '')
            )
            uncategorized_count = uncategorized_mask.sum()
            
            if uncategorized_count > 0:
                # Show warning dialog
                response = messagebox.askyesno(
                    "Uncategorized Rows Detected",
                    f"Some rows ({uncategorized_count}) are not categorized. These will be excluded from category totals in Summarize.\n\n"
                    "Would you like to categorize them now?",
                    icon='question'
                )
                
                if response:
                    # User wants to categorize uncategorized rows
                    uncategorized_rows = dataframe[uncategorized_mask].copy()
                    
                    # Get file_mapping for context
                    current_tab_id = self.notebook.select()
                    file_mapping = self.tab_id_to_file_mapping.get(current_tab_id)
                    
                    if file_mapping:
                        # Store original DataFrame for merging
                        file_mapping._original_full_dataframe = dataframe.copy()
                        file_mapping._is_subset_categorization = True
                        
                        # Start categorization
                        self._update_status(f"Categorizing {len(uncategorized_rows)} uncategorized rows...")
                        self._start_categorization(uncategorized_rows, file_mapping)
                        return
                    else:
                        messagebox.showwarning(
                            "Warning",
                            "Could not find file mapping. Please categorize rows manually or proceed with summarize."
                        )
                
                # User chose to proceed anyway - show info and continue
                messagebox.showinfo(
                    "Proceeding with Summarize",
                    f"Summarize will proceed. {uncategorized_count} uncategorized rows will be excluded from category totals."
                )
            
            # Proceed with summarize (either no uncategorized rows or user chose to proceed)
            self._summarize_categorized_data(dataframe)
            
        except Exception as e:
            logger.error(f"Error checking for uncategorized rows: {e}")
            # On error, proceed with summarize anyway
            self._summarize_categorized_data(dataframe)

    def _debug_export_datasets_before_merge(self, master_df, comparison_df, offer_info):
        """DEBUG: Export both datasets to Excel for debugging merge process"""
        try:
            import pandas as pd
            from datetime import datetime
            import os
            import logging
            logger = logging.getLogger(__name__)
            
            # DEBUG: Log original columns
            logger.info(f"Master DataFrame original columns: {list(master_df.columns)}")
            logger.info(f"Comparison DataFrame original columns: {list(comparison_df.columns)}")
            
            # Get offer name
            offer_name = offer_info.get('offer_name', 'Comparison')
            
            # Create filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"DEBUG_Datasets_Before_Merge_{offer_name}_{timestamp}.xlsx"
            
            # Get the directory of the current working directory
            export_dir = os.getcwd()
            filepath = os.path.join(export_dir, filename)
            
            # Function to determine MERGE/ADD decision for each comparison row
            def determine_merge_add_decision(comparison_df, master_df):
                """Determine whether each comparison row would be MERGE or ADD"""
                decisions = []
                
                # Create a mapping of description to instance counts for correct ordering
                description_instance_counts = {}
                
                for idx, comp_row in comparison_df.iterrows():
                    description = str(comp_row.get('Description', '')).strip()
                    
                    if not description:
                        decisions.append('INVALID - No Description')
                        continue
                    
                    # Get master instances using normalized whitespace matching first
                    # This handles cases where descriptions have different whitespace/newline characters
                    normalized_desc = ' '.join(description.split())  # Normalize whitespace
                    master_instances = master_df[
                        master_df['Description'].str.lower().apply(lambda x: ' '.join(str(x).split())) == normalized_desc.lower()
                    ]
                    
                    # If no matches found with normalized matching, try exact matching as fallback
                    if len(master_instances) == 0:
                        master_instances = master_df[master_df['Description'] == description]
                        
                        # If still no matches, try case-insensitive matching
                        if len(master_instances) == 0:
                            master_instances = master_df[
                                master_df['Description'].str.lower() == description.lower()
                            ]
                    
                    # Initialize instance count for this description if not seen before
                    if description not in description_instance_counts:
                        description_instance_counts[description] = 0
                    
                    # Get the current instance number for this description
                    comp_instance_number = description_instance_counts[description]
                    description_instance_counts[description] += 1
                    
                    # Decision logic: if instance number < master instances, MERGE; else ADD
                    if comp_instance_number < len(master_instances):
                        master_idx = master_instances.index[comp_instance_number]
                        decisions.append(f'MERGE Instance {comp_instance_number + 1} (Master Row {master_idx + 2})')
                    else:
                        decisions.append(f'ADD Instance {comp_instance_number + 1} (New Row)')
                
                return decisions
            
            # Prepare datasets with proper formatting (same as main export)
            def prepare_dataset_for_export(df):
                """Prepare dataset with same formatting as main export"""
                export_df = df.copy()
                
                # DEBUG: Log columns before filtering
                logger.info(f"Columns before filtering: {list(export_df.columns)}")
                
                # Remove unwanted columns - comprehensive list of internal processing columns
                # Note: Keep Source_Sheet for debug export to show which sheet each row comes from
                columns_to_remove = [
                    'ignore', 'Position', 'Labour', 
                    'ignore_14', 'ignore_15', '_3', 'scope'
                ]
                
                # Also remove any columns that start with 'ignore_' or are just numbers
                # AND remove any columns that are just numbers (like _1, _2, _3, etc.)
                for col in list(export_df.columns):
                    if (col.startswith('ignore_') or 
                        col.isdigit() or 
                        col in columns_to_remove or
                        (col.startswith('_') and col[1:].isdigit())):  # Remove _1, _2, _3, etc.
                        export_df = export_df.drop(columns=[col])
                
                # DEBUG: Log columns after removing unwanted columns
                logger.info(f"Columns after removing unwanted: {list(export_df.columns)}")
                
                # Define the correct essential column order (same as main export)
                # Include Source_Sheet for debug export to show which sheet each row comes from
                essential_columns = [
                    'code', 'Category', 'Description', 'unit', 'quantity', 'unit_price', 'total_price',
                    'manhours', 'wage', 'Source_Sheet'
                ]
                
                # Keep only essential columns that exist in the dataframe
                available_essential_columns = [col for col in essential_columns if col in export_df.columns]
                
                # Create final column list: ONLY essential columns, no remaining columns
                final_columns = available_essential_columns.copy()
                
                # DEBUG: Log final columns
                logger.info(f"Final columns: {final_columns}")
                logger.info(f"Remaining non-essential columns: {[col for col in export_df.columns if col not in essential_columns]}")
                
                # Apply the column order
                export_df = export_df[final_columns]
                
                # Convert numeric columns to proper numeric values for Excel
                numeric_columns = ['quantity', 'unit_price', 'total_price', 'manhours', 'wage']
                for col in numeric_columns:
                    if col in export_df.columns:
                        # Convert to numeric, handling any formatting
                        export_df[col] = pd.to_numeric(export_df[col], errors='coerce')
                
                return export_df
            
            # Prepare both datasets
            master_export_df = prepare_dataset_for_export(master_df)
            comparison_export_df = prepare_dataset_for_export(comparison_df)
            
            # Add MERGE/ADD decision column to comparison dataset
            merge_add_decisions = determine_merge_add_decision(comparison_export_df, master_export_df)
            comparison_export_df['MERGE_ADD_Decision'] = merge_add_decisions
            
            # Create Excel writer with proper formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Export master dataset
                master_export_df.to_excel(writer, sheet_name='Master_Dataset', index=False)
                
                # Export comparison dataset
                comparison_export_df.to_excel(writer, sheet_name='Comparison_Dataset', index=False)
                
                # Get the workbook
                workbook = writer.book
                
                # Apply formatting to both sheets
                for sheet_name in ['Master_Dataset', 'Comparison_Dataset']:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Apply European number formatting to numeric columns
                    numeric_columns = ['quantity', 'unit_price', 'total_price', 'manhours', 'wage']
                    for col_idx, col_name in enumerate(worksheet[1], 1):
                        if col_name.value in numeric_columns:
                            # Apply European number format (dot for thousands, comma for decimals)
                            from openpyxl.styles import NamedStyle
                            from openpyxl.utils import get_column_letter
                            
                            # Create European number style
                            euro_style = NamedStyle(name=f"euro_number_{col_name.value}")
                            euro_style.number_format = '#,##0.00'
                            
                            # Apply to the column
                            col_letter = get_column_letter(col_idx)
                            for row in range(2, worksheet.max_row + 1):  # Skip header row
                                cell = worksheet[f'{col_letter}{row}']
                                cell.style = euro_style
                    
                    # Add data validation for Category column
                    if 'Category' in [cell.value for cell in worksheet[1]]:
                        category_col_idx = None
                        for col_idx, cell in enumerate(worksheet[1], 1):
                            if cell.value == 'Category':
                                category_col_idx = col_idx
                                break
                        
                        if category_col_idx:
                            category_col_letter = get_column_letter(category_col_idx)
                            
                            # Get available categories in the correct order
                            try:
                                category_order = ['General Costs', 'Site Costs', 'Civil Works', 'Earth Movement', 'Roads', 'OEM Building', 'Electrical Works', 'Solar Cables', 'LV Cables', 'MV Cables', 'Trenching', 'PV Mod. Installation', 'Cleaning and Cabling of PV Mod.', 'Tracker Inst.', 'Other']
                                available_categories = category_order
                                
                                # Create data validation for Category column
                                from openpyxl.worksheet.datavalidation import DataValidation
                                
                                # Create validation rule
                                dv = DataValidation(
                                    type="list",
                                    formula1=f'"{",".join(available_categories)}"',
                                    allow_blank=True,
                                    showErrorMessage=True,
                                    errorTitle="Invalid Category",
                                    error="Please select a category from the dropdown list.",
                                    showInputMessage=True,
                                    promptTitle="Category Selection",
                                    prompt="Select a category from the dropdown list."
                                )
                                
                                # Add validation to worksheet
                                worksheet.add_data_validation(dv)
                                
                                # Apply validation to Category column (skip header row)
                                for row in range(2, worksheet.max_row + 1):
                                    dv.add(f'{category_col_letter}{row}')
                                
                            except Exception as e:
                                logger.warning(f"Could not add category validation: {e}")
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Format header row
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Log the export
            logger.info(f"DEBUG: Exported datasets to {filepath}")
            logger.info(f"DEBUG: Master dataset shape: {master_export_df.shape}")
            logger.info(f"DEBUG: Comparison dataset shape: {comparison_export_df.shape}")
            
            # Show success message
            messagebox.showinfo("Debug Export", f"Datasets exported for debugging:\n{filepath}")
            
            # Note: File will be opened by the calling function, so we don't open it here
                
        except Exception as e:
            logger.error(f"Error in debug export: {e}")
            messagebox.showerror("Debug Export Error", f"Failed to export debug data: {str(e)}")

    def _export_categorized_data(self, final_dataframe):
        """Export categorized data to Excel with proper formatting and data validation"""
        try:
            filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
            filename = filedialog.asksaveasfilename(
                title="Export Categorized Data",
                filetypes=filetypes,
                defaultextension=".xlsx"
            )
            
            if filename:
                # Prepare data for export (same as display)
                export_df = final_dataframe.copy()
                
                # Remove unwanted columns
                columns_to_remove = ['ignore', 'Position']
                for col in columns_to_remove:
                    if col in export_df.columns:
                        export_df = export_df.drop(columns=[col])
                
                # Define desired column order
                desired_order = ['code', 'Category', 'Description', 'unit', 
                               'quantity', 'unit_price', 'total_price', 'manhours', 'wage']
                
                # Reorder columns (only include columns that exist)
                existing_columns = [col for col in desired_order if col in export_df.columns]
                other_columns = [col for col in export_df.columns if col not in desired_order]
                final_columns = existing_columns + other_columns
                
                export_df = export_df[final_columns]
                
                # Convert numeric columns to proper numeric values for Excel
                numeric_columns = ['quantity', 'unit_price', 'total_price', 'manhours', 'wage']
                for col in numeric_columns:
                    if col in export_df.columns:
                        # Convert to numeric, handling any formatting
                        export_df[col] = pd.to_numeric(export_df[col], errors='coerce')
                
                # Export with proper Excel formatting and data validation
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='BOQ Data')
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['BOQ Data']
                    
                    # Apply European number formatting to numeric columns
                    for col_idx, col_name in enumerate(export_df.columns, 1):
                        if col_name in numeric_columns:
                            # Apply European number format (dot for thousands, comma for decimals)
                            from openpyxl.styles import NamedStyle
                            from openpyxl.utils import get_column_letter
                            
                            # Create European number style
                            euro_style = NamedStyle(name=f"euro_number_{col_name}")
                            euro_style.number_format = '#,##0.00'
                            
                            # Apply to the column
                            col_letter = get_column_letter(col_idx)
                            for row in range(2, len(export_df) + 2):  # Skip header row
                                cell = worksheet[f'{col_letter}{row}']
                                cell.style = euro_style
                    
                    # Add data validation for Category column
                    if 'Category' in export_df.columns:
                        category_col_idx = list(export_df.columns).index('Category') + 1
                        category_col_letter = get_column_letter(category_col_idx)
                        
                        # Get available categories in the correct order (from category_order)
                        try:
                            category_order = ['General Costs', 'Site Costs', 'Civil Works', 'Earth Movement', 'Roads', 'OEM Building', 'Electrical Works', 'Solar Cables', 'LV Cables', 'MV Cables', 'Trenching', 'PV Mod. Installation', 'Cleaning and Cabling of PV Mod.', 'Tracker Inst.', 'Other']
                            available_categories = category_order
                            
                            # Create data validation for Category column
                            from openpyxl.worksheet.datavalidation import DataValidation
                            
                            # Create validation rule
                            dv = DataValidation(
                                type="list",
                                formula1=f'"{",".join(available_categories)}"',
                                allow_blank=True,
                                showErrorMessage=True,
                                errorTitle="Invalid Category",
                                error="Please select a category from the dropdown list.",
                                showInputMessage=True,
                                promptTitle="Category Selection",
                                prompt="Select a category from the dropdown list."
                            )
                            
                            # Add validation to worksheet
                            worksheet.add_data_validation(dv)
                            
                            # Apply validation to Category column (skip header row)
                            for row in range(2, len(export_df) + 2):
                                dv.add(f'{category_col_letter}{row}')
                            
                        except Exception as e:
                            logger.warning(f"Could not add category validation: {e}")
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Format header row
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Get offer name for summary sheet
                    offer_name = "Current Offer"
                    current_tab_path = self.notebook.select()
                    for file_key, file_data in self.controller.current_files.items():
                        if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                            if hasattr(file_data['file_mapping'], 'offer_info') and file_data['file_mapping'].offer_info:
                                offer_info = file_data['file_mapping'].offer_info
                                offer_name = offer_info.get('offer_name', offer_name)
                            elif 'offer_info' in file_data:
                                offer_info = file_data['offer_info']
                                offer_name = offer_info.get('offer_name', offer_name)
                            break
                    
                    # Add summary sheet with formulas
                    self._add_summary_sheet_with_formulas(workbook, export_df, offer_name)
                
                messagebox.showinfo("Export Complete", f"Categorized data exported to {filename}")
                
        except Exception as e:
            logger.error(f"Error exporting categorized data: {e}")
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")

    def _add_summary_sheet_with_formulas(self, workbook, dataframe, offer_name):
        """Add a summary sheet with formulas that calculate totals for each category"""
        try:
            logger.info(f"Creating summary sheet for offer: {offer_name}")
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            
            # Define category order (same as in _summarize_categorized_data)
            category_order = [
                "General Costs",
                "Site Costs", 
                "Civil Works",
                "Earth Movement",
                "Roads",
                "OEM Building",
                "Electrical Works",
                "Solar Cables",
                "LV Cables", 
                "MV Cables",
                "Trenching",
                "PV Mod. Installation",
                "Cleaning and Cabling of PV Mod.",
                "Tracker Inst.",
                "Other"
            ]
            
            # Create headers
            headers = ['Offer Name'] + category_order
            for col_idx, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            
            # Add offer name
            summary_sheet.cell(row=2, column=1, value=offer_name)
            
            # Add formulas for each category
            for col_idx, category in enumerate(category_order, 2):  # Start from column 2 (after Offer Name)
                # Find the Category column in the main data sheet
                category_col_idx = None
                for idx, col_name in enumerate(dataframe.columns, 1):
                    if col_name == 'Category':
                        category_col_idx = idx
                        break
                
                # Find the total_price column in the main data sheet
                total_price_col_idx = None
                for idx, col_name in enumerate(dataframe.columns, 1):
                    if col_name == 'total_price':
                        total_price_col_idx = idx
                        break
                
                if category_col_idx and total_price_col_idx:
                    # Create SUMIFS formula: =SUMIFS('BOQ Data'!F:F,'BOQ Data'!C:C,"General Costs")
                    # Where F is total_price column and C is Category column
                    category_col_letter = openpyxl.utils.get_column_letter(category_col_idx)
                    total_price_col_letter = openpyxl.utils.get_column_letter(total_price_col_idx)
                    
                    formula = f'=SUMIFS(\'BOQ Data\'!{total_price_col_letter}:{total_price_col_letter},\'BOQ Data\'!{category_col_letter}:{category_col_letter},"{category}")'
                    
                    cell = summary_sheet.cell(row=2, column=col_idx)
                    cell.value = formula
                    
                    # Apply European number formatting
                    cell.number_format = '#,##0.00'
                else:
                    # If columns not found, set to 0
                    summary_sheet.cell(row=2, column=col_idx, value=0)
                    summary_sheet.cell(row=2, column=col_idx).number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error adding summary sheet with formulas: {e}")
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
            # Don't raise the exception, just log it so the main export can continue

    def _add_summary_sheet_with_formulas_comparison(self, workbook, dataframe, offer_info):
        """Add a summary sheet with formulas for comparison results"""
        try:
            offer_name = offer_info.get('offer_name', 'Comparison Results') if offer_info else 'Comparison Results'
            logger.info(f"Creating summary sheet for comparison: {offer_name}")
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            
            # Define category order (same as in _summarize_categorized_data)
            category_order = [
                "General Costs",
                "Site Costs", 
                "Civil Works",
                "Earth Movement",
                "Roads",
                "OEM Building",
                "Electrical Works",
                "Solar Cables",
                "LV Cables", 
                "MV Cables",
                "Trenching",
                "PV Mod. Installation",
                "Cleaning and Cabling of PV Mod.",
                "Tracker Inst.",
                "Other"
            ]
            
            # Create headers
            headers = ['Offer Name'] + category_order
            for col_idx, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            
            # Add offer name
            offer_name = offer_info.get('offer_name', 'Comparison Results') if offer_info else 'Comparison Results'
            summary_sheet.cell(row=2, column=1, value=offer_name)
            
            # Add formulas for each category
            for col_idx, category in enumerate(category_order, 2):  # Start from column 2 (after Offer Name)
                # Find the Category column in the main data sheet
                category_col_idx = None
                for idx, col_name in enumerate(dataframe.columns, 1):
                    if col_name == 'Category':
                        category_col_idx = idx
                        break
                
                # Find the total_price column in the main data sheet
                total_price_col_idx = None
                for idx, col_name in enumerate(dataframe.columns, 1):
                    if col_name == 'total_price':
                        total_price_col_idx = idx
                        break
                
                if category_col_idx and total_price_col_idx:
                    # Create SUMIFS formula: =SUMIFS('Comparison Results'!F:F,'Comparison Results'!C:C,"General Costs")
                    # Where F is total_price column and C is Category column
                    category_col_letter = openpyxl.utils.get_column_letter(category_col_idx)
                    total_price_col_letter = openpyxl.utils.get_column_letter(total_price_col_idx)
                    
                    formula = f'=SUMIFS(\'Comparison Results\'!{total_price_col_letter}:{total_price_col_letter},\'Comparison Results\'!{category_col_letter}:{category_col_letter},"{category}")'
                    
                    cell = summary_sheet.cell(row=2, column=col_idx)
                    cell.value = formula
                    
                    # Apply European number formatting
                    cell.number_format = '#,##0.00'
                else:
                    # If columns not found, set to 0
                    summary_sheet.cell(row=2, column=col_idx, value=0)
                    summary_sheet.cell(row=2, column=col_idx).number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error adding summary sheet with formulas for comparison: {e}")
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
            # Don't raise the exception, just log it so the main export can continue

    def _summarize_categorized_data(self, dataframe):
        # Use the dataframe with comparison columns if available
        display_dataframe = getattr(self, '_current_dataframe_with_comparison', dataframe)
        """Show detailed summary of categorized data in a new section below the summary"""
        try:
            # Get the current tab
            current_tab_path = self.notebook.select()
            current_tab = self.notebook.nametowidget(self.notebook.select())
            
            # Find the main frame in the current tab
            main_frame = None
            for widget in current_tab.winfo_children():
                if isinstance(widget, ttk.Frame):
                    main_frame = widget
                    break
            
            if not main_frame:
                logger.error("Could not find main frame")
                return
            
            # Clear existing category summary if it exists
            for widget in main_frame.winfo_children():
                if hasattr(widget, 'category_summary'):
                    widget.destroy()
            
            # Create category summary frame
            category_frame = ttk.LabelFrame(main_frame, text="Category Summary", padding="5")
            category_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)
            category_frame.columnconfigure(0, weight=1)
            category_frame.rowconfigure(0, weight=1)
            
            # Define category order (from manual_categorizer.py)
            category_order = [
                "General Costs",
                "Site Costs", 
                "Civil Works",
                "Earth Movement",
                "Roads",
                "OEM Building",
                "Electrical Works",
                "Solar Cables",
                "LV Cables", 
                "MV Cables",
                "Trenching",
                "PV Mod. Installation",
                "Cleaning and Cabling of PV Mod.",
                "Tracker Inst.",
                "Other"
            ]
            
            # Create treeview for categories
            columns = ['Offer Name'] + category_order
            category_tree = ttk.Treeview(category_frame, columns=columns, show='headings', height=5)
            category_tree.category_summary = True  # Mark for identification
            
            # Configure columns
            category_tree.heading('Offer Name', text='Offer Name')
            category_tree.column('Offer Name', width=150, minwidth=100)
            
            for category in category_order:
                category_tree.heading(category, text=category)
                category_tree.column(category, width=120, minwidth=80)
            
            # Add scrollbars
            category_vsb = ttk.Scrollbar(category_frame, orient="vertical", command=category_tree.yview)
            category_hsb = ttk.Scrollbar(category_frame, orient="horizontal", command=category_tree.xview)
            category_tree.configure(yscrollcommand=category_vsb.set, xscrollcommand=category_hsb.set)
            
            # Grid layout
            category_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
            category_vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
            category_hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
            
            # Collect all offers and their category costs
            offers_data = []
            
            # Get current offer name
            current_offer_name = "Current Offer"
            
            # First try to get from current_offer_info (for Use Mapping workflow)
            if hasattr(self, 'current_offer_info') and self.current_offer_info:
                offer_info = self.current_offer_info
                current_offer_name = offer_info.get('offer_name', current_offer_name)
                logger.debug(f"Found offer name from current_offer_info: {current_offer_name}")
            else:
                # Fallback: search through controller's current_files
                for file_key, file_data in self.controller.current_files.items():
                    if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                        if hasattr(file_data['file_mapping'], 'offer_info') and file_data['file_mapping'].offer_info:
                            offer_info = file_data['file_mapping'].offer_info
                            current_offer_name = offer_info.get('offer_name', current_offer_name)
                            logger.debug(f"Found offer name from file_mapping: {current_offer_name}")
                        elif 'offer_info' in file_data:
                            offer_info = file_data['offer_info']
                            current_offer_name = offer_info.get('offer_name', current_offer_name)
                            logger.debug(f"Found offer name from current_files: {current_offer_name}")
                        break
            
            # Calculate category costs for current offer
            current_offer_data = [current_offer_name]
            for category in category_order:
                try:
                    category_data = dataframe[dataframe['Category'] == category]
                    if len(category_data) > 0:
                        category_prices = pd.to_numeric(category_data['total_price'], errors='coerce')
                        category_cost = category_prices.sum()
                        if category_cost > 0:
                            formatted_cost = format_number_eu(category_cost)
                        else:
                            formatted_cost = "0,00"
                        current_offer_data.append(formatted_cost)
                    else:
                        current_offer_data.append("0,00")
                except Exception as e:
                    logger.warning(f"Error calculating cost for category {category}: {e}")
                    current_offer_data.append("0,00")
            
            offers_data.append(current_offer_data)
            
            # Check for comparison offers in the current dataframe (regardless of workflow flag)
            comparison_columns = [col for col in display_dataframe.columns if '[' in col and ']' in col and 'total_price' in col]
            logger.info(f"DEBUG: display_dataframe columns: {list(display_dataframe.columns)}")
            logger.info(f"DEBUG: comparison_columns found: {comparison_columns}")
            
            if comparison_columns:
                # Look for comparison offers in the current dataframe
                
                for col in comparison_columns:
                    try:
                        # Extract offer name from column name (e.g., "total_price[OfferName]" -> "OfferName")
                        offer_name_from_col = col.split('[')[1].split(']')[0]
                        
                        # Calculate category costs for this comparison offer
                        comparison_offer_data = [offer_name_from_col]
                        
                        for category in category_order:
                            try:
                                category_data = display_dataframe[display_dataframe['Category'] == category]
                                if len(category_data) > 0:
                                    # Use the comparison column for this offer
                                    category_prices = pd.to_numeric(category_data[col], errors='coerce')
                                    category_cost = category_prices.sum()
                                    if category_cost > 0:
                                        formatted_cost = format_number_eu(category_cost)
                                    else:
                                        formatted_cost = "0,00"
                                    comparison_offer_data.append(formatted_cost)
                                else:
                                    comparison_offer_data.append("0,00")
                            except Exception as e:
                                logger.warning(f"Error calculating cost for category {category} in comparison offer {offer_name_from_col}: {e}")
                                comparison_offer_data.append("0,00")
                        
                        offers_data.append(comparison_offer_data)
                        
                    except Exception as e:
                        logger.warning(f"Error processing comparison offer column {col}: {e}")
            
            # Insert all offers into category treeview
            for offer_data in offers_data:
                category_tree.insert('', 'end', values=offer_data)
            
            # Configure grid weights
            category_frame.columnconfigure(0, weight=1)
            category_frame.rowconfigure(0, weight=1)
            
        except Exception as e:
            logger.error(f"Error summarizing categorized data: {e}")
            messagebox.showerror("Error", f"Failed to summarize data: {str(e)}")

    def _save_mapping_for_categorized_data(self, dataframe):
        """Save the column mapping for future use"""
        try:
            filetypes = [("Pickle files", "*.pkl"), ("All files", "*.*")]
            filename = filedialog.asksaveasfilename(
                title="Save Column Mapping",
                filetypes=filetypes,
                defaultextension=".pkl"
            )
            
            if filename:
                # Get current sheet categories and column mappings
                current_sheet_categories = getattr(self, 'current_sheet_categories', {})
                current_column_mappings = {}
                
                # Get column mappings from current file mapping
                current_tab_path = self.notebook.select()
                for file_key, file_data in self.controller.current_files.items():
                    if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                        if hasattr(file_data['file_mapping'], 'sheets'):
                            for sheet in file_data['file_mapping'].sheets:
                                if hasattr(sheet, 'column_mappings'):
                                    current_column_mappings[sheet.sheet_name] = sheet.column_mappings
                        break
                
                # Create mapping data
                mapping_data = {
                    'columns': list(dataframe.columns),
                    'column_order': ['code', 'Source_Sheet', 'Category', 'Description', 'unit', 
                                   'quantity', 'unit_price', 'total_price', 'manhours', 'wage'],
                    'sheet_categories': current_sheet_categories,
                    'column_mappings': current_column_mappings,
                    'timestamp': time.time(),
                    'dataframe_shape': dataframe.shape
                }
                
                with open(filename, 'wb') as f:
                    pickle.dump(mapping_data, f)
                
                messagebox.showinfo("Mapping Saved", f"Column mapping saved to {filename}")
                
        except Exception as e:
            logger.error(f"Error saving mapping: {e}")
            messagebox.showerror("Error", f"Failed to save mapping: {str(e)}")

    def _save_analysis_for_categorized_data(self, dataframe, categorization_result):
        """Save the complete analysis including categorization results"""
        try:
            filetypes = [("Pickle files", "*.pkl"), ("All files", "*.*")]
            filename = filedialog.asksaveasfilename(
                title="Save Analysis",
                filetypes=filetypes,
                defaultextension=".pkl"
            )
            
            if filename:
                # Create analysis data
                total_cost = 0.0
                if 'total_price' in dataframe.columns:
                    try:
                        numeric_prices = pd.to_numeric(dataframe['total_price'], errors='coerce')
                        total_cost = numeric_prices.sum()
                    except Exception as e:
                        logger.warning(f"Error calculating total cost for analysis: {e}")
                        total_cost = 0.0
                
                # Get the current offer info from the controller
                current_offer_info = None
                current_tab_path = self.notebook.select()
                for file_key, file_data in self.controller.current_files.items():
                    if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                        if 'offer_info' in file_data:
                            current_offer_info = file_data['offer_info']
                        elif hasattr(file_data['file_mapping'], 'offer_info'):
                            current_offer_info = file_data['file_mapping'].offer_info
                        break
                
                analysis_data = {
                    'dataframe': dataframe,
                    'categorization_result': categorization_result,
                    'offer_info': current_offer_info or {
                        'offer_name': 'Unknown',
                        'project_name': 'Unknown',
                        'project_size': 'N/A',
                        'date': datetime.now().strftime('%Y-%m-%d')
                    },
                    'timestamp': time.time(),
                    'total_rows': len(dataframe),
                    'total_cost': total_cost,
                    'category_counts': dataframe['Category'].value_counts().to_dict() if 'Category' in dataframe.columns else {}
                }
                
                with open(filename, 'wb') as f:
                    pickle.dump(analysis_data, f)
                
                messagebox.showinfo("Analysis Saved", f"Complete analysis saved to {filename}")
                
        except Exception as e:
            logger.error(f"Error saving analysis: {e}")
            messagebox.showerror("Error", f"Failed to save analysis: {str(e)}")

    def _create_dataframe_from_mapping(self, file_mapping):
        """Create a DataFrame from file mapping data"""
        try:
            import pandas as pd
            
            # First, try to use the dataframe attribute if it exists
            if hasattr(file_mapping, 'dataframe') and file_mapping.dataframe is not None:
                logger.debug(f"Using existing dataframe from file mapping with {len(file_mapping.dataframe)} rows")
                return file_mapping.dataframe.copy()
            
            # Fallback: create DataFrame from sheet data
            rows = []
            for sheet in file_mapping.sheets:
                if hasattr(sheet, 'sheet_data') and sheet.sheet_data:
                    # Create a simple DataFrame from sheet data
                    # This is a basic implementation - you might need to enhance it
                    for i, row_data in enumerate(sheet.sheet_data):
                        if i == 0:  # Skip header row
                            continue
                        row_dict = {
                            'Source_Sheet': sheet.sheet_name,
                            'Description': row_data[0] if len(row_data) > 0 else '',
                            'code': row_data[1] if len(row_data) > 1 else '',
                            'unit': row_data[2] if len(row_data) > 2 else '',
                            'quantity': row_data[3] if len(row_data) > 3 else '',
                            'unit_price': row_data[4] if len(row_data) > 4 else '',
                            'total_price': row_data[5] if len(row_data) > 5 else '',
                        }
                        rows.append(row_dict)
            
            if rows:
                df = pd.DataFrame(rows)
                logger.debug(f"Created DataFrame with {len(df)} rows from file mapping sheet data")
                return df
            else:
                logger.warning("No data found in file mapping")
                return None
                
        except Exception as e:
            logger.error(f"Error creating DataFrame from mapping: {e}")
            return None

    def _create_dataframe_from_boq_results(self, boq_results):
        """Create a DataFrame from BOQ processor results"""
        try:
            import pandas as pd
            
            rows = []
            
            # Extract BOQ data from all sheets
            for sheet_name, sheet_data in boq_results.get('sheets_data', {}).items():
                if 'boq_data' in sheet_data and sheet_data['boq_data']:
                    for item in sheet_data['boq_data']:
                        row_dict = {
                            'Source_Sheet': sheet_name,
                            'Description': item.get('description', ''),
                            'code': item.get('code', ''),
                            'unit': item.get('unit', ''),
                            'quantity': item.get('quantity', ''),
                            'unit_price': item.get('unit_price', ''),
                            'total_price': item.get('total_price', ''),
                            'Category': item.get('classification', ''),
                        }
                        rows.append(row_dict)
            
            if rows:
                df = pd.DataFrame(rows)
                logger.debug(f"Created DataFrame with {len(df)} rows from BOQ results")
                return df
            else:
                logger.warning("No data found in BOQ results")
                return None
                
        except Exception as e:
            logger.error(f"Error creating DataFrame from BOQ results: {e}")
            return None

    def _compare_full_from_categorized_data(self, tab):
        """Start comparison workflow from categorized data"""
        try:
            # Store the current dataframe in the tab for comparison
            current_tab_path = self.notebook.select()
            current_tab = self.notebook.nametowidget(self.notebook.select())
            
            # Find the matching file data and store the dataframe
            for file_key, file_data in self.controller.current_files.items():
                if hasattr(file_data['file_mapping'], 'tab') and str(file_data['file_mapping'].tab) == str(current_tab_path):
                    # Get the dataframe from the current display
                    # This would need to be passed from the button command
                    # For now, we'll use the stored final_dataframe
                    if 'final_dataframe' in file_data:
                        file_data['file_mapping'].dataframe = file_data['final_dataframe']
                        self._compare_full(tab)
                    else:
                        messagebox.showerror("Error", "No categorized data available for comparison")
                    break
            else:
                messagebox.showerror("Error", "Could not find current file data")
                
        except Exception as e:
            logger.error(f"Error starting comparison: {e}")
            messagebox.showerror("Error", f"Failed to start comparison: {str(e)}")

    def _process_comparison_file_optimized(self, filepath, offer_info, master_file_mapping):
        """
        Process comparison file using master BOQ mapping for maximum efficiency
        
        Args:
            filepath: Path to comparison file
            offer_info: Offer information dictionary
            master_file_mapping: Master BOQ file mapping with known structure
        
        Returns:
            FileMapping object or None if failed
        """
        try:
            from core.file_processor import ExcelProcessor
            import pandas as pd
            
            # Create Excel processor
            excel_processor = ExcelProcessor()
            
            # Load the file
            if not excel_processor.load_file(filepath):
                return None
            
            # Get the exact sheets that exist in the master BOQ
            master_sheet_names = set()
            if hasattr(master_file_mapping, 'sheets') and master_file_mapping.sheets:
                master_sheet_names = {sheet.sheet_name for sheet in master_file_mapping.sheets}
            else:
                # Fallback: get visible sheets
                master_sheet_names = set(excel_processor.get_visible_sheets())
            
            # Get visible sheets from comparison file
            visible_sheets = excel_processor.get_visible_sheets()
            if not visible_sheets:
                return None
            
            # Check if all master sheets exist in comparison file
            missing_sheets = master_sheet_names - set(visible_sheets)
            if missing_sheets:
                error_msg = f"Comparison file is missing required sheets: {missing_sheets}"
                logger.error(error_msg)
                messagebox.showerror("Structure Mismatch", error_msg)
                return None
            
            # Only process the exact sheets from master BOQ
            sheets_to_process = list(master_sheet_names)
            logger.info(f"Processing {len(sheets_to_process)} sheets based on master BOQ structure")
            
            # Extract data directly using master mapping information
            all_data = []
            
            for sheet_name in sheets_to_process:
                try:
                    # Get sheet data with proper header detection
                    sheet_data = excel_processor.get_sheet_data(sheet_name, max_rows=10000)
                    if not sheet_data or len(sheet_data) < 2:  # Need at least header + 1 row
                        continue
                    
                    # Find the actual header row (skip empty rows at the top)
                    header_row_idx = 0
                    for i, row in enumerate(sheet_data):
                        if any(cell and str(cell).strip() for cell in row):
                            header_row_idx = i
                            break
                    
                    # Get headers and data
                    headers = sheet_data[header_row_idx]
                    data_rows = sheet_data[header_row_idx + 1:]
                    
                    # Clean headers (remove duplicates, handle empty headers)
                    clean_headers = []
                    for i, header in enumerate(headers):
                        if not header or pd.isna(header) or str(header).strip() == '':
                            # Try to find a meaningful header from the data
                            meaningful_header = None
                            for data_row in data_rows[:5]:  # Check first 5 rows
                                if i < len(data_row) and data_row[i] and str(data_row[i]).strip():
                                    meaningful_header = f"Column_{str(data_row[i]).strip()[:20]}"
                                    break
                            if not meaningful_header:
                                meaningful_header = f'Column_{i}'
                            clean_headers.append(meaningful_header)
                        else:
                            clean_headers.append(str(header).strip())
                    
                    # Ensure unique headers
                    unique_headers = []
                    seen_headers = set()
                    for header in clean_headers:
                        if header in seen_headers:
                            counter = 1
                            while f"{header}_{counter}" in seen_headers:
                                counter += 1
                            header = f"{header}_{counter}"
                        unique_headers.append(header)
                        seen_headers.add(header)
                    
                    # Convert to DataFrame with clean headers
                    df = pd.DataFrame(data_rows, columns=unique_headers)
                    if df.empty:
                        continue
                    
                    # Add sheet name column
                    df['Source_Sheet'] = sheet_name
                    
                    # Reset index to avoid conflicts
                    df = df.reset_index(drop=True)
                    
                    # Append to all data
                    all_data.append(df)
                    
                    logger.debug(f"Extracted {len(df)} rows from sheet '{sheet_name}' with headers: {unique_headers[:5]}...")
                    
                except Exception as e:
                    logger.warning(f"Failed to process sheet '{sheet_name}': {e}")
                    continue
            
            if not all_data:
                messagebox.showerror("Error", "No data could be extracted from comparison file")
                return None
            
            # Combine all data with proper error handling
            try:
                combined_df = pd.concat(all_data, ignore_index=True, sort=False)
                logger.info(f"Combined {len(combined_df)} total rows from comparison file")
            except Exception as e:
                logger.error(f"Error combining DataFrames: {e}")
                # Fallback: try with different approach
                try:
                    # Create a list of dictionaries and then DataFrame
                    all_rows = []
                    for df in all_data:
                        for _, row in df.iterrows():
                            all_rows.append(row.to_dict())
                    
                    combined_df = pd.DataFrame(all_rows)
                    logger.info(f"Combined {len(combined_df)} total rows using fallback method")
                except Exception as e2:
                    logger.error(f"Fallback method also failed: {e2}")
                    messagebox.showerror("Error", f"Failed to combine data from comparison file: {str(e)}")
                    return None
            
            # Create file mapping object
            file_mapping = type('MockFileMapping', (), {
                'dataframe': combined_df,
                'offer_info': offer_info,
                'sheets': []  # Add empty sheets list for compatibility
            })()
            
            # Store offer information
            file_mapping.offer_info = offer_info
            
            return file_mapping
            
        except Exception as e:
            logger.error(f"Error processing comparison file: {e}")
            return None

    def _process_comparison_file_with_master_mappings(self, filepath, master_file_mapping, offer_info):
        """
        Process comparison file using master BOQ mappings (header row index and column mapping from master)
        Args:
            filepath: Path to comparison file
            master_file_mapping: Master BOQ file mapping with known structure
            offer_info: Offer information dictionary
        Returns:
            FileMapping object or None if failed
        """
        try:
            from core.file_processor import ExcelProcessor
            import pandas as pd
            
            excel_processor = ExcelProcessor()
            if not excel_processor.load_file(filepath):
                return None
            
            # Get master mapping info
            master_sheets = getattr(master_file_mapping, 'sheets', [])
            if not master_sheets:
                logger.error("Master file mapping has no sheets info")
                return None
            
            visible_sheets = excel_processor.get_visible_sheets()
            if not visible_sheets:
                return None
            
            all_data = []
            for master_sheet in master_sheets:
                sheet_name = getattr(master_sheet, 'sheet_name', None)
                if not sheet_name or sheet_name not in visible_sheets:
                    continue
                header_row_idx = getattr(master_sheet, 'header_row_index', 0)
                column_mappings = getattr(master_sheet, 'column_mappings', [])
                if not column_mappings:
                    logger.warning(f"No column mappings for sheet {sheet_name}")
                    continue
                
                # Get sheet data
                sheet_data = excel_processor.get_sheet_data(sheet_name, max_rows=10000)
                if not sheet_data or len(sheet_data) <= header_row_idx:
                    continue
                
                # Get raw headers and data (no enhancement needed - we'll use column indices)
                headers = sheet_data[header_row_idx]
                data_rows = sheet_data[header_row_idx + 1:]
                
                # Build mapping from column index to canonical name using master mapping
                column_index_map = {}
                for cm in column_mappings:
                    col_idx = getattr(cm, 'column_index', None)
                    canon = getattr(cm, 'mapped_type', None)
                    if col_idx is not None and canon:
                        # mapped_type may be an Enum or str
                        canon_val = canon.value if hasattr(canon, 'value') else str(canon)
                        # Ensure correct case for canonical names
                        if canon_val.lower() == 'description':
                            canon_val = 'Description'
                        elif canon_val.lower() == 'quantity':
                            canon_val = 'quantity'
                        elif canon_val.lower() == 'unit_price':
                            canon_val = 'unit_price'
                        elif canon_val.lower() == 'total_price':
                            canon_val = 'total_price'
                        elif canon_val.lower() == 'unit':
                            canon_val = 'unit'
                        elif canon_val.lower() == 'code':
                            canon_val = 'code'
                        column_index_map[col_idx] = canon_val
                
                # Map columns by index - only include mapped columns
                mapped_columns = []
                valid_column_indices = []  # Track which indices are actually valid
                for col_idx in sorted(column_index_map.keys()):
                    if col_idx < len(headers):  # Ensure column index is valid
                        mapped_columns.append(column_index_map[col_idx])
                        valid_column_indices.append(col_idx)  # Track valid indices
                
                # Ensure unique column names
                unique_columns = []
                seen_columns = set()
                for col in mapped_columns:
                    if col in seen_columns:
                        counter = 1
                        while f"{col}_{counter}" in seen_columns:
                            counter += 1
                        col = f"{col}_{counter}"
                    unique_columns.append(col)
                    seen_columns.add(col)
                
                # Filter data rows to match only mapped columns by index
                # Use only valid_column_indices to ensure row length matches unique_columns length
                filtered_data_rows = []
                for row in data_rows:
                    filtered_row = []
                    for col_idx in valid_column_indices:  # Only use indices that were included in mapped_columns
                        if col_idx < len(row):  # Ensure column index is valid
                            filtered_row.append(row[col_idx] if row[col_idx] is not None else '')
                        else:
                            filtered_row.append('')  # Fill missing columns
                    filtered_data_rows.append(filtered_row)
                
                # Build DataFrame with only mapped columns
                df = pd.DataFrame(filtered_data_rows, columns=unique_columns)
                if df.empty:
                    continue
                
                # Add required columns that may be missing
                if 'Category' not in df.columns:
                    df['Category'] = None
                if 'Source_Sheet' not in df.columns:
                    df['Source_Sheet'] = sheet_name
                
                df = df.reset_index(drop=True)
                all_data.append(df)
                logger.debug(f"Extracted {len(df)} rows from sheet '{sheet_name}' using master mapping")
            
            if not all_data:
                logger.error("No data could be extracted from comparison file using master mappings")
                return None
            
            # Combine all data
            try:
                combined_df = pd.concat(all_data, ignore_index=True, sort=False)
                logger.info(f"Combined {len(combined_df)} total rows from comparison file (master mapping)")
            except Exception as e:
                logger.error(f"Error combining DataFrames: {e}")
                return None
            
            file_mapping = type('MockFileMapping', (), {
                'dataframe': combined_df,
                'offer_info': offer_info,
                'sheets': []
            })()
            file_mapping.offer_info = offer_info
            return file_mapping
        except Exception as e:
            logger.error(f"Error processing comparison file with master mappings: {e}")
            return None

    def _create_unified_dataframe(self, file_mapping, is_master=True):
        """
        Create a unified DataFrame from file mapping data with consistent structure
        
        Args:
            file_mapping: FileMapping object containing the data
            is_master: Whether this is the master dataset (affects logging)
            
        Returns:
            DataFrame with consistent column structure
        """
        try:
            import pandas as pd
            
            dataset_type = "master" if is_master else "comparison"
            logger.info(f"Creating unified DataFrame for {dataset_type} dataset")
            
            # First, try to use the dataframe attribute if it exists
            if hasattr(file_mapping, 'dataframe') and file_mapping.dataframe is not None:
                df = file_mapping.dataframe.copy()
                logger.info(f"Using existing dataframe for {dataset_type}: {len(df)} rows, columns: {list(df.columns)}")
                
                # Normalize Description column - handle both 'Description' and 'description'
                if 'description' in df.columns and 'Description' in df.columns:
                    # Both exist - merge them, preferring 'Description' but filling from 'description' where empty
                    logger.info(f"Both 'Description' and 'description' columns found, merging them")
                    df['Description'] = df['Description'].fillna('').astype(str)
                    df['description'] = df['description'].fillna('').astype(str)
                    # Use 'Description' if it has content, otherwise use 'description'
                    mask = (df['Description'].str.strip() == '') | (df['Description'].isna())
                    df.loc[mask, 'Description'] = df.loc[mask, 'description']
                    # Drop the lowercase version
                    df = df.drop(columns=['description'])
                elif 'description' in df.columns and 'Description' not in df.columns:
                    # Only lowercase exists - rename it
                    logger.info(f"Only 'description' column found, renaming to 'Description'")
                    df = df.rename(columns={'description': 'Description'})
                
                # Ensure we have the required columns
                required_columns = ['Description', 'code', 'unit', 'quantity', 'unit_price', 'total_price']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    logger.warning(f"Missing columns in {dataset_type} dataset: {missing_columns}")
                    # Add missing columns with empty values
                    for col in missing_columns:
                        df[col] = ''
                
                # Ensure Source_Sheet column is populated for debug export
                if 'Source_Sheet' not in df.columns:
                    logger.info(f"Adding Source_Sheet column to {dataset_type} dataset")
                    # Try to populate Source_Sheet from sheet information if available
                    if hasattr(file_mapping, 'sheets') and file_mapping.sheets:
                        # Create a mapping of row indices to sheet names
                        sheet_mapping = {}
                        current_row = 0
                        for sheet in file_mapping.sheets:
                            if hasattr(sheet, 'sheet_data') and sheet.sheet_data:
                                sheet_rows = len(sheet.sheet_data) - 1  # Exclude header row
                                for i in range(sheet_rows):
                                    sheet_mapping[current_row + i] = sheet.sheet_name
                                current_row += sheet_rows
                        
                        # Populate Source_Sheet column
                        df['Source_Sheet'] = df.index.map(lambda x: sheet_mapping.get(x, 'Unknown'))
                    else:
                        # Fallback: use a default value
                        df['Source_Sheet'] = 'Unknown'
                elif df['Source_Sheet'].isna().all() or (df['Source_Sheet'] == '').all():
                    logger.warning(f"Source_Sheet column exists but is empty in {dataset_type} dataset")
                    # Try to populate from sheet information
                    if hasattr(file_mapping, 'sheets') and file_mapping.sheets:
                        sheet_mapping = {}
                        current_row = 0
                        for sheet in file_mapping.sheets:
                            if hasattr(sheet, 'sheet_data') and sheet.sheet_data:
                                sheet_rows = len(sheet.sheet_data) - 1  # Exclude header row
                                for i in range(sheet_rows):
                                    sheet_mapping[current_row + i] = sheet.sheet_name
                                current_row += sheet_rows
                        
                        df['Source_Sheet'] = df.index.map(lambda x: sheet_mapping.get(x, 'Unknown'))
                    else:
                        df['Source_Sheet'] = 'Unknown'
                
                # Ensure consistent column order
                final_columns = required_columns + [col for col in df.columns if col not in required_columns]
                df = df[final_columns]
                
                # Log sample descriptions to verify they're not empty
                if 'Description' in df.columns:
                    sample_descriptions = df['Description'].head(5).tolist()
                    logger.info(f"Sample descriptions in {dataset_type} dataset: {sample_descriptions}")
                else:
                    logger.error(f"No Description column found in {dataset_type} dataset!")
                
                return df
            
            # Fallback: create DataFrame from sheet data
            logger.warning(f"No dataframe attribute found for {dataset_type}, falling back to sheet data")
            rows = []
            for sheet in file_mapping.sheets:
                if hasattr(sheet, 'sheet_data') and sheet.sheet_data:
                    for i, row_data in enumerate(sheet.sheet_data):
                        if i == 0:  # Skip header row
                            continue
                        row_dict = {
                            'Source_Sheet': sheet.sheet_name,
                            'Description': row_data[0] if len(row_data) > 0 else '',
                            'code': row_data[1] if len(row_data) > 1 else '',
                            'unit': row_data[2] if len(row_data) > 2 else '',
                            'quantity': row_data[3] if len(row_data) > 3 else '',
                            'unit_price': row_data[4] if len(row_data) > 4 else '',
                            'total_price': row_data[5] if len(row_data) > 5 else '',
                        }
                        rows.append(row_dict)
            
            if rows:
                df = pd.DataFrame(rows)
                logger.info(f"Created DataFrame from sheet data for {dataset_type}: {len(df)} rows")
                return df
            else:
                logger.error(f"No data found in {dataset_type} file mapping")
                return None
                
        except Exception as e:
            logger.error(f"Error creating unified DataFrame for {dataset_type}: {e}")
            return None

    def run(self):
        """Start the main window event loop"""
        try:
            # Start the main event loop
            self.root.mainloop()
        except Exception as e:
            logger.error(f"Error in main window event loop: {e}")
            raise
    
    def _process_comparison_file_with_sheet_structure(self, filepath, offer_info):
        """Process comparison file with sheet structure, like master BOQ workflow."""
        # Use the same logic as _process_file, but for the comparison file
        file_mapping = self.controller.process_file(
            Path(filepath),
            progress_callback=lambda p, m: self.update_progress(p, m)
        )
        file_mapping.offer_info = offer_info
        return file_mapping

    def _show_comparison_row_review(self, file_mapping, offer_info):
        """Show row review for comparison file, then create unified dataset and export after confirmation."""
        # Show the row review dialog (reusing the master workflow)
        self._show_row_review(file_mapping)
        # After user confirms row review, _on_confirm_row_review will be called as usual
        # We can hook into that to trigger the debug export for comparison
        self._pending_comparison_export = True
        self._pending_comparison_offer_info = offer_info

    def _update_tab_with_comparison_data(self, tab, updated_df):
        print("=== ENTERED _update_tab_with_comparison_data ===")
        logger.info("Updating tab with comparison data")
        logger.info(f"Tab children count: {len(tab.winfo_children())}")
        
        # Find the existing treeview in the tab
        treeview = None
        for widget in tab.winfo_children():
            logger.info(f"Widget type: {type(widget)}")
            if isinstance(widget, ttk.Frame):
                logger.info(f"Frame children count: {len(widget.winfo_children())}")
                for child in widget.winfo_children():
                    logger.info(f"Child type: {type(child)}")
                    if isinstance(child, ttk.Treeview):
                        treeview = child
                        logger.info("Found treeview!")
                        break
                if treeview:
                    break
        
        if not treeview:
            logger.warning("No treeview found in tab")
            return
        
        # Clear existing data
        for item in treeview.get_children():
            treeview.delete(item)
        
        # Get the current columns from the treeview
        current_columns = treeview['columns']
        
        # Filter out ignore columns from the updated dataframe
        meaningful_columns = [col for col in updated_df.columns if not (col.startswith('ignore') or col == 'ignore')]
        
        # Normalize column names to handle case differences
        # Map common variations to standard names
        column_mapping = {
            'description': 'Description',
            'Description': 'Description',
            'DESCRIPTION': 'Description',
            'category': 'Category',
            'Category': 'Category',
            'CATEGORY': 'Category',
            'code': 'code',
            'unit': 'unit',
            'quantity': 'quantity',
            'unit_price': 'unit_price',
            'total_price': 'total_price',
            'manhours': 'manhours',
            'wage': 'wage',
            'source_sheet': 'Source_Sheet',
            'Source_Sheet': 'Source_Sheet',
            'SOURCE_SHEET': 'Source_Sheet'
        }
        
        # Rename columns to standard names
        for col in updated_df.columns:
            if col.lower() in column_mapping:
                new_name = column_mapping[col.lower()]
                if col != new_name:
                    updated_df = updated_df.rename(columns={col: new_name})
        
        # Update meaningful_columns after renaming
        meaningful_columns = [col for col in updated_df.columns if not (col.startswith('ignore') or col == 'ignore')]
        
        # Define the exact column order from original master BOQ (same as main.py)
        # Source_Sheet should be the first column
        # Use consistent naming - match the actual DataFrame column names
        display_column_order = ["Source_Sheet", "code", "Category", "Description", "scope", "unit", "quantity", "unit_price", "total_price", "manhours", "wage"]
         
        # Debug: Log column information
        logger.info(f"Updated dataframe columns: {list(updated_df.columns)}")
        logger.info(f"Meaningful columns: {meaningful_columns}")
        logger.info(f"Display column order: {display_column_order}")
        
        # Check which display columns are actually in meaningful_columns
        available_display_columns = [col for col in display_column_order if col in meaningful_columns]
        logger.info(f"Available display columns: {available_display_columns}")
        
        # Ensure Source_Sheet column is present
        if 'Source_Sheet' not in meaningful_columns:
            # Add Source_Sheet column with default value
            updated_df['Source_Sheet'] = 'Unknown'
            meaningful_columns.append('Source_Sheet')
            logger.info("Added Source_Sheet column with default value")
        
        # Add any offer-specific columns (those with [offer_name] pattern)
        offer_columns = [col for col in meaningful_columns if '[' in col and ']' in col]
        
        # Create final column order: standard columns first, then offer-specific columns
        final_columns = []
        for col in display_column_order:
            if col in meaningful_columns:
                final_columns.append(col)
        
        # Add offer-specific columns at the end
        for col in offer_columns:
            if col not in final_columns:
                final_columns.append(col)
        
        # Store the updated dataframe with comparison columns for summary methods
        # This ensures the comparison columns are available when summary methods are called
        self._current_dataframe_with_comparison = updated_df.copy()
        
        # Add any remaining meaningful columns
        for col in meaningful_columns:
            if col not in final_columns:
                final_columns.append(col)
        
        logger.info(f"Final column order: {final_columns}")
        
        # Update treeview columns if needed
        if list(treeview['columns']) != final_columns:
            treeview['columns'] = final_columns
            
            # Apply light blue selection style to the treeview
            style = ttk.Style(treeview)
            style.map('Treeview', 
                     background=[('selected', '#B3E5FC')])  # Light blue background
            
            # Configure columns with appropriate widths (same as master window)
            column_widths = {
                'Source_Sheet': 120,
                'code': 100,
                'Category': 150,
                'Description': 400,  # Wider for descriptions
                'scope': 100,
                'unit': 80,
                'quantity': 100,
                'unit_price': 120,
                'total_price': 120,
                'manhours': 100,
                'wage': 80
            }
            
            # Set default width for offer-specific columns
            default_offer_width = 120
            
            for col in final_columns:
                # Display pretty names in headers
                pretty_name = self._get_pretty_column_name(col)
                treeview.heading(col, text=pretty_name)
                
                # Use specific width if defined, otherwise use default for offer columns
                if col in column_widths:
                    width = column_widths[col]
                elif '[' in col and ']' in col:
                    width = default_offer_width
                else:
                    width = 100
                treeview.column(col, width=width, minwidth=80, stretch=False)
        
        # Populate with data using exact same formatting as row review
        for idx, row in updated_df.iterrows():
            values = []
            for col in final_columns:
                val = row.get(col, '')
                
                # Apply exact same formatting as row review
                if pd.notna(val) and val != '':
                    # Check if this is a comparison column (has [offer_name] pattern)
                    is_comparison_col = '[' in col and ']' in col
                    
                    # Extract base column name for comparison columns
                    base_col = col.split('[')[0] if is_comparison_col else col
                    
                    if base_col in ['unit_price', 'total_price', 'wage']:
                        # Currency formatting (same as row review) - applies to both master and comparison
                        val = format_number_eu(val)
                    elif base_col in ['quantity', 'manhours']:
                        # Use standard European number formatting for both quantity and manhours
                        val = format_number_eu(val)
                    else:
                        val = str(val)
                else:
                    val = ''
                
                values.append(val)
            
            treeview.insert('', 'end', values=values)
        
        logger.info("Tab updated successfully with comparison data")
        
        # Refresh the summary to show all offers including comparison offers
        try:
            self._refresh_summary_after_comparison(tab, updated_df)
            logger.info("Summary refreshed after comparison data update")
        except Exception as e:
            logger.error(f"Error refreshing summary after comparison: {e}")
        
    def _refresh_summary_after_comparison(self, tab, updated_df):
        """Refresh the summary frame to show all offers after comparison data is loaded"""
        try:
            # Find the summary frame in the current tab
            main_frame = None
            for widget in tab.winfo_children():
                if isinstance(widget, ttk.Frame):
                    main_frame = widget
                    break
            
            if not main_frame:
                logger.error("Could not find main frame for summary refresh")
                return
            
            # Find the summary frame
            summary_frame = None
            for widget in main_frame.winfo_children():
                if isinstance(widget, ttk.LabelFrame) and widget.cget('text') == 'Summary':
                    summary_frame = widget
                    break
            
            if not summary_frame:
                logger.error("Could not find summary frame")
                return
            
            # Find the summary treeview
            summary_tree = None
            for widget in summary_frame.winfo_children():
                if isinstance(widget, ttk.Treeview):
                    summary_tree = widget
                    break
            
            if not summary_tree:
                logger.error("Could not find summary treeview")
                return
            
            # Clear existing summary data
            for item in summary_tree.get_children():
                summary_tree.delete(item)
            
            # Collect all offers and their total costs
            offers_data = []
            
            # Get current offer info
            current_tab_path = self.notebook.select()
            offer_name = "Current Offer"
            project_name = "Current Project"
            project_size = "N/A"
            date = datetime.now().strftime('%Y-%m-%d')
            
            
            # Try to get actual offer info
            if hasattr(self, 'current_file_mapping') and self.current_file_mapping and hasattr(self.current_file_mapping, 'offer_info'):
                offer_info = self.current_file_mapping.offer_info
                offer_name = offer_info.get('offer_name', offer_name)
                project_name = offer_info.get('project_name', project_name)
                project_size = offer_info.get('project_size', project_size)
                date = offer_info.get('date', date)
            elif hasattr(self, 'file_mapping') and self.file_mapping and hasattr(self.file_mapping, 'offer_info'):
                offer_info = self.file_mapping.offer_info
                offer_name = offer_info.get('offer_name', offer_name)
                project_name = offer_info.get('project_name', project_name)
                project_size = offer_info.get('project_size', project_size)
                date = offer_info.get('date', date)
            elif hasattr(self, 'current_offer_info') and self.current_offer_info:
                offer_info = self.current_offer_info
                offer_name = offer_info.get('offer_name', offer_name)
                project_name = offer_info.get('project_name', project_name)
                project_size = offer_info.get('project_size', project_size)
                date = offer_info.get('date', date)
            
            # Calculate total cost for current offer
            total_cost = 0.0
            if 'total_price' in updated_df.columns:
                try:
                    numeric_prices = pd.to_numeric(updated_df['total_price'], errors='coerce')
                    total_cost = numeric_prices.sum()
                except Exception as e:
                    logger.warning(f"Error calculating total cost: {e}")
                    total_cost = 0.0
            
            # Format total cost
            if total_cost > 0:
                formatted_total_cost = format_number_eu(total_cost)
            else:
                formatted_total_cost = "0,00"
            
            # Add current offer
            current_offer_data = {
                'offer_name': offer_name,
                'project_name': project_name,
                'project_size': project_size,
                'date': date,
                'total_cost': total_cost,
                'formatted_total_cost': formatted_total_cost
            }
            offers_data.append(current_offer_data)
            
            # Add comparison offers
            comparison_columns = [col for col in updated_df.columns if '[' in col and ']' in col and 'total_price' in col]
            
            for col in comparison_columns:
                try:
                    # Extract offer name from column name
                    offer_name_from_col = col.split('[')[1].split(']')[0]
                    
                    # Calculate total cost for this offer
                    offer_total_cost = 0.0
                    if col in updated_df.columns:
                        numeric_prices = pd.to_numeric(updated_df[col], errors='coerce')
                        offer_total_cost = numeric_prices.sum()
                    
                    # Format total cost
                    if offer_total_cost > 0:
                        formatted_offer_cost = format_number_eu(offer_total_cost)
                    else:
                        formatted_offer_cost = "0,00"
                    
                    # Add comparison offer data - use the user's entered project info
                    # Get the comparison offer info from the controller's current_files
                    comparison_offer_info = None
                    for file_key, file_data in self.controller.current_files.items():
                        if 'offers' in file_data and offer_name_from_col in file_data['offers']:
                            comparison_offer_info = file_data['offers'][offer_name_from_col]
                            break
                    
                    # Use user's entered project info if available, otherwise use defaults
                    if comparison_offer_info:
                        project_name = comparison_offer_info.get('project_name', f"Project {offer_name_from_col}")
                        project_size = comparison_offer_info.get('project_size', 'N/A')
                        date = comparison_offer_info.get('date', date)
                    else:
                        project_name = f"Project {offer_name_from_col}"
                        project_size = "N/A"
                    
                    comparison_offer_data = {
                        'offer_name': offer_name_from_col,
                        'project_name': project_name,
                        'project_size': project_size,
                        'date': date,
                        'total_cost': offer_total_cost,
                        'formatted_total_cost': formatted_offer_cost
                    }
                    offers_data.append(comparison_offer_data)
                    
                except Exception as e:
                    logger.warning(f"Error processing comparison offer column {col}: {e}")
            
            # Sort offers by total cost (lowest first)
            offers_data.sort(key=lambda x: x['total_cost'])
            
            # Insert all offers into summary treeview (ONLY ONCE)
            for offer_data in offers_data:
                summary_tree.insert('', 'end', values=(
                    offer_data['offer_name'],
                    offer_data['project_name'],
                    offer_data['project_size'],
                    offer_data['date'],
                    offer_data['formatted_total_cost']
                ))
            
            logger.info(f"Summary refreshed with {len(offers_data)} offers")
            
        except Exception as e:
            logger.error(f"Error refreshing summary after comparison: {e}")
            messagebox.showerror("Error", f"Failed to refresh summary: {str(e)}")

    def _get_pretty_column_name(self, column_name):
        """
        Convert internal column names to pretty display names
        
        Args:
            column_name: Internal column name
            
        Returns:
            Pretty display name for the column
        """
        # Handle comparison columns (those with [offer_name] pattern)
        if '[' in column_name and ']' in column_name:
            base_col = column_name.split('[')[0]
            offer_name = column_name.split('[')[1].split(']')[0]
            pretty_base = self._get_pretty_column_name(base_col)
            return f"{pretty_base} [{offer_name}]"
        
        # Map internal names to pretty names
        pretty_names = {
            'source_sheet': 'Source Sheet',
            'Source_Sheet': 'Source Sheet',
            'code': 'Code',
            'category': 'Category',
            'Category': 'Category',
            'description': 'Description',
            'Description': 'Description',
            'scope': 'Scope',
            'unit': 'Unit',
            'quantity': 'Quantity',
            'unit_price': 'Unit Price',
            'total_price': 'Total Price',
            'manhours': 'Man Hours',
            'wage': 'Wage'
        }
        
        return pretty_names.get(column_name, column_name)



