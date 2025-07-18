"""
Excel and Data Export Utilities for BOQ Tools
Handles export of processed BOQ data to various formats like Excel, CSV, and JSON.
"""

import logging
import json
from pathlib import Path
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Handles exporting processed BOQ data to different file formats.
    """

    def __init__(self):
        """Initialize the exporter."""
        logger.info("ExcelExporter initialized.")

    def export_data(self, file_mapping: Dict[str, Any], export_path: Path, format_type: str) -> bool:
        """
        Export data to the specified format.

        Args:
            file_mapping: The processed file mapping data.
            export_path: The destination path for the exported file.
            format_type: The format to export to ('normalized_excel', 'summary_excel', 'json', 'csv').

        Returns:
            True if export was successful, False otherwise.
        """
        export_methods = {
            'normalized_excel': self.export_normalized_boq,
            'summary_excel': self.export_summary_report,
            'json': self.export_to_json,
            'csv': self.export_to_csv,
        }

        method = export_methods.get(format_type)
        if not method:
            logger.error(f"Unknown export format: {format_type}")
            return False

        try:
            return method(file_mapping, export_path)
        except Exception as e:
            logger.error(f"Failed to export to {format_type}: {e}", exc_info=True)
            return False

    def export_to_json(self, file_mapping: Dict[str, Any], export_path: Path) -> bool:
        """Export mapping data to a JSON file."""
        export_path.parent.mkdir(parents=True, exist_ok=True)
        with open(export_path, 'w', encoding='utf-8') as f:
            json.dump(file_mapping, f, indent=2, ensure_ascii=False)
        logger.info(f"Successfully exported data to JSON: {export_path}")
        return True

    def export_to_csv(self, file_mapping: Dict[str, Any], export_path: Path) -> bool:
        """Export normalized data to a CSV file."""
        df = self._create_dataframe(file_mapping)
        export_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(export_path, index=False, encoding='utf-8')
        logger.info(f"Successfully exported data to CSV: {export_path}")
        return True

    def export_normalized_boq(self, file_mapping: Dict[str, Any], export_path: Path) -> bool:
        """Export normalized BOQ data to an Excel file."""
        df = self._create_dataframe(file_mapping)
        export_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Normalized BOQ', index=False)
            # Formatting
            ws = writer.sheets['Normalized BOQ']
            self._format_excel_sheet(ws)
            
            # Add summary sheet with formulas
            self._add_summary_sheet_with_formulas(writer.book, df, "Normalized BOQ")
            
        logger.info(f"Successfully exported normalized BOQ to Excel: {export_path}")
        return True

    def export_summary_report(self, file_mapping: Dict[str, Any], export_path: Path) -> bool:
        """Export a summary report to an Excel file."""
        # This can be expanded with more summary details
        summary_data = []
        for sheet_name, sheet_data in file_mapping.get('sheets', {}).items():
            item_count = len(sheet_data.get('items', []))
            total_value = sum(
                float(item.get('total_price', 0) or 0)
                for item in sheet_data.get('items', [])
            )
            summary_data.append({
                "Sheet Name": sheet_name,
                "Item Count": item_count,
                "Total Value": total_value,
            })
        df = pd.DataFrame(summary_data)
        export_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summary Report', index=False)
            ws = writer.sheets['Summary Report']
            self._format_excel_sheet(ws)
        logger.info(f"Successfully exported summary report to Excel: {export_path}")
        return True
        
    def _create_dataframe(self, file_mapping: Dict[str, Any]) -> pd.DataFrame:
        """Create a pandas DataFrame from the file mapping."""
        records = []
        for sheet_name, sheet_data in file_mapping.get('sheets', {}).items():
            for item in sheet_data.get('items', []):
                record = {
                    "Item No": item.get('item_no'),
                    "Description": item.get('description'),
                    "Unit": item.get('unit'),
                    "Quantity": item.get('quantity'),
                    "Unit Price": item.get('unit_price'),
                    "Total Price": item.get('total_price'),
                    "Category": item.get('category'),
                    "Source Sheet": sheet_name,
                }
                records.append(record)
        return pd.DataFrame(records)

    def _format_excel_sheet(self, ws):
        """Apply standard formatting to an Excel worksheet with European number formatting and data validation."""
        from openpyxl.styles import NamedStyle
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Format header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Apply European number formatting to numeric columns
        numeric_columns = []
        for col_idx, col_name in enumerate(ws[1], 1):
            col_lower = str(col_name.value).lower()
            if any(keyword in col_lower for keyword in ['price', 'quantity', 'manhours', 'wage']):
                numeric_columns.append((col_idx, col_name.value))
        
        # Apply formatting to numeric columns
        for col_idx, col_name in numeric_columns:
            # Create European number style
            euro_style = NamedStyle(name=f"euro_number_{col_name}")
            euro_style.number_format = '#,##0.00'
            
            # Apply to the column
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):  # Skip header row
                cell = ws[f'{col_letter}{row}']
                cell.style = euro_style
        
        # Add data validation for Category column if it exists
        category_col_idx = None
        for col_idx, col_name in enumerate(ws[1], 1):
            if str(col_name.value).lower() == 'category':
                category_col_idx = col_idx
                break
        
        if category_col_idx:
            category_col_letter = get_column_letter(category_col_idx)
            
            # Get available categories in the correct order (from category_order)
            try:
                category_order = ['General Costs', 'Site Costs', 'Civil Works', 'Earth Movement', 'Roads', 'OEM Building', 'Electrical Works', 'Solar Cables', 'LV Cables', 'MV Cables', 'Trenching', 'PV Mod. Installation', 'Cleaning and Cabling of PV Mod.', 'Tracker Inst.', 'Other']
                available_categories = category_order
                
                # Create data validation for Category column
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(available_categories)}"',
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Invalid Category",
                    error="Please select a category from the dropdown list.",
                    showInputMessage=True,
                    promptTitle="Category Selection",
                    prompt="Select a category from the dropdown list."
                )
                
                # Add validation to worksheet
                ws.add_data_validation(dv)
                
                # Apply validation to Category column (skip header row)
                for row in range(2, ws.max_row + 1):
                    dv.add(f'{category_col_letter}{row}')
                
            except Exception as e:
                logger.warning(f"Could not add category validation: {e}")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_summary_sheet_with_formulas(self, workbook, dataframe, sheet_name):
        """Add a summary sheet with formulas that calculate totals for each category"""
        try:
            # Create summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            
            # Define category order (same as in main_window.py)
            category_order = [
                "General Costs",
                "Site Costs", 
                "Civil Works",
                "Earth Movement",
                "Roads",
                "OEM Building",
                "Electrical Works",
                "Solar Cables",
                "LV Cables", 
                "MV Cables",
                "Trenching",
                "PV Mod. Installation",
                "Cleaning and Cabling of PV Mod.",
                "Tracker Inst.",
                "Other"
            ]
            
            # Create headers
            headers = ['Offer Name'] + category_order
            for col_idx, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add offer name
            summary_sheet.cell(row=2, column=1, value=sheet_name)
            
            # Add formulas for each category
            for col_idx, category in enumerate(category_order, 2):  # Start from column 2 (after Offer Name)
                # Find the Category column in the main data sheet
                category_col_idx = None
                for idx, col_name in enumerate(dataframe.columns, 1):
                    if col_name == 'Category':
                        category_col_idx = idx
                        break
                
                # Find the total_price column in the main data sheet
                total_price_col_idx = None
                for idx, col_name in enumerate(dataframe.columns, 1):
                    if col_name == 'Total Price':
                        total_price_col_idx = idx
                        break
                
                if category_col_idx and total_price_col_idx:
                    # Create SUMIFS formula
                    category_col_letter = get_column_letter(category_col_idx)
                    total_price_col_letter = get_column_letter(total_price_col_idx)
                    
                    formula = f'=SUMIFS(\'{sheet_name}\'!{total_price_col_letter}:{total_price_col_letter},\'{sheet_name}\'!{category_col_letter}:{category_col_letter},"{category}")'
                    
                    cell = summary_sheet.cell(row=2, column=col_idx)
                    cell.value = formula
                    
                    # Apply European number formatting
                    cell.number_format = '#,##0.00'
                else:
                    # If columns not found, set to 0
                    summary_sheet.cell(row=2, column=col_idx, value=0)
                    summary_sheet.cell(row=2, column=col_idx).number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error adding summary sheet with formulas: {e}")
            # Don't raise the exception, just log it so the main export can continue 