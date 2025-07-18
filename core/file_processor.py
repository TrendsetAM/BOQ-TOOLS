"""
Excel File Processor for BOQ Tools
Comprehensive Excel file handling with metadata extraction and memory management
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass
import warnings

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils.exceptions import InvalidFileException
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Excel processing will not work.")

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logging.warning("xlrd not available. Legacy .xls files will not be supported.")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("pandas not available. CSV processing will not work.")

logger = logging.getLogger(__name__)


@dataclass
class SheetMetadata:
    """Metadata information for an Excel sheet"""
    name: str
    is_visible: bool
    row_count: int
    column_count: int
    data_density: float
    first_data_row: int
    last_data_row: int
    first_data_column: int
    last_data_column: int
    empty_rows_count: int
    empty_columns_count: int
    file_format: str
    estimated_size_mb: float


@dataclass
class ContentSample:
    """Content sample from an Excel sheet"""
    sheet_name: str
    rows: List[List[str]]
    headers: List[str]
    sample_size: int
    has_data: bool


class ExcelProcessor:
    """
    Comprehensive Excel file processor with metadata extraction and memory management
    """
    
    def __init__(self, max_memory_mb: int = 512, chunk_size: int = 1000):
        """
        Initialize the Excel processor
        
        Args:
            max_memory_mb: Maximum memory usage in MB
            chunk_size: Number of rows to process in chunks
        """
        self.max_memory_mb = max_memory_mb
        self.chunk_size = chunk_size
        self.workbook: Optional[Workbook] = None
        self.file_path: Optional[Path] = None
        self.file_format: Optional[str] = None
        self._sheet_metadata_cache: Dict[str, SheetMetadata] = {}
        
        # Validate dependencies
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing")
    
    def load_file(self, filepath: Union[str, Path]) -> bool:
        """
        Safely load an Excel file with comprehensive error handling
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            True if file loaded successfully, False otherwise
            
        Raises:
            FileNotFoundError: If file doesn't exist
            InvalidFileException: If file is corrupted or unsupported
            MemoryError: If file exceeds memory limits
        """
        filepath = Path(filepath)
        
        try:
            # Validate file exists
            if not filepath.exists():
                raise FileNotFoundError(f"File not found: {filepath}")
            
            # Check file size
            file_size_mb = filepath.stat().st_size / (1024 * 1024)
            if file_size_mb > self.max_memory_mb:
                raise MemoryError(f"File size ({file_size_mb:.1f}MB) exceeds memory limit ({self.max_memory_mb}MB)")
            
            # Determine file format
            self.file_format = self._detect_file_format(filepath)
            logger.info(f"Detected file format: {self.file_format}")
            
            # Load workbook based on format
            if self.file_format == "xlsx":
                self._load_xlsx_file(filepath)
            elif self.file_format == "csv":
                self._load_csv_file(filepath)
            elif self.file_format == "xls":
                self._load_xls_file(filepath)
            else:
                raise InvalidFileException(f"Unsupported file format: {self.file_format}")
            
            self.file_path = filepath
            logger.info(f"Successfully loaded Excel file: {filepath.name}")
            if self.workbook:
                logger.info(f"Total sheets: {len(self.workbook.sheetnames)}")
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to load Excel file {filepath}: {str(e)}")
            self._cleanup()
            raise
    
    def _detect_file_format(self, filepath: Path) -> str:
        """Detect Excel file format based on extension and content"""
        extension = filepath.suffix.lower()
        
        if extension == ".xlsx":
            return "xlsx"
        elif extension == ".csv":
            if not PANDAS_AVAILABLE:
                raise ImportError("pandas is required for .csv files")
            return "csv"
        elif extension == ".xls":
            if not XLRD_AVAILABLE:
                raise ImportError("xlrd is required for .xls files")
            return "xls"
        else:
            # Try to detect by content
            try:
                with open(filepath, 'rb') as f:
                    header = f.read(8)
                    if header.startswith(b'PK\x03\x04'):
                        return "xlsx"
                    elif header.startswith(b'\xd0\xcf\x11\xe0'):
                        return "xls"
            except Exception:
                pass
            
            raise InvalidFileException(f"Unable to detect file format for: {filepath}")
    
    def _load_xlsx_file(self, filepath: Path) -> None:
        """Load .xlsx file using openpyxl"""
        try:
            # Suppress warnings for better user experience
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self.workbook = load_workbook(
                    filename=filepath,
                    read_only=True,  # Memory efficient
                    data_only=True,  # Get values instead of formulas
                    keep_vba=False   # Don't load VBA code
                )
        except Exception as e:
            raise InvalidFileException(f"Failed to load .xlsx file: {str(e)}")
    
    def _load_csv_file(self, filepath: Path) -> None:
        """Load .csv file using pandas into a virtual workbook structure."""
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas is required for .csv files")
        try:
            df = pd.read_csv(filepath, keep_default_na=False, dtype=str)
            self.workbook = Workbook()
            if self.workbook.active:
                self.workbook.remove(self.workbook.active) # Remove default sheet
            
            sheet_name = filepath.stem
            ws = self.workbook.create_sheet(title=sheet_name)
            ws.append(list(df.columns))
            for _, row in df.iterrows():
                ws.append(list(row))
            setattr(ws, '_dataframe', df)
        except Exception as e:
            raise InvalidFileException(f"Failed to load .csv file: {str(e)}")
    
    def _load_xls_file(self, filepath: Path) -> None:
        """Load .xls file using xlrd (legacy format)"""
        if not XLRD_AVAILABLE:
            raise ImportError("xlrd is required for .xls files")
        
        try:
            self.workbook = Workbook()
            if self.workbook.active:
                self.workbook.remove(self.workbook.active)
            
            xlrd_book = xlrd.open_workbook(str(filepath))
            
            for sheet_name in xlrd_book.sheet_names():
                sheet = xlrd_book.sheet_by_name(sheet_name)
                ws = self.workbook.create_sheet(title=sheet_name)
                
                # Copy data (limited to prevent memory issues)
                max_rows = min(sheet.nrows, 1000)  # Limit for memory
                max_cols = min(sheet.ncols, 100)
                
                for row_idx in range(max_rows):
                    for col_idx in range(max_cols):
                        try:
                            cell_value = sheet.cell_value(row_idx, col_idx)
                            if cell_value:
                                ws.cell(row=row_idx + 1, column=col_idx + 1, value=str(cell_value))
                        except Exception:
                            continue
                            
        except Exception as e:
            raise InvalidFileException(f"Failed to load .xls file: {str(e)}")
    
    def get_visible_sheets(self) -> List[str]:
        """
        Get list of visible sheet names
        
        Returns:
            List of visible sheet names
        """
        if not self.workbook:
            raise RuntimeError("No workbook loaded. Call load_file() first.")
        
        visible_sheets = []
        
        for sheet_name in self.workbook.sheetnames:
            try:
                worksheet = self.workbook[sheet_name]
                if not worksheet.sheet_state == 'hidden':
                    visible_sheets.append(sheet_name)
            except Exception as e:
                logger.warning(f"Error checking visibility of sheet '{sheet_name}': {e}")
                # Assume visible if we can't determine
                visible_sheets.append(sheet_name)
        
        logger.info(f"Found {len(visible_sheets)} visible sheets out of {len(self.workbook.sheetnames)} total")
        return visible_sheets
    
    def get_sheet_metadata(self, sheet_name: str) -> SheetMetadata:
        """
        Extract comprehensive metadata for a specific sheet
        
        Args:
            sheet_name: Name of the sheet to analyze
            
        Returns:
            SheetMetadata object with comprehensive information
        """
        if not self.workbook:
            raise RuntimeError("No workbook loaded. Call load_file() first.")
        
        # Check cache first
        if sheet_name in self._sheet_metadata_cache:
            return self._sheet_metadata_cache[sheet_name]
        
        try:
            worksheet = self.workbook[sheet_name]
            
            # Get basic dimensions
            max_row = worksheet.max_row
            max_column = worksheet.max_column
            
            # Calculate data boundaries
            first_data_row, last_data_row, first_data_col, last_data_col = self._find_data_boundaries(worksheet)
            
            # Calculate data density
            data_density = self._calculate_data_density(worksheet, first_data_row, last_data_row, 
                                                       first_data_col, last_data_col)
            
            # Count empty rows and columns
            empty_rows = self._count_empty_rows(worksheet, first_data_row, last_data_row, first_data_col, last_data_col)
            empty_cols = self._count_empty_columns(worksheet, first_data_row, last_data_row, first_data_col, last_data_col)
            
            # Estimate file size
            estimated_size = self._estimate_sheet_size(worksheet, last_data_row, last_data_col)
            
            # Create metadata object
            metadata = SheetMetadata(
                name=sheet_name,
                is_visible=worksheet.sheet_state != 'hidden',
                row_count=max_row,
                column_count=max_column,
                data_density=data_density,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                first_data_column=first_data_col,
                last_data_column=last_data_col,
                empty_rows_count=empty_rows,
                empty_columns_count=empty_cols,
                file_format=self.file_format or "unknown",
                estimated_size_mb=estimated_size
            )
            
            # Cache the result
            self._sheet_metadata_cache[sheet_name] = metadata
            
            logger.debug(f"Extracted metadata for sheet '{sheet_name}': "
                        f"{metadata.row_count} rows, {metadata.column_count} columns, "
                        f"density: {metadata.data_density:.2%}")
            
            return metadata
            
        except Exception as e:
            logger.error(f"Failed to extract metadata for sheet '{sheet_name}': {str(e)}")
            raise
    
    def sample_sheet_content(self, sheet_name: str, rows: int = 20) -> ContentSample:
        """
        Sample content from a sheet for analysis
        
        Args:
            sheet_name: Name of the sheet to sample
            rows: Number of rows to sample (default: 20)
            
        Returns:
            ContentSample object with sampled data
        """
        if not self.workbook:
            raise RuntimeError("No workbook loaded. Call load_file() first.")
        
        try:
            worksheet = self.workbook[sheet_name]
            
            # Get metadata to determine sampling range
            metadata = self.get_sheet_metadata(sheet_name)
            
            if metadata.last_data_row < metadata.first_data_row:
                # No data found
                return ContentSample(
                    sheet_name=sheet_name,
                    rows=[],
                    headers=[],
                    sample_size=0,
                    has_data=False
                )
            
            # Determine sampling range
            start_row = metadata.first_data_row
            end_row = min(start_row + rows - 1, metadata.last_data_row)
            
            # Extract headers (first row)
            headers = []
            for col in range(metadata.first_data_column, metadata.last_data_column + 1):
                cell_value = worksheet.cell(row=start_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else "")
            
            # Extract sample rows
            sample_rows = []
            for row in range(start_row + 1, end_row + 1):
                row_data = []
                for col in range(metadata.first_data_column, metadata.last_data_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                sample_rows.append(row_data)
            
            content_sample = ContentSample(
                sheet_name=sheet_name,
                rows=sample_rows,
                headers=headers,
                sample_size=len(sample_rows),
                has_data=len(sample_rows) > 0
            )
            
            logger.debug(f"Sampled {len(sample_rows)} rows from sheet '{sheet_name}'")
            return content_sample
            
        except Exception as e:
            logger.error(f"Failed to sample content from sheet '{sheet_name}': {str(e)}")
            raise
    
    def _find_data_boundaries(self, worksheet: Worksheet) -> Tuple[int, int, int, int]:
        """Find the boundaries of actual data in the worksheet using optimized scanning"""
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if max_row == 0 or max_col == 0:
            return 0, 0, 0, 0
        
        # For very large sheets, use sampling to find boundaries faster
        if max_row > 1000 or max_col > 100:
            return self._find_data_boundaries_sampled(worksheet, max_row, max_col)
        
        # For smaller sheets, use the original method
        return self._find_data_boundaries_full(worksheet, max_row, max_col)
    
    def _find_data_boundaries_sampled(self, worksheet: Worksheet, max_row: int, max_col: int) -> Tuple[int, int, int, int]:
        """Find data boundaries using sampling for large sheets"""
        # Sample every 10th row and column for large sheets
        row_sample = 10
        col_sample = 10
        
        # Find first data row (sample first 1000 rows)
        first_data_row = max_row + 1
        search_rows = min(1000, max_row)
        for row in range(1, search_rows + 1, row_sample):
            for col in range(1, min(max_col + 1, 100), col_sample):
                if worksheet.cell(row=row, column=col).value is not None:
                    first_data_row = min(first_data_row, row)
                    break
        
        # Find last data row (sample last 1000 rows)
        last_data_row = 0
        search_start = max(1, max_row - 1000)
        for row in range(max_row, search_start - 1, -row_sample):
            for col in range(1, min(max_col + 1, 100), col_sample):
                if worksheet.cell(row=row, column=col).value is not None:
                    last_data_row = max(last_data_row, row)
                    break
        
        # Find first data column (sample first 100 columns)
        first_data_col = max_col + 1
        search_cols = min(100, max_col)
        for col in range(1, search_cols + 1, col_sample):
            for row in range(first_data_row, min(last_data_row + 1, first_data_row + 100), row_sample):
                if worksheet.cell(row=row, column=col).value is not None:
                    first_data_col = min(first_data_col, col)
                    break
        
        # Find last data column (sample last 100 columns)
        last_data_col = 0
        search_start = max(1, max_col - 100)
        for col in range(max_col, search_start - 1, -col_sample):
            for row in range(first_data_row, min(last_data_row + 1, first_data_row + 100), row_sample):
                if worksheet.cell(row=row, column=col).value is not None:
                    last_data_col = max(last_data_col, col)
                    break
        
        return first_data_row, last_data_row, first_data_col, last_data_col
    
    def _find_data_boundaries_full(self, worksheet: Worksheet, max_row: int, max_col: int) -> Tuple[int, int, int, int]:
        """Find data boundaries using full scanning for smaller sheets"""
        # Find first data row
        first_data_row = max_row + 1
        for row in range(1, min(max_row + 1, 1000)):  # Limit search for performance
            for col in range(1, min(max_col + 1, 100)):
                if worksheet.cell(row=row, column=col).value is not None:
                    first_data_row = min(first_data_row, row)
                    break
        
        # Find last data row
        last_data_row = 0
        for row in range(max_row, max(1, first_data_row - 1), -1):
            for col in range(1, min(max_col + 1, 100)):
                if worksheet.cell(row=row, column=col).value is not None:
                    last_data_row = max(last_data_row, row)
                    break
        
        # Find first data column
        first_data_col = max_col + 1
        for col in range(1, min(max_col + 1, 100)):
            for row in range(first_data_row, min(last_data_row + 1, first_data_row + 100)):
                if worksheet.cell(row=row, column=col).value is not None:
                    first_data_col = min(first_data_col, col)
                    break
        
        # Find last data column
        last_data_col = 0
        for col in range(max_col, max(1, first_data_col - 1), -1):
            for row in range(first_data_row, min(last_data_row + 1, first_data_row + 100)):
                if worksheet.cell(row=row, column=col).value is not None:
                    last_data_col = max(last_data_col, col)
                    break
        
        return first_data_row, last_data_row, first_data_col, last_data_col
    
    def _calculate_data_density(self, worksheet: Worksheet, first_row: int, last_row: int, 
                               first_col: int, last_col: int) -> float:
        """Calculate the percentage of cells that contain data"""
        if first_row > last_row or first_col > last_col:
            return 0.0
        
        total_cells = (last_row - first_row + 1) * (last_col - first_col + 1)
        if total_cells == 0:
            return 0.0
        
        filled_cells = 0
        # Sample for performance (check every 10th cell)
        sample_factor = 10
        
        for row in range(first_row, last_row + 1, sample_factor):
            for col in range(first_col, last_col + 1, sample_factor):
                if worksheet.cell(row=row, column=col).value is not None:
                    filled_cells += 1
        
        # Extrapolate to full range
        sampled_cells = ((last_row - first_row + 1) // sample_factor + 1) * \
                       ((last_col - first_col + 1) // sample_factor + 1)
        
        if sampled_cells == 0:
            return 0.0
        
        density = (filled_cells / sampled_cells)
        return min(density, 1.0)  # Cap at 100%
    
    def _count_empty_rows(self, worksheet: Worksheet, first_row: int, last_row: int, 
                          first_col: int, last_col: int) -> int:
        """Count completely empty rows in the data range"""
        empty_rows = 0
        
        for row in range(first_row, last_row + 1):
            is_empty = True
            for col in range(first_col, last_col + 1):
                if worksheet.cell(row=row, column=col).value is not None:
                    is_empty = False
                    break
            if is_empty:
                empty_rows += 1
        
        return empty_rows
    
    def _count_empty_columns(self, worksheet: Worksheet, first_row: int, last_row: int, 
                            first_col: int, last_col: int) -> int:
        """Count completely empty columns in the data range"""
        empty_cols = 0
        
        for col in range(first_col, last_col + 1):
            is_empty = True
            for row in range(first_row, last_row + 1):
                if worksheet.cell(row=row, column=col).value is not None:
                    is_empty = False
                    break
            if is_empty:
                empty_cols += 1
        
        return empty_cols
    
    def _estimate_sheet_size(self, worksheet: Worksheet, last_row: int, last_col: int) -> float:
        """Estimate the size of the sheet in MB"""
        # Rough estimation: assume average cell size of 50 bytes
        total_cells = last_row * last_col
        estimated_bytes = total_cells * 50
        return estimated_bytes / (1024 * 1024)  # Convert to MB
    
    def get_sheet_data(self, sheet_name: str, max_rows: int = 1000) -> List[List[str]]:
        """
        Get all data from a specific sheet as a list of rows.

        Handles both standard Excel sheets and virtual CSV sheets.
        
        Args:
            sheet_name: Name of the sheet to get data from
            max_rows: Maximum number of rows to retrieve
            
        Returns:
            List of rows, where each row is a list of cell values
        """
        if not self.workbook:
            raise RuntimeError("No workbook loaded. Call load_file() first.")
        
        try:
            worksheet = self.workbook[sheet_name]

            # If it's a CSV, the data is already loaded in rows
            if self.file_format == 'csv':
                # The data is already in the openpyxl worksheet structure
                data = []
                for row in worksheet.iter_rows(max_row=max_rows):
                    data.append([cell.value for cell in row])
                return data

            # For regular excel files, iterate and extract
            data = []
            # Use iter_rows for memory efficiency
            for row in worksheet.iter_rows(max_row=max_rows):
                # Convert cells to string, handling None
                data.append([str(cell.value) if cell.value is not None else "" for cell in row])
                
            return data
            
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
        except Exception as e:
            logger.error(f"Failed to get data for sheet '{sheet_name}': {e}")
            return []
    
    def get_all_sheets_data(self, max_rows: int = 1000) -> Dict[str, List[List[str]]]:
        """
        Get data from all visible sheets
        
        Args:
            max_rows: Maximum number of rows to extract per sheet
            
        Returns:
            Dictionary mapping sheet names to their data
        """
        visible_sheets = self.get_visible_sheets()
        sheets_data = {}
        
        for sheet_name in visible_sheets:
            try:
                sheets_data[sheet_name] = self.get_sheet_data(sheet_name, max_rows)
            except Exception as e:
                logger.error(f"Failed to get data for sheet '{sheet_name}': {e}")
                sheets_data[sheet_name] = []
        
        return sheets_data
    
    def get_file_info(self) -> Dict[str, Any]:
        """
        Get basic file information
        
        Returns:
            Dictionary with file information
        """
        if not self.workbook:
            raise RuntimeError("No workbook loaded. Call load_file() first.")
        
        return {
            "file_path": str(self.file_path) if self.file_path else None,
            "file_format": self.file_format,
            "total_sheets": len(self.workbook.sheetnames),
            "visible_sheets": self.get_visible_sheets(),
            "file_size_mb": self.file_path.stat().st_size / (1024 * 1024) if self.file_path else 0
        }
    
    def get_all_sheets_metadata(self) -> Dict[str, SheetMetadata]:
        """
        Get metadata for all visible sheets
        
        Returns:
            Dictionary mapping sheet names to their metadata
        """
        visible_sheets = self.get_visible_sheets()
        metadata_dict = {}
        
        for sheet_name in visible_sheets:
            try:
                metadata_dict[sheet_name] = self.get_sheet_metadata(sheet_name)
            except Exception as e:
                logger.error(f"Failed to get metadata for sheet '{sheet_name}': {e}")
        
        return metadata_dict
    
    def close(self) -> None:
        """Close the workbook and clean up resources"""
        self._cleanup()
    
    def _cleanup(self) -> None:
        """Clean up resources"""
        if self.workbook:
            try:
                self.workbook.close()
            except Exception:
                pass
            finally:
                self.workbook = None
        
        self.file_path = None
        self.file_format = None
        self._sheet_metadata_cache.clear()
    
    def __enter__(self):
        """Context manager entry"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.close()


# Convenience function for quick file analysis
def analyze_excel_file(filepath: Union[str, Path], sample_rows: int = 20) -> Dict[str, Any]:
    """
    Quick analysis of an Excel file
    
    Args:
        filepath: Path to the Excel file
        sample_rows: Number of rows to sample per sheet
        
    Returns:
        Dictionary with analysis results
    """
    with ExcelProcessor() as processor:
        if not processor.load_file(filepath):
            return {"error": "Failed to load file"}
        
        visible_sheets = processor.get_visible_sheets()
        all_metadata = processor.get_all_sheets_metadata()
        
        # Sample content from first sheet
        first_sheet_sample = None
        if visible_sheets:
            try:
                first_sheet_sample = processor.sample_sheet_content(visible_sheets[0], sample_rows)
            except Exception as e:
                logger.warning(f"Failed to sample first sheet: {e}")
        
        return {
            "file_path": str(filepath),
            "file_format": processor.file_format,
            "visible_sheets": visible_sheets,
            "total_sheets": len(processor.workbook.sheetnames) if processor.workbook else 0,
            "sheets_metadata": {
                name: {
                    "row_count": meta.row_count,
                    "column_count": meta.column_count,
                    "data_density": meta.data_density,
                    "estimated_size_mb": meta.estimated_size_mb,
                    "is_visible": meta.is_visible
                }
                for name, meta in all_metadata.items()
            },
            "first_sheet_sample": {
                "headers": first_sheet_sample.headers if first_sheet_sample else [],
                "sample_rows": first_sheet_sample.rows if first_sheet_sample else [],
                "has_data": first_sheet_sample.has_data if first_sheet_sample else False
            } if first_sheet_sample else None
        } 