"""
Manual Categorizer for BOQ Tools
Generates Excel files for manual categorization of unmatched descriptions
"""

import logging
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import shutil
import tempfile

from core.auto_categorizer import UnmatchedDescription
from core.category_dictionary import CategoryDictionary

logger = logging.getLogger(__name__)


def generate_manual_categorization_excel(review_descriptions: List,
                                        available_categories: List[str],
                                        output_dir: Optional[Path] = None) -> Path:
    """
    Generate Excel file for manual categorization of descriptions that need review
    (unmatched, fuzzy, and partial matches)
    
    Args:
        review_descriptions: List of ManualReviewDescription or UnmatchedDescription objects
        available_categories: List of available categories for dropdown
        output_dir: Output directory (default: current directory)
        
    Returns:
        Path to the created Excel file
    """
    logger.info(f"Generating manual categorization Excel file for {len(review_descriptions)} descriptions")
    
    # Create output directory if it doesn't exist
    if output_dir is None:
        output_dir = Path.cwd()
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"manual_categorization_{timestamp}.xlsx"
    filepath = output_dir / filename
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet:
        wb.remove(default_sheet)
    
    # Create main categorization sheet
    ws_categorize = wb.create_sheet("Categorization", 0)
    
    # Create instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 1)
    
    # Set up the main categorization sheet
    _setup_categorization_sheet(ws_categorize, review_descriptions, available_categories)
    
    # Set up the instructions sheet
    _setup_instructions_sheet(ws_instructions, available_categories)
    
    # Save the workbook
    wb.save(filepath)
    logger.info(f"Manual categorization Excel file created: {filepath}")
    print(f"[DEBUG] Manual categorization Excel created at: {filepath}, exists: {filepath.exists()}")
    
    return filepath


def _setup_categorization_sheet(worksheet, review_descriptions: List, 
                               available_categories: List[str]):
    """Set up the main categorization worksheet"""
    
    # If there are no descriptions to review, show a message and return early
    if not review_descriptions:
        worksheet['A2'] = "No descriptions to categorize."
        return
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set up headers - include auto-category and match type for ManualReviewDescription objects
    headers = ["Description", "Source_Sheet", "Frequency", "Auto_Category", "Match_Type", "Confidence", "Category", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, desc in enumerate(review_descriptions, 2):
        worksheet.cell(row=row, column=1, value=desc.description)
        worksheet.cell(row=row, column=2, value=desc.source_sheet_name)
        worksheet.cell(row=row, column=3, value=desc.frequency)
        
        # Handle both UnmatchedDescription and ManualReviewDescription objects
        if hasattr(desc, 'auto_category'):
            # ManualReviewDescription object
            worksheet.cell(row=row, column=4, value=desc.auto_category or "")
            worksheet.cell(row=row, column=5, value=desc.match_type)
            worksheet.cell(row=row, column=6, value=f"{desc.confidence:.2f}" if desc.confidence > 0 else "")
            worksheet.cell(row=row, column=7, value=desc.auto_category or "")  # Pre-fill with auto-category
            worksheet.cell(row=row, column=8, value="")  # Notes (to be filled manually)
        else:
            # UnmatchedDescription object (backward compatibility)
            worksheet.cell(row=row, column=4, value="")  # Auto_Category
            worksheet.cell(row=row, column=5, value="none")  # Match_Type
            worksheet.cell(row=row, column=6, value="")  # Confidence
            worksheet.cell(row=row, column=7, value="")  # Category (to be filled manually)
            worksheet.cell(row=row, column=8, value="")  # Notes (to be filled manually)
        
        # Apply borders to all cells
        for col in range(1, 9):
            worksheet.cell(row=row, column=col).border = border
    
    # Set up data validation for Category column
    category_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(get_manual_categorization_categories())}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Category",
        error="Please select a category from the dropdown list.",
        showInputMessage=True,
        promptTitle="Category Selection",
        prompt="Select a category from the dropdown list."
    )
    worksheet.add_data_validation(category_validation)
    
    # Apply validation to Category column (column G)
    category_validation.add(f'G2:G{len(review_descriptions) + 1}')
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 60  # Description
    worksheet.column_dimensions['B'].width = 20  # Source_Sheet
    worksheet.column_dimensions['C'].width = 12  # Frequency
    worksheet.column_dimensions['D'].width = 25  # Auto_Category
    worksheet.column_dimensions['E'].width = 15  # Match_Type
    worksheet.column_dimensions['F'].width = 12  # Confidence
    worksheet.column_dimensions['G'].width = 25  # Category
    worksheet.column_dimensions['H'].width = 30  # Notes
    
    # Add conditional formatting for frequency
    from openpyxl.formatting.rule import ColorScaleRule
    frequency_rule = ColorScaleRule(
        start_type='min',
        start_color='FFFFFF',
        end_type='max',
        end_color='FF6B6B'
    )
    worksheet.conditional_formatting.add(f'C2:C{len(review_descriptions) + 1}', frequency_rule)
    
    # Add conditional formatting for confidence
    confidence_rule = ColorScaleRule(
        start_type='min',
        start_color='FF6B6B',  # Red for low confidence
        end_type='max',
        end_color='4CAF50'     # Green for high confidence
    )
    worksheet.conditional_formatting.add(f'F2:F{len(review_descriptions) + 1}', confidence_rule)
    
    # Freeze the header row
    worksheet.freeze_panes = "A2"


def _setup_instructions_sheet(worksheet, available_categories: List[str]):
    """Set up the instructions worksheet"""
    
    # Title
    title_cell = worksheet['A1']
    title_cell.value = "Manual Categorization Instructions"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Instructions
    instructions = [
        ("Purpose", "This file contains descriptions that could not be automatically categorized. Please assign appropriate categories to each description."),
        ("", ""),
        ("How to Use", "1. Go to the 'Categorization' sheet"),
        ("", "2. For each description, select a category from the dropdown in the 'Category' column"),
        ("", "3. Optionally add notes in the 'Notes' column"),
        ("", "4. Save the file when complete"),
        ("", ""),
        ("Available Categories", f"The following categories are available: {', '.join(get_manual_categorization_categories())}"),
        ("", ""),
        ("Tips", "- Descriptions are sorted by frequency (most frequent first)"),
        ("", "- Use the 'Notes' column to record any special considerations"),
        ("", "- If a description doesn't fit any category, leave it blank"),
        ("", "- You can add new categories by editing the dropdown list"),
        ("", ""),
        ("File Information", f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
        ("", f"Total descriptions to categorize: {len(get_manual_categorization_categories())}")
    ]
    
    # Add instructions to worksheet
    for row, (title, content) in enumerate(instructions, 3):
        if title:
            cell = worksheet.cell(row=row, column=1, value=title)
            cell.font = Font(bold=True)
        if content:
            worksheet.cell(row=row, column=2, value=content)
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 80


def load_manual_categorization_results(filepath: Path) -> List[dict]:
    """
    Load manual categorization results from Excel file
    
    Args:
        filepath: Path to the Excel file with manual categorizations
        
    Returns:
        List of dictionaries with categorization results
    """
    logger.info(f"Loading manual categorization results from {filepath}")
    
    if not filepath.exists():
        raise FileNotFoundError(f"Manual categorization file not found: {filepath}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["Categorization"]
        
        results = []
        
        # Read data starting from row 2 (skip header)
        for row in range(2, ws.max_row + 1):
            description = ws.cell(row=row, column=1).value
            source_sheet = ws.cell(row=row, column=2).value
            frequency = ws.cell(row=row, column=3).value
            category = ws.cell(row=row, column=4).value
            notes = ws.cell(row=row, column=5).value
            
            if description and category:  # Only include rows with both description and category
                # Safely convert frequency to int
                freq_value = 1
                if frequency is not None:
                    try:
                        freq_value = int(float(str(frequency)))
                    except (ValueError, TypeError):
                        freq_value = 1
                
                results.append({
                    'description': str(description).strip(),
                    'source_sheet': str(source_sheet).strip() if source_sheet else 'Unknown',
                    'frequency': freq_value,
                    'category': str(category).strip(),
                    'notes': str(notes).strip() if notes else ''
                })
        
        logger.info(f"Loaded {len(results)} manual categorizations")
        return results
        
    except Exception as e:
        logger.error(f"Error loading manual categorization results: {e}")
        raise


def apply_manual_categorizations(dataframe, manual_results: List[dict], 
                                description_column: str = 'Description',
                                category_column: str = 'Category'):
    """
    Apply manual categorizations to the DataFrame
    
    Args:
        dataframe: DataFrame to update
        manual_results: List of manual categorization results
        description_column: Name of the description column
        category_column: Name of the category column
        
    Returns:
        Updated DataFrame with manual categorizations applied
    """
    logger.info(f"Applying {len(manual_results)} manual categorizations to DataFrame")
    
    # Create a copy to avoid modifying the original
    df = dataframe.copy()
    
    # Create a mapping from description to category
    manual_mapping = {}
    for result in manual_results:
        description = result['description'].lower()
        category = result['category']
        manual_mapping[description] = category
    
    # Apply manual categorizations
    updated_count = 0
    for index, row in df.iterrows():
        description = str(row[description_column]).strip().lower()
        
        if description in manual_mapping:
            df.at[index, category_column] = manual_mapping[description]
            updated_count += 1
    
    logger.info(f"Updated {updated_count} rows with manual categorizations")
    return df


def create_categorization_summary(dataframe, category_column: str = 'Category') -> dict:
    """
    Create a summary of categorization results
    
    Args:
        dataframe: Categorized DataFrame
        category_column: Name of the category column
        
    Returns:
        Dictionary with categorization summary
    """
    total_rows = len(dataframe)
    categorized_rows = len(dataframe[dataframe[category_column].notna() & (dataframe[category_column] != '')])
    uncategorized_rows = total_rows - categorized_rows
    
    category_distribution = dataframe[category_column].value_counts().to_dict()
    
    # Remove empty categories from distribution
    if '' in category_distribution:
        del category_distribution['']
    
    summary = {
        'total_rows': total_rows,
        'categorized_rows': categorized_rows,
        'uncategorized_rows': uncategorized_rows,
        'categorization_rate': categorized_rows / total_rows if total_rows > 0 else 0.0,
        'category_distribution': category_distribution,
        'unique_categories': len(category_distribution)
    }
    
    return summary 


def process_manual_categorizations(excel_filepath: Path, 
                                  description_column: str = "Description",
                                  category_column: str = "Category",
                                  source_sheet_column: str = "Source_Sheet",
                                  notes_column: str = "Notes") -> Dict[str, str]:
    """
    Process user-completed manual categorization Excel file
    
    Args:
        excel_filepath: Path to the completed Excel file
        description_column: Name of the description column
        category_column: Name of the category column
        source_sheet_column: Name of the source sheet column
        notes_column: Name of the notes column
        
    Returns:
        Dictionary mapping descriptions to their manually assigned categories
        
    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the file is corrupted or missing required columns
        KeyError: If required data is missing
    """
    logger.info(f"Processing manual categorizations from {excel_filepath}")
    
    # Validate file exists
    if not excel_filepath.exists():
        raise FileNotFoundError(f"Manual categorization file not found: {excel_filepath}")
    
    try:
        # Read the Excel file using pandas
        df: pd.DataFrame = pd.read_excel(excel_filepath, sheet_name="Categorization")
        print(f"[DEBUG] Loaded manual categorization file columns: {list(df.columns)}")
        print(f"[DEBUG] First few rows:\n{df.head()}")
        logger.info(f"Successfully loaded Excel file with {len(df)} rows")
        
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        raise ValueError(f"Failed to read Excel file {excel_filepath}: {e}")
    
    # Validate required columns exist
    required_columns = [description_column, category_column, source_sheet_column]
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        available_columns = list(df.columns)
        logger.error(f"Missing required columns: {missing_columns}")
        logger.error(f"Available columns: {available_columns}")
        raise ValueError(f"Missing required columns: {missing_columns}. Available columns: {available_columns}")
    
    # Clean and validate data
    logger.info("Cleaning and validating data...")
    
    # Remove rows where description is empty or null
    initial_count = len(df)
    df = df.dropna(subset=[description_column])
    df = df[df[description_column].astype(str).str.strip() != '']
    
    if len(df) < initial_count:
        logger.warning(f"Removed {initial_count - len(df)} rows with empty descriptions")
    
    # Clean description column
    df[description_column] = df[description_column].astype(str).str.strip()
    
    # Clean category column and filter for rows with categories
    df[category_column] = df[category_column].astype(str).str.strip()
    df = df[df[category_column] != '']
    df = df[df[category_column] != 'nan']
    
    if len(df) == 0:
        logger.warning("No valid categorizations found in the file")
        return {}
    
    # Clean source sheet column
    if source_sheet_column in df.columns:
        df[source_sheet_column] = df[source_sheet_column].astype(str).str.strip()
        df[source_sheet_column] = df[source_sheet_column].replace('nan', 'Unknown')
    
    # Clean notes column
    if notes_column in df.columns:
        df[notes_column] = df[notes_column].astype(str).str.strip()
        df[notes_column] = df[notes_column].replace('nan', '')
    
    # Create mapping dictionary
    categorization_mapping = {}
    duplicate_descriptions = []
    
    for index, row in df.iterrows():
        description = str(row[description_column]).lower()  # Normalize to lowercase
        category = str(row[category_column])
        
        # Check for duplicates
        if description in categorization_mapping:
            duplicate_descriptions.append(description)
            logger.warning(f"Duplicate description found: '{description}' with categories '{categorization_mapping[description]}' and '{category}'")
            # Keep the last occurrence (most recent)
        
        categorization_mapping[description] = category
    
    # Log summary
    logger.info(f"Successfully processed {len(categorization_mapping)} manual categorizations")
    if duplicate_descriptions:
        logger.warning(f"Found {len(set(duplicate_descriptions))} duplicate descriptions")
    
    # Validate mapping quality
    _validate_categorization_mapping(categorization_mapping)
    
    return categorization_mapping


def _validate_categorization_mapping(mapping: Dict[str, str]) -> None:
    """
    Validate the categorization mapping for quality issues
    
    Args:
        mapping: Dictionary mapping descriptions to categories
        
    Raises:
        ValueError: If validation fails
    """
    if not mapping:
        logger.warning("Empty categorization mapping")
        return
    
    # Check for empty or invalid categories
    invalid_categories = []
    for description, category in mapping.items():
        if not category or category.strip() == '':
            invalid_categories.append(description)
    
    if invalid_categories:
        logger.warning(f"Found {len(invalid_categories)} descriptions with empty categories")
        for desc in invalid_categories[:5]:  # Show first 5
            logger.warning(f"  Empty category for: '{desc}'")
    
    # Check for very short descriptions (potential data quality issues)
    short_descriptions = [desc for desc in mapping.keys() if len(desc.strip()) < 3]
    if short_descriptions:
        logger.warning(f"Found {len(short_descriptions)} very short descriptions (< 3 chars)")
        for desc in short_descriptions[:5]:  # Show first 5
            logger.warning(f"  Very short description: '{desc}'")
    
    # Check category distribution
    category_counts = {}
    for category in mapping.values():
        category_counts[category] = category_counts.get(category, 0) + 1
    
    logger.info(f"Category distribution: {category_counts}")
    
    # Warn if any category has very few items (potential typos)
    for category, count in category_counts.items():
        if count == 1:
            logger.info(f"Category '{category}' has only 1 item - verify spelling")


def validate_excel_file_structure(filepath: Path) -> Dict[str, Any]:
    """
    Validate the structure of a manual categorization Excel file
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        Dictionary with validation results
    """
    logger.info(f"Validating Excel file structure: {filepath}")
    
    validation_result = {
        'is_valid': False,
        'errors': [],
        'warnings': [],
        'file_info': {},
        'sheet_info': {}
    }
    
    try:
        # Check if file exists
        if not filepath.exists():
            validation_result['errors'].append(f"File not found: {filepath}")
            return validation_result
        
        # Load workbook
        wb = openpyxl.load_workbook(filepath, read_only=True)
        
        # Check required sheets
        required_sheets = ["Categorization", "Instructions"]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
        
        if missing_sheets:
            validation_result['errors'].append(f"Missing required sheets: {missing_sheets}")
            validation_result['errors'].append(f"Available sheets: {wb.sheetnames}")
        else:
            validation_result['sheet_info']['available_sheets'] = wb.sheetnames
        
        # Check Categorization sheet structure
        if "Categorization" in wb.sheetnames:
            ws = wb["Categorization"]
            
            # Check headers
            expected_headers = ["Description", "Source_Sheet", "Frequency", "Auto_Category", "Match_Type", "Confidence", "Category", "Notes"]
            actual_headers = []
            
            for col in range(1, 9):  # First 8 columns
                cell_value = ws.cell(row=1, column=col).value
                actual_headers.append(str(cell_value) if cell_value else "")
            
            missing_headers = [h for h in expected_headers if h not in actual_headers]
            if missing_headers:
                validation_result['errors'].append(f"Missing headers: {missing_headers}")
                validation_result['warnings'].append(f"Actual headers: {actual_headers}")
            else:
                validation_result['sheet_info']['headers'] = actual_headers
            
            # Check data rows
            data_row_count = 0
            for row in range(2, ws.max_row + 1):
                description = ws.cell(row=row, column=1).value
                if description and str(description).strip():
                    data_row_count += 1
            
            validation_result['sheet_info']['data_rows'] = data_row_count
            
            if data_row_count == 0:
                validation_result['warnings'].append("No data rows found in Categorization sheet")
        
        # File information
        validation_result['file_info'] = {
            'file_size': filepath.stat().st_size,
            'last_modified': datetime.fromtimestamp(filepath.stat().st_mtime),
            'file_path': str(filepath)
        }
        
        # Determine if file is valid
        validation_result['is_valid'] = len(validation_result['errors']) == 0
        
        logger.info(f"Validation completed. Valid: {validation_result['is_valid']}")
        if validation_result['errors']:
            logger.error(f"Validation errors: {validation_result['errors']}")
        if validation_result['warnings']:
            logger.warning(f"Validation warnings: {validation_result['warnings']}")
        
        return validation_result
        
    except Exception as e:
        validation_result['errors'].append(f"Error during validation: {e}")
        logger.error(f"Validation error: {e}")
        return validation_result


def get_categorization_statistics(mapping: Dict[str, str]) -> Dict[str, Any]:
    """
    Generate statistics about manual categorizations
    
    Args:
        mapping: Dictionary mapping descriptions to categories
        
    Returns:
        Dictionary with statistics
    """
    if not mapping:
        return {
            'total_categorizations': 0,
            'unique_categories': 0,
            'category_distribution': {},
            'average_description_length': 0,
            'shortest_description': '',
            'longest_description': ''
        }
    
    # Basic counts
    total_categorizations = len(mapping)
    unique_categories = len(set(mapping.values()))
    
    # Category distribution
    category_distribution = {}
    for category in mapping.values():
        category_distribution[category] = category_distribution.get(category, 0) + 1
    
    # Description length statistics
    description_lengths = [len(desc) for desc in mapping.keys()]
    avg_length = sum(description_lengths) / len(description_lengths)
    
    shortest_desc = min(mapping.keys(), key=len)
    longest_desc = max(mapping.keys(), key=len)
    
    return {
        'total_categorizations': total_categorizations,
        'unique_categories': unique_categories,
        'category_distribution': category_distribution,
        'average_description_length': round(avg_length, 2),
        'shortest_description': shortest_desc,
        'longest_description': longest_desc,
        'description_length_range': f"{min(description_lengths)} - {max(description_lengths)}"
    }


def apply_manual_categories(dataframe: pd.DataFrame,
                           manual_categorizations: Dict[str, str],
                           description_column: str = 'Description',
                           category_column: str = 'Category',
                           case_sensitive: bool = False) -> Dict[str, Any]:
    """
    Apply manual categorizations to the main DataFrame
    
    Args:
        dataframe: Main DataFrame to update
        manual_categorizations: Dictionary mapping descriptions to categories
        description_column: Name of the description column
        category_column: Name of the category column
        case_sensitive: Whether to perform case-sensitive matching
        
    Returns:
        Dictionary containing:
        - 'updated_dataframe': DataFrame with manual categorizations applied
        - 'statistics': Categorization statistics
        - 'updated_count': Number of rows updated
        - 'remaining_unmatched': List of descriptions still unmatched
        - 'coverage_rate': Percentage of rows now categorized
    """
    logger.info(f"Applying {len(manual_categorizations)} manual categorizations to DataFrame")
    
    # Validate inputs
    if description_column not in dataframe.columns:
        raise ValueError(f"Description column '{description_column}' not found in DataFrame")
    
    if category_column not in dataframe.columns:
        logger.warning(f"Category column '{category_column}' not found, creating it")
        dataframe[category_column] = ''
    
    # Create a copy to avoid modifying the original
    df = dataframe.copy()
    
    # Initialize tracking variables
    initial_unmatched_count = 0
    updated_count = 0
    exact_matches = 0
    case_insensitive_matches = 0
    remaining_unmatched = []
    
    # Count initial unmatched rows
    unmatched_mask = df[category_column].isna() | (df[category_column] == '') | (df[category_column].isnull())
    initial_unmatched_count = unmatched_mask.sum()
    
    logger.info(f"Initial unmatched rows: {initial_unmatched_count}")
    
    # Process each row
    for index, row in df.iterrows():
        description = str(row[description_column]).strip()
        
        # Skip empty descriptions
        if not description or description.lower() in ['nan', 'none', '']:
            continue
        
        # Check if this row is currently unmatched
        current_category = str(row[category_column]).strip()
        if current_category and current_category.lower() not in ['nan', 'none', '']:
            continue  # Already categorized
        
        # Try to find a match in manual categorizations
        matched_category = None
        
        if case_sensitive:
            # Case-sensitive matching
            if description in manual_categorizations:
                matched_category = manual_categorizations[description]
                exact_matches += 1
                logger.debug(f"Exact case-sensitive match: '{description}' → '{matched_category}'")
        else:
            # Case-insensitive matching
            description_lower = description.lower()
            
            # First try exact case-insensitive match
            if description_lower in manual_categorizations:
                matched_category = manual_categorizations[description_lower]
                exact_matches += 1
                logger.debug(f"Exact case-insensitive match: '{description}' → '{matched_category}'")
            else:
                # Try to find a match by comparing normalized descriptions
                for manual_desc, category in manual_categorizations.items():
                    if description_lower == manual_desc.lower():
                        matched_category = category
                        case_insensitive_matches += 1
                        logger.debug(f"Case-insensitive match: '{description}' → '{matched_category}' (manual: '{manual_desc}')")
                        break
        
        # Apply the category if found
        if matched_category:
            df.at[index, category_column] = matched_category
            updated_count += 1
            logger.debug(f"Updated row {index}: '{description[:50]}...' → '{matched_category}'")
        else:
            # Track remaining unmatched descriptions
            if description not in remaining_unmatched:
                remaining_unmatched.append(description)
    
    # Calculate final statistics
    final_unmatched_mask = df[category_column].isna() | (df[category_column] == '') | (df[category_column].isnull())
    final_unmatched_count = final_unmatched_mask.sum()
    
    total_rows = len(df)
    categorized_rows = total_rows - final_unmatched_count
    coverage_rate = categorized_rows / total_rows if total_rows > 0 else 0.0
    
    # Create statistics
    statistics = {
        'total_rows': total_rows,
        'initial_unmatched': initial_unmatched_count,
        'final_unmatched': final_unmatched_count,
        'rows_updated': updated_count,
        'categorized_rows': categorized_rows,
        'coverage_rate': coverage_rate,
        'exact_matches': exact_matches,
        'case_insensitive_matches': case_insensitive_matches,
        'manual_categorizations_applied': len(manual_categorizations),
        'remaining_unmatched_count': len(remaining_unmatched)
    }
    
    # Log results
    logger.info(f"Manual categorization application completed:")
    logger.info(f"  Total rows: {total_rows}")
    logger.info(f"  Initial unmatched: {initial_unmatched_count}")
    logger.info(f"  Rows updated: {updated_count}")
    logger.info(f"  Final unmatched: {final_unmatched_count}")
    logger.info(f"  Coverage rate: {coverage_rate:.1%}")
    logger.info(f"  Exact matches: {exact_matches}")
    logger.info(f"  Case-insensitive matches: {case_insensitive_matches}")
    
    if remaining_unmatched:
        logger.info(f"  Remaining unmatched descriptions: {len(remaining_unmatched)}")
        logger.info("  Sample remaining unmatched:")
        for desc in remaining_unmatched[:5]:
            logger.info(f"    - '{desc[:60]}...'")
    
    # Create result dictionary
    result = {
        'updated_dataframe': df,
        'statistics': statistics,
        'updated_count': updated_count,
        'remaining_unmatched': remaining_unmatched,
        'coverage_rate': coverage_rate
    }
    
    return result


def get_categorization_coverage_report(dataframe: pd.DataFrame,
                                      category_column: str = 'Category',
                                      description_column: str = 'Description') -> Dict[str, Any]:
    """
    Generate a comprehensive categorization coverage report
    
    Args:
        dataframe: DataFrame to analyze
        category_column: Name of the category column
        description_column: Name of the description column
        
    Returns:
        Dictionary with detailed coverage statistics
    """
    logger.info(f"Generating categorization coverage report for {len(dataframe)} rows")
    
    total_rows = len(dataframe)
    
    # Count categorized vs uncategorized rows
    categorized_mask = ~(dataframe[category_column].isna() | (dataframe[category_column] == '') | (dataframe[category_column].isnull()))
    categorized_count = categorized_mask.sum()
    uncategorized_count = total_rows - categorized_count
    
    # Get category distribution
    category_distribution = {}
    if categorized_count > 0:
        category_distribution = dataframe[category_column].value_counts().to_dict()
    
    # Analyze uncategorized descriptions
    uncategorized_df = dataframe[~categorized_mask]
    unique_uncategorized = []
    if len(uncategorized_df) > 0:
        unique_uncategorized = uncategorized_df[description_column].dropna().unique().tolist()
    
    # Calculate statistics
    coverage_rate = categorized_count / total_rows if total_rows > 0 else 0.0
    unique_categories = len(category_distribution)
    
    # Find most common categories
    top_categories = []
    if category_distribution:
        sorted_categories = sorted(category_distribution.items(), key=lambda x: x[1], reverse=True)
        top_categories = sorted_categories[:10]  # Top 10 categories
    
    # Find categories with only one item (potential issues)
    single_item_categories = [cat for cat, count in category_distribution.items() if count == 1]
    
    report = {
        'summary': {
            'total_rows': total_rows,
            'categorized_rows': categorized_count,
            'uncategorized_rows': uncategorized_count,
            'coverage_rate': coverage_rate,
            'unique_categories': unique_categories
        },
        'category_distribution': category_distribution,
        'top_categories': top_categories,
        'single_item_categories': single_item_categories,
        'uncategorized_analysis': {
            'total_uncategorized': uncategorized_count,
            'unique_uncategorized_descriptions': len(unique_uncategorized),
            'sample_uncategorized': unique_uncategorized[:10]  # First 10
        },
        'quality_metrics': {
            'categories_with_single_item': len(single_item_categories),
            'average_items_per_category': categorized_count / unique_categories if unique_categories > 0 else 0,
            'most_common_category_count': max(category_distribution.values()) if category_distribution else 0
        }
    }
    
    # Log report summary
    logger.info(f"Categorization coverage report:")
    logger.info(f"  Coverage: {coverage_rate:.1%} ({categorized_count}/{total_rows})")
    logger.info(f"  Unique categories: {unique_categories}")
    logger.info(f"  Uncategorized: {uncategorized_count}")
    logger.info(f"  Single-item categories: {len(single_item_categories)}")
    
    return report


def export_categorization_report(dataframe: pd.DataFrame,
                                report_data: Dict[str, Any],
                                output_path: Path,
                                include_samples: bool = True) -> bool:
    """
    Export categorization report to Excel file
    
    Args:
        dataframe: Categorized DataFrame
        report_data: Report data from get_categorization_coverage_report
        output_path: Output file path
        include_samples: Whether to include sample data in the report
        
    Returns:
        True if export was successful
    """
    logger.info(f"Exporting categorization report to {output_path}")
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Add summary data
        summary = report_data['summary']
        ws_summary['A1'] = "Categorization Coverage Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        summary_data = [
            ("Total Rows", summary['total_rows']),
            ("Categorized Rows", summary['categorized_rows']),
            ("Uncategorized Rows", summary['uncategorized_rows']),
            ("Coverage Rate", f"{summary['coverage_rate']:.1%}"),
            ("Unique Categories", summary['unique_categories'])
        ]
        
        for row, (label, value) in enumerate(summary_data, 3):
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
        
        # Category distribution sheet
        ws_distribution = wb.create_sheet("Category Distribution", 1)
        ws_distribution['A1'] = "Category"
        ws_distribution['B1'] = "Count"
        ws_distribution['C1'] = "Percentage"
        ws_distribution['A1'].font = Font(bold=True)
        ws_distribution['B1'].font = Font(bold=True)
        ws_distribution['C1'].font = Font(bold=True)
        
        total_categorized = summary['categorized_rows']
        for row, (category, count) in enumerate(report_data['top_categories'], 2):
            ws_distribution[f'A{row}'] = category
            ws_distribution[f'B{row}'] = count
            ws_distribution[f'C{row}'] = f"{count/total_categorized:.1%}" if total_categorized > 0 else "0%"
        
        # Uncategorized samples sheet
        if include_samples and report_data['uncategorized_analysis']['sample_uncategorized']:
            ws_uncategorized = wb.create_sheet("Uncategorized Samples", 2)
            ws_uncategorized['A1'] = "Uncategorized Descriptions"
            ws_uncategorized['A1'].font = Font(bold=True)
            
            for row, desc in enumerate(report_data['uncategorized_analysis']['sample_uncategorized'], 2):
                ws_uncategorized[f'A{row}'] = desc
        
        # Quality metrics sheet
        ws_quality = wb.create_sheet("Quality Metrics", 3)
        ws_quality['A1'] = "Quality Metrics"
        ws_quality['A1'].font = Font(size=14, bold=True)
        
        quality_data = [
            ("Categories with Single Item", report_data['quality_metrics']['categories_with_single_item']),
            ("Average Items per Category", f"{report_data['quality_metrics']['average_items_per_category']:.1f}"),
            ("Most Common Category Count", report_data['quality_metrics']['most_common_category_count'])
        ]
        
        for row, (label, value) in enumerate(quality_data, 3):
            ws_quality[f'A{row}'] = label
            ws_quality[f'B{row}'] = value
            ws_quality[f'A{row}'].font = Font(bold=True)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Categorization report exported successfully to {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error exporting categorization report: {e}")
        return False


def update_master_dictionary(
    category_dict: CategoryDictionary,
    manual_categorizations: dict,
    backup_dir: Path = Path('config/backups')
) -> dict:
    """
    Update the master CategoryDictionary with new manual categorizations.
    - Adds new description-category mappings
    - Prevents duplicates and handles conflicts
    - Creates a backup of the old dictionary before updating
    - Saves the updated dictionary to file
    - Logs all additions and conflicts

    Args:
        category_dict: CategoryDictionary object
        manual_categorizations: dict mapping description (str) to category (str)
        backup_dir: Directory to store backups (default: config/backups)
    Returns:
        dict with summary of additions, conflicts, and backup path
    """
    logger.info(f"Updating master dictionary with {len(manual_categorizations)} manual categorizations...")
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = None
    additions = []
    conflicts = []
    skipped = []

    # Backup the old dictionary file
    dict_file = category_dict.dictionary_file
    if dict_file.exists():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = backup_dir / f"category_dictionary_{timestamp}.json"
        shutil.copy2(dict_file, backup_path)
        logger.info(f"Backup of old dictionary created at: {backup_path}")
    else:
        logger.warning(f"Dictionary file {dict_file} does not exist, skipping backup.")

    # Add new mappings
    for desc, cat in manual_categorizations.items():
        desc_norm = desc.strip().lower()
        cat_norm = cat.strip()
        if not desc_norm or not cat_norm:
            continue
        # Check for existing mapping
        existing = category_dict.mappings.get(desc_norm)
        if existing:
            if existing.category == cat_norm:
                skipped.append(desc)
                continue  # Already present, skip
            else:
                # Conflict: description exists with different category
                conflicts.append({'description': desc, 'existing_category': existing.category, 'new_category': cat_norm})
                logger.warning(f"Conflict for '{desc}': existing category '{existing.category}', new '{cat_norm}' (skipped)")
                continue  # Do not overwrite, just log
        # Add new mapping
        added = category_dict.add_mapping(desc_norm, cat_norm)
        if added:
            additions.append({'description': desc, 'category': cat_norm})
            logger.info(f"Added mapping: '{desc}' → '{cat_norm}'")
        else:
            logger.error(f"Failed to add mapping: '{desc}' → '{cat_norm}'")

    # Save updated dictionary
    saved = category_dict.save_dictionary()
    if saved:
        logger.info(f"Updated dictionary saved to {dict_file}")
    else:
        logger.error(f"Failed to save updated dictionary to {dict_file}")

    summary = {
        'additions': additions,
        'conflicts': conflicts,
        'skipped': skipped,
        'backup_path': str(backup_path) if backup_path else None,
        'saved': saved,
        'total_added': len(additions),
        'total_conflicts': len(conflicts),
        'total_skipped': len(skipped)
    }
    return summary 


def execute_row_categorization(
    mapped_df,
    category_dict_path: Path = Path('config/category_dictionary.json'),
    output_dir: Path = Path('examples'),
    user_manual_excel: Path = None,
    cleanup_temp: bool = True,
    progress_callback=None
) -> dict:
    """
    Orchestrate the full row categorization process.
    Args:
        mapped_df: The mapped DataFrame to categorize
        category_dict_path: Path to the category dictionary JSON
        output_dir: Directory for outputs and temp files
        user_manual_excel: Optional path to user-completed manual categorization Excel
        cleanup_temp: Whether to clean up temp files
        progress_callback: Optional function(percent, message) for progress updates
    Returns:
        dict with 'final_dataframe', 'summary', 'all_stats', 'rollback_performed', 'error' (if any)
    """
    import shutil
    import traceback
    from datetime import datetime
    from core.category_dictionary import CategoryDictionary
    
    temp_files = []
    rollback_performed = False
    error = None
    summary = {}
    all_stats = {}
    final_df = None
    backup_dict_path = None
    
    def update_progress(percent, message):
        if progress_callback:
            progress_callback(percent, message)
        else:
            print(f"[{percent:.0f}%] {message}")
    
    try:
        update_progress(0, "Loading category dictionary...")
        category_dict = CategoryDictionary(category_dict_path)
        summary['category_dict_loaded'] = True
        all_stats['initial_dict_size'] = len(category_dict.mappings)
        
        update_progress(5, "Auto-categorizing rows...")
        from core.auto_categorizer import auto_categorize_dataset, collect_unmatched_descriptions
        auto_result = auto_categorize_dataset(mapped_df, category_dict)
        print(f"[DEBUG] auto_categorize_dataset: total_rows={auto_result.total_rows}, matched_rows={auto_result.matched_rows}, unmatched_rows={auto_result.unmatched_rows}")
        auto_df = auto_result.dataframe
        all_stats['auto_stats'] = auto_result.match_statistics
        summary['auto_categorized'] = True
        
        update_progress(20, f"Collecting descriptions for manual review...")
        # Pass the sheet name column if it exists
        sheet_name_column = None
        for col in auto_df.columns:
            if col.lower() in ["sheet_name", "source_sheet", "sheet"]:
                sheet_name_column = col
                break
        from core.auto_categorizer import collect_descriptions_for_manual_review
        review_list = collect_descriptions_for_manual_review(
            auto_df, 
            category_dict, 
            category_column='Category', 
            description_column='Description', 
            sheet_name_column=sheet_name_column,
            confidence_threshold=0.8
        )
        print(f"[DEBUG] review_list: {len(review_list)} descriptions for manual review")
        all_stats['review_count'] = len(review_list)
        summary['review_collected'] = True
        
        # Check if manual categorization is needed
        if len(review_list) == 0:
            # All rows were auto-categorized successfully
            update_progress(100, "All rows auto-categorized successfully - no manual categorization needed.")
            summary['all_auto_categorized'] = True
            summary['manual_categorization_needed'] = False
            
            return {
                'final_dataframe': auto_df,  # Return auto-categorized DataFrame as final result
                'summary': summary,
                'all_stats': all_stats,
                'rollback_performed': rollback_performed,
                'error': None,
                'manual_excel_path': None,
                'review_list': [],
                'category_dict': category_dict
            }
        
        # Generate manual categorization Excel
        update_progress(30, "Generating manual categorization Excel file...")
        from core.manual_categorizer import generate_manual_categorization_excel
        available_categories = list(category_dict.get_all_categories())
        print(f"[DEBUG] About to call generate_manual_categorization_excel with {len(review_list)} review descriptions, output_dir={output_dir}")
        excel_path = generate_manual_categorization_excel(review_list, available_categories, output_dir=output_dir)
        temp_files.append(excel_path)
        summary['manual_excel_generated'] = str(excel_path)
        summary['manual_categorization_needed'] = True
        update_progress(40, f"Manual categorization Excel generated at {excel_path}")
        
        # Stop here - let the UI handle manual categorization
        # The user will need to complete the Excel file and upload it back
        update_progress(100, "Manual categorization Excel file ready for user input.")
        
        return {
            'final_dataframe': auto_df,  # Return auto-categorized DataFrame
            'summary': summary,
            'all_stats': all_stats,
            'rollback_performed': rollback_performed,
            'error': None,
            'manual_excel_path': str(excel_path),
            'review_list': review_list,
            'category_dict': category_dict
        }
    except Exception as exc:
        error = str(exc)
        traceback_str = traceback.format_exc()
        print(f"[ERROR] {error}\n{traceback_str}")
        summary['error'] = error
        # Rollback: restore dictionary from backup if it was made
        if backup_dict_path and Path(backup_dict_path).exists():
            try:
                shutil.copy2(backup_dict_path, category_dict_path)
                rollback_performed = True
                print(f"[ROLLBACK] Restored category dictionary from backup: {backup_dict_path}")
            except Exception as e:
                print(f"[ROLLBACK ERROR] Failed to restore backup: {e}")
        # Clean up temp files
        if cleanup_temp:
            for f in temp_files:
                try:
                    Path(f).unlink(missing_ok=True)
                except Exception as e:
                    print(f"Warning: Could not delete temp file {f}: {e}")
        return {
            'final_dataframe': None,
            'summary': summary,
            'all_stats': all_stats,
            'rollback_performed': rollback_performed,
            'error': error,
            'traceback': traceback_str
        } 

def get_manual_categorization_categories() -> List[str]:
    """Get the list of categories for manual categorization in the desired order and format"""
    return [
        "General Costs",
        "Site Costs",
        "Civil Works",
        "Earth Movement",
        "Roads",
        "OEM Building",
        "Electrical Works", 
        "Solar Cables",
        "LV Cables",
        "MV Cables",
        "Trenching",
        "PV Mod. Installation",
        "Cleaning and Cabling of PV Mod.",
        "Tracker Inst.",
        "Other",
    ] 