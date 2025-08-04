import openpyxl
import os
import shutil
import pickle

# --- Configuration ---
ORIGINAL_FILE_PATH = "examples/GRE.EEC.F.27.IT.P.18371.00.098.02 - PONTESTURA 9,69 MW_cBOQ PV_rev 9 giu (ENEMEK).xlsx"
OUTPUT_DIR = "test_outputs"
MAPPING_FILE = "examples/pontestura mapping.pkl"
TARGET_SHEET_NAME = "ACCESS ROADS"

# Load mapping data
try:
    with open(MAPPING_FILE, 'rb') as f:
        mapping_data = pickle.load(f)
    
    sheet_mapping = mapping_data.get(TARGET_SHEET_NAME, {})
    
    START_ROW = sheet_mapping.get('start_row', 2)
    DESCRIPTION_COLUMN = sheet_mapping.get('description_col', 'B')
    QUANTITY_COLUMN = sheet_mapping.get('quantity_col', 'C')
    UNIT_COLUMN = sheet_mapping.get('unit_col', 'D')

    print(f"Loaded mapping for '{TARGET_SHEET_NAME}':")
    print(f"  START_ROW: {START_ROW}")
    print(f"  DESCRIPTION_COLUMN: {DESCRIPTION_COLUMN}")
    print(f"  QUANTITY_COLUMN: {QUANTITY_COLUMN}")
    print(f"  UNIT_COLUMN: {UNIT_COLUMN}")

except FileNotFoundError:
    print(f"Error: Mapping file not found at {MAPPING_FILE}. Using default column/row settings.")
    DESCRIPTION_COLUMN = 'B'
    QUANTITY_COLUMN = 'C'
    UNIT_COLUMN = 'D'
    START_ROW = 2
except Exception as e:
    print(f"Error loading mapping file: {e}. Using default column/row settings.")
    DESCRIPTION_COLUMN = 'B'
    QUANTITY_COLUMN = 'C'
    UNIT_COLUMN = 'D'
    START_ROW = 2

def safe_set_cell_value(ws, row, col, value):
    """
    Safely sets the value of a cell, handling merged cells.
    If the cell is part of a merged range, it temporarily unmerges, sets value, and re-merges.
    """
    cell_coords = openpyxl.utils.get_column_letter(col) + str(row)
    
    merged_cell_ranges_to_unmerge = []
    for merged_range in ws.merged_cells.ranges:
        if cell_coords in merged_range:
            merged_cell_ranges_to_unmerge.append(merged_range)
    
    for merged_range in merged_cell_ranges_to_unmerge:
        ws.unmerge_cells(str(merged_range))

    ws.cell(row=row, column=col).value = value

    for merged_range in merged_cell_ranges_to_unmerge:
        ws.merge_cells(str(merged_range))


def copy_and_load_workbook(test_number, test_name, sheet_name=TARGET_SHEET_NAME):
    """Copies the original workbook to a new file and loads it."""
    output_filename = os.path.join(OUTPUT_DIR, f"{test_number}_{test_name}.xlsx")
    shutil.copy(ORIGINAL_FILE_PATH, output_filename)
    wb = openpyxl.load_workbook(output_filename)
    return wb, wb[sheet_name], output_filename

# --- Test Cases ---

def test_minor_text_changes():
    """
    TEST 1: Minor Text Changes: A single character change, a typo correction,
    or a change in capitalization (e.g., "Concrete" vs. "concrete").
    """
    print("Running Test 1: Minor Text Changes")
    wb, ws, output_path = copy_and_load_workbook(1, "minor_text_changes")

    # Find the first non-empty description cell and modify it
    for row_idx in range(START_ROW, ws.max_row + 1):
        cell = ws[f"{DESCRIPTION_COLUMN}{row_idx}"]
        if cell.value and isinstance(cell.value, str):
            original_text = cell.value
            # Change capitalization of the first word
            modified_text = original_text[0].lower() + original_text[1:] if original_text else ""
            cell.value = modified_text
            print(f"  Modified cell {DESCRIPTION_COLUMN}{row_idx}: '{original_text}' -> '{modified_text}'")
            break # Only modify the first one

    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_reordered_words():
    """
    TEST 2: Reordered Words: "Steel Reinforcement Bar" vs. "Reinforcement Steel Bar".
    """
    print("Running Test 2: Reordered Words")
    wb, ws, output_path = copy_and_load_workbook(2, "reordered_words")

    # Find a suitable description and reorder words
    for row_idx in range(START_ROW, ws.max_row + 1):
        cell = ws[f"{DESCRIPTION_COLUMN}{row_idx}"]
        if cell.value and isinstance(cell.value, str) and len(cell.value.split()) >= 2:
            words = cell.value.split()
            if len(words) >= 2:
                # Swap the first two words
                modified_words = [words[1], words[0]] + words[2:]
                modified_text = " ".join(modified_words)
                original_text = cell.value
                cell.value = modified_text
                print(f"  Modified cell {DESCRIPTION_COLUMN}{row_idx}: '{original_text}' -> '{modified_text}'")
                break
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_added_removed_keywords():
    """
    TEST 3: Added/Removed Keywords: "Concrete C25/30" vs. "Concrete C25/30 (Pumpable)".
    """
    print("Running Test 3: Added/Removed Keywords")
    wb, ws, output_path = copy_and_load_workbook(3, "added_removed_keywords")

    for row_idx in range(START_ROW, ws.max_row + 1):
        cell = ws[f"{DESCRIPTION_COLUMN}{row_idx}"]
        if cell.value and isinstance(cell.value, str):
            original_text = cell.value
            # Add a keyword
            modified_text = f"{original_text} (Modified Keyword)"
            cell.value = modified_text
            print(f"  Modified cell {DESCRIPTION_COLUMN}{row_idx}: '{original_text}' -> '{modified_text}'")
            break
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_different_units_of_measure():
    """
    TEST 4: Different Units of Measure: "100 m" vs. "100 linear meters" for the same item.
    """
    print("Running Test 4: Different Units of Measure")
    wb, ws, output_path = copy_and_load_workbook(4, "different_units_of_measure")

    for row_idx in range(START_ROW, ws.max_row + 1):
        unit_cell = ws[f"{UNIT_COLUMN}{row_idx}"]
        if unit_cell.value and isinstance(unit_cell.value, str):
            original_unit = unit_cell.value
            # Simple replacement for common units
            if original_unit.lower() == "m":
                modified_unit = "linear meters"
            elif original_unit.lower() == "sqm":
                modified_unit = "square meters"
            elif original_unit.lower() == "cum":
                modified_unit = "cubic meters"
            else:
                modified_unit = f"{original_unit} (alt)" # Generic alteration
            
            unit_cell.value = modified_unit
            print(f"  Modified unit cell {UNIT_COLUMN}{row_idx}: '{original_unit}' -> '{modified_unit}'")
            break
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_special_characters_encoding():
    """
    TEST 5: Special Characters/Encoding: Descriptions containing unusual characters, symbols, or different language encodings.
    """
    print("Running Test 5: Special Characters/Encoding")
    wb, ws, output_path = copy_and_load_workbook(5, "special_characters_encoding")

    for row_idx in range(START_ROW, ws.max_row + 1):
        cell = ws[f"{DESCRIPTION_COLUMN}{row_idx}"]
        if cell.value and isinstance(cell.value, str):
            original_text = cell.value
            # Add some special characters
            modified_text = f"{original_text} - €§©®™"
            cell.value = modified_text
            print(f"  Modified cell {DESCRIPTION_COLUMN}{row_idx}: '{original_text}' -> '{modified_text}'")
            break
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_reordered_items():
    """
    TEST 6: Reordered Items: The same items exist, but their order in the BoQ has changed.
    This version attempts to swap the content of the first two *unmerged* description cells.
    """
    print("Running Test 6: Reordered Items (simplified)")
    wb, ws, output_path = copy_and_load_workbook(6, "reordered_items_simplified")

    # Find the first two unmerged description cells
    cell1_coords = None
    cell2_coords = None
    
    for row_idx in range(START_ROW, ws.max_row + 1):
        cell_coords = f"{DESCRIPTION_COLUMN}{row_idx}"
        # Check if cell is not part of a merged range
        if cell_coords not in ws.merged_cells and ws[cell_coords].value is not None:
            if cell1_coords is None:
                cell1_coords = cell_coords
            elif cell2_coords is None:
                cell2_coords = cell_coords
                break # Found two unmerged cells

    if cell1_coords and cell2_coords:
        # Swap the values
        value1 = ws[cell1_coords].value
        value2 = ws[cell2_coords].value
        ws[cell1_coords].value = value2
        ws[cell2_coords].value = value1
        print(f"  Swapped content of {cell1_coords} and {cell2_coords}.")
    else:
        print("  Could not find two unmerged description cells to reorder.")

    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_split_rows():
    """
    TEST 7: Split Rows: One item is split into two or more distinct items.
    This test will take one item, reduce its quantity, and add a new row
    with the remaining quantity and a slightly modified description.
    """
    print("Running Test 7: Split Rows")
    wb, ws, output_path = copy_and_load_workbook(7, "split_rows")

    target_row_idx = -1
    original_quantity = 0
    original_description = ""
    original_unit = ""

    # Find a row with a quantity > 1 to split
    for row_idx in range(START_ROW, ws.max_row + 1):
        quantity_cell = ws[f"{QUANTITY_COLUMN}{row_idx}"]
        if quantity_cell.value and isinstance(quantity_cell.value, (int, float)) and quantity_cell.value > 1:
            target_row_idx = row_idx
            original_quantity = quantity_cell.value
            original_description = ws[f"{DESCRIPTION_COLUMN}{row_idx}"].value
            original_unit = ws[f"{UNIT_COLUMN}{row_idx}"].value
            break
    
    if target_row_idx != -1:
        # Reduce original quantity
        new_quantity_part1 = original_quantity // 2
        new_quantity_part2 = original_quantity - new_quantity_part1
        ws[f"{QUANTITY_COLUMN}{target_row_idx}"].value = new_quantity_part1

        # Insert a new row below the modified one
        ws.insert_rows(target_row_idx + 1)
        
        # Populate the new row
        new_row_description = f"{original_description} (Part 2)"
        ws[f"{DESCRIPTION_COLUMN}{target_row_idx + 1}"].value = new_row_description
        ws[f"{QUANTITY_COLUMN}{target_row_idx + 1}"].value = new_quantity_part2
        ws[f"{UNIT_COLUMN}{target_row_idx + 1}"].value = original_unit
        print(f"  Split row {target_row_idx}: Original quantity {original_quantity} split into {new_quantity_part1} and {new_quantity_part2}.")
    else:
        print("  No suitable row found to split (quantity <= 1).")

    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_missing_required_fields():
    """
    TEST 8: Missing Required Fields: What if an item is missing its quantity, unit, or description?
    This test will clear the quantity, unit, and description for a single item.
    """
    print("Running Test 8: Missing Required Fields")
    wb, ws, output_path = copy_and_load_workbook(8, "missing_required_fields")

    # Find the first data row and clear its quantity, unit, and description
    for row_idx in range(START_ROW, ws.max_row + 1):
        desc_col_idx = openpyxl.utils.column_index_from_string(DESCRIPTION_COLUMN)
        qty_col_idx = openpyxl.utils.column_index_from_string(QUANTITY_COLUMN)
        unit_col_idx = openpyxl.utils.column_index_from_string(UNIT_COLUMN)

        if ws.cell(row=row_idx, column=desc_col_idx).value: # Ensure it's a data row
            safe_set_cell_value(ws, row_idx, desc_col_idx, None)
            safe_set_cell_value(ws, row_idx, qty_col_idx, None)
            safe_set_cell_value(ws, row_idx, unit_col_idx, None)
            print(f"  Cleared description, quantity, and unit for row {row_idx}.")
            break
    
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_invalid_data_types():
    """
    TEST 9: Invalid Data Types: Non-numeric values in quantity fields, or dates in text fields.
    This test will put a string in the quantity field and a number in the description field.
    """
    print("Running Test 9: Invalid Data Types")
    wb, ws, output_path = copy_and_load_workbook(9, "invalid_data_types")

    for row_idx in range(START_ROW, ws.max_row + 1):
        desc_col_idx = openpyxl.utils.column_index_from_string(DESCRIPTION_COLUMN)
        qty_col_idx = openpyxl.utils.column_index_from_string(QUANTITY_COLUMN)

        if ws.cell(row=row_idx, column=desc_col_idx).value: # Ensure it's a data row
            # Put a string in quantity
            original_quantity = ws.cell(row=row_idx, column=qty_col_idx).value
            safe_set_cell_value(ws, row_idx, qty_col_idx, "INVALID_QUANTITY")
            print(f"  Modified quantity in row {row_idx}: '{original_quantity}' -> 'INVALID_QUANTITY'")

            # Put a number in description
            original_description = ws.cell(row=row_idx, column=desc_col_idx).value
            safe_set_cell_value(ws, row_idx, desc_col_idx, 12345)
            print(f"  Modified description in row {row_idx}: '{original_description}' -> '12345'")
            break
    
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

def test_only_metadata_changes():
    """
    TEST 10: Only Metadata Changes: Changes only in header information or non-item-specific data.
    This test will modify a cell in the first row (header/metadata) that is not part of the BoQ item data.
    Assuming cell A1 is a good candidate for metadata.
    """
    print("Running Test 10: Only Metadata Changes")
    wb, ws, output_path = copy_and_load_workbook(10, "only_metadata_changes")

    # Modify a cell in the first row (e.g., A1)
    original_a1 = ws['A1'].value
    ws['A1'].value = f"Modified Metadata - {original_a1}"
    print(f"  Modified cell A1: '{original_a1}' -> '{ws['A1'].value}'")
    
    wb.save(output_path)
    print(f"  Saved to: {output_path}")

# --- Main Execution ---
if __name__ == "__main__":
    if not os.path.exists(ORIGINAL_FILE_PATH):
        print(f"Error: Original file not found at {ORIGINAL_FILE_PATH}")
        print("Please ensure the path is correct and the file exists.")
    elif not os.path.exists(MAPPING_FILE):
        print(f"Error: Mapping file not found at {MAPPING_FILE}.")
        print("Please ensure the path is correct.")
    else:
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
            print(f"Created output directory: {OUTPUT_DIR}")

        print("Starting test file generation...")
        test_minor_text_changes()
        test_reordered_words()
        test_added_removed_keywords()
        test_different_units_of_measure()
        test_special_characters_encoding()
        test_reordered_items()
        test_split_rows()
        test_missing_required_fields()
        test_invalid_data_types()
        test_only_metadata_changes()
        print("All test files generated.")
        print(f"Please check the '{OUTPUT_DIR}' directory for the generated files.")
        print("\nTo run this script, you might need to install openpyxl:")
        print("pip install openpyxl")